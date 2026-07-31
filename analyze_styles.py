import openpyxl

EXCEL_PATH = r"c:\Users\夏瑞泽\xwechat_files\wxid_vnl456wm0r9h12_0c22\msg\file\2026-07\2025-2026学年夏五周到夏八周礼服分配(1)(1).xlsx"

wb = openpyxl.load_workbook(EXCEL_PATH)

print("=== DELTA CELL FILL ANALYSIS (装备分配) ===")
ws = wb['装备分配']
# Check every cell for non-default fill
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        fg = cell.fill.fgColor
        bg = cell.fill.bgColor
        pt = cell.fill.patternType
        if pt and pt != 'none':
            fg_info = f"fg=rgb:{fg.rgb}/idx:{fg.indexed}/theme:{fg.theme}/tint:{fg.tint}" if fg else "fg=None"
            bg_info = f"bg=rgb:{bg.rgb}/idx:{bg.indexed}/theme:{bg.theme}/tint:{bg.tint}" if bg else "bg=None"
            print(f"  [{row},{col}] {cell.value} | pattern={pt} | {fg_info} | {bg_info}")

print("\n=== RED FONT ANALYSIS (礼服腰带摆放) ===")
ws2 = wb['礼服腰带摆放']
for row in range(1, ws2.max_row + 1):
    for col in range(1, ws2.max_column + 1):
        cell = ws2.cell(row=row, column=col)
        if cell.value is not None:
            fc = cell.font.color
            is_red = False
            try:
                if fc:
                    is_red = (fc.rgb and 'FF0000' in str(fc.rgb)) or (fc.indexed == 2)
            except:
                pass
            if is_red or (fc and fc.rgb and fc.rgb != '00000000'):
                fc_info = f"rgb={fc.rgb}/idx={fc.indexed}/theme={fc.theme}/tint={fc.tint}/type={fc.type}"
                print(f"  [{row},{col}] {cell.value} | font.color: {fc_info}")

print("\n=== RED FONT ANALYSIS (礼帽摆放) ===")
ws3 = wb['礼帽摆放']
for row in range(1, ws3.max_row + 1):
    for col in range(1, ws3.max_column + 1):
        cell = ws3.cell(row=row, column=col)
        if cell.value is not None:
            fc = cell.font.color
            is_red = False
            try:
                if fc:
                    is_red = (fc.rgb and 'FF0000' in str(fc.rgb)) or (fc.indexed == 2)
            except:
                pass
            if is_red or (fc and fc.rgb and fc.rgb != '00000000'):
                fc_info = f"rgb={fc.rgb}/idx={fc.indexed}/theme={fc.theme}/tint={fc.tint}/type={fc.type}"
                print(f"  [{row},{col}] {cell.value} | font.color: {fc_info}")

print("\n=== ROW HEIGHTS & COLUMN WIDTHS (ALL SHEETS) ===")
for sname in wb.sheetnames:
    ws = wb[sname]
    rh = {}
    for r in range(1, min(ws.max_row+1, 50)):
        if r in ws.row_dimensions and ws.row_dimensions[r].height:
            rh[r] = ws.row_dimensions[r].height
    cw = {}
    for col_letter, dim in ws.column_dimensions.items():
        if dim.width:
            cw[col_letter] = dim.width
    print(f"  {sname}: row_heights={rh}")
    print(f"  {sname}: col_widths={cw}")
