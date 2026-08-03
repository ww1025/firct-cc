"""离线验证 X 聚类 + 列放大的完整流程"""
import cv2, numpy as np, easyocr, re, sys
from PIL import Image
from difflib import SequenceMatcher

sys.path.insert(0, r'C:\Users\夏瑞泽\Desktop\firct cc\uniform-assigner')
from server import force_correct_name, fuzzy_match, fuzzy_match_aggressive

image_path = r'c:\Users\夏瑞泽\xwechat_files\wxid_vnl456wm0r9h12_0c22\temp\RWTemp\2026-07\9e20f478899dc29eb19741386f9343c8\cfd5ca841af9d1bf15e098f16c5f3147.png'
img = Image.open(image_path).convert('RGB')
w, h = img.size

TARGET_W = min(w, 2000)
scale = TARGET_W / w
new_w, new_h = TARGET_W, int(h * scale)
img = img.resize((new_w, new_h), Image.LANCZOS)
print(f'Image: {new_w}x{new_h}')

reader = easyocr.Reader(['ch_sim'], gpu=False)
img_np = np.array(img)
gray = np.mean(img_np, axis=2).astype(np.uint8)

results = reader.readtext(img_np, detail=1, low_text=0.2, text_threshold=0.3)
print(f'OCR text blocks: {len(results)}')

# Filter edges
filtered = []
for bbox, text, conf in results:
    x1, y1 = bbox[0]; x3, y3 = bbox[2]
    t = text.strip()
    if not t: continue
    if x1 <= 5 and y1 <= 5: continue
    if x3 >= new_w - 5 and y3 >= new_h - 5: continue
    filtered.append((bbox, t, conf))

# Blocks
blocks = []
for bbox, text, conf in filtered:
    x1, y1 = bbox[0]; x3, y3 = bbox[2]
    blocks.append({'text': text, 'x1': x1, 'x3': x3, 'cx': (x1+x3)/2, 'y1': y1, 'y3': y3})

# X clustering
blocks.sort(key=lambda b: b['cx'])
x_clusters = []
if blocks:
    x_clusters = [[blocks[0]]]
    for b in blocks[1:]:
        if b['cx'] - x_clusters[-1][-1]['cx'] < 40:
            x_clusters[-1].append(b)
        else:
            x_clusters.append([b])
print(f'X-clusters: {len(x_clusters)}')

# Load known names
import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\夏瑞泽\Desktop\firct cc\uniform-assigner\本月礼服分配.xlsx')
known_names = []
for row in range(2, wb['装备分配'].max_row + 1):
    n = wb['装备分配'].cell(row=row, column=2).value
    if n and str(n).strip():
        known_names.append(str(n).strip())
wb.close()

DAY_LABELS = ['周一', '周二', '周三', '周四', '周五']

# Now process each column - crop and re-OCR at higher resolution
for ci, cluster in enumerate(x_clusters):
    cluster.sort(key=lambda b: b['y1'])
    col_x1 = max(0, int(min(b['x1'] for b in cluster)) - 5)
    col_x2 = min(new_w, int(max(b['x3'] for b in cluster)) + 5)
    col_y1 = max(0, int(min(b['y1'] for b in cluster)) - 5)
    col_y2 = min(new_h, int(max(b['y3'] for b in cluster)) + 5)

    col_img = gray[col_y1:col_y2, col_x1:col_x2]
    if col_img.size == 0: continue

    # Scale up
    scale_col = max(1.0, 600.0 / col_img.shape[0])  # Bigger - 600px
    col_big = cv2.resize(col_img,
        (int(col_img.shape[1] * scale_col), int(col_img.shape[0] * scale_col)),
        interpolation=cv2.INTER_LANCZOS4)

    # Re-OCR the column at large size
    col_results = reader.readtext(col_big, detail=1, low_text=0.1, text_threshold=0.15)

    day = DAY_LABELS[ci] if ci < 5 else f'C{ci}'

    # Try: check top texts for day label
    top_texts = []
    for _, t, conf in col_results:
        t = t.strip()
        if len(t) >= 2:
            top_texts.append(t)
        if len(top_texts) >= 3:
            break
    print(f'\n=== {day} (x {col_x1}-{col_x2}) | top: {top_texts[:3]}')

    # Apply name correction
    seen = set()
    for _, t, conf in col_results:
        t = t.strip()
        if len(t) < 2: continue
        corrected = force_correct_name(t, known_names)
        if not corrected:
            corrected = fuzzy_match_aggressive(t, known_names)
        if corrected and corrected not in seen:
            seen.add(corrected)
            print(f'  {t} ({conf:.2f}) -> {corrected}')
