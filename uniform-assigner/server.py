"""礼服分配系统 - Flask版，双击启动.bat 即可使用"""
import base64, io, os, re, tempfile
from flask import Flask, request, jsonify
from difflib import SequenceMatcher
import cv2, numpy as np

app = Flask(__name__)

YELLOW = 'FFFFFF00'
PORT = 8765

# ── 人名模糊匹配 + 自动修正 ──
def fuzzy_match(ocr_name, known_names):
    """将 OCR 误识别的名字匹配到最接近的已知人名"""
    if not known_names:
        return ocr_name
    if ocr_name in known_names:
        return ocr_name
    # 去掉 OCR 噪声字符
    cleaned = re.sub(r'[_|^~\s\d]+', '', ocr_name)
    if cleaned in known_names:
        return cleaned
    # 前两字相同优先 —— 取最短的作为更精确匹配
    candidates_2char = []
    for ref in known_names:
        if len(ref) >= 2 and len(cleaned) >= 2 and cleaned[:2] == ref[:2]:
            candidates_2char.append(ref)
    if candidates_2char:
        return min(candidates_2char, key=len)
    best, best_score = None, 0
    for ref in known_names:
        common = len(set(cleaned) & set(ref))
        ratio = SequenceMatcher(None, cleaned, ref).ratio()
        score = common * 3 + ratio * 5
        if score > best_score:
            best_score, best = score, ref
    if best and best_score >= 6:
        return best
    return re.sub(r'[_|^~\s\d]+', '', ocr_name)  # 至少去噪

def fuzzy_match_aggressive(ocr_name, known_names):
    """超级宽松匹配：只用字符重叠，专门对付 EasyOCR 离谱的手写字 """
    if not ocr_name or not known_names:
        return None
    if ocr_name in known_names:
        return ocr_name
    cleaned = re.sub(r'[_|^~\s\d\s]+', '', ocr_name)
    if not cleaned:
        return None
    if cleaned in known_names:
        return cleaned
    # 前两字相同
    candidates_2char = []
    for ref in known_names:
        if len(ref) >= 2 and len(cleaned) >= 2 and cleaned[:2] == ref[:2]:
            candidates_2char.append(ref)
    if candidates_2char:
        return min(candidates_2char, key=len)
    # 至少有一个字重叠 + 长度相近
    best, best_score = None, 0
    cc = set(cleaned)
    for ref in known_names:
        rc = set(ref)
        common = len(cc & rc)
        if common == 0:
            continue
        overlap = common / max(len(cc), len(rc))
        # 长度差惩罚
        len_penalty = abs(len(ref) - len(cleaned)) * 1.5
        score = overlap * 10 - len_penalty
        if score > best_score:
            best_score = score
            best = ref
    if best and best_score >= 2.0:
        return best
    return None

def force_correct_name(ocr_text, known_names):
    """强制纠错：OCR结果必须匹配已知人名库，无法可信匹配返回 None。
    用于日常排班 OCR —— 绝不输出错字，宁缺毋滥。"""
    if not ocr_text or len(ocr_text) < 2:
        return None
    if ocr_text in known_names:
        return ocr_text
    cleaned = re.sub(r'[_|^~\s\d]+', '', ocr_text)
    if cleaned in known_names:
        return cleaned
    # 前两字相同优先（中文名常见 OCR 错误模式）
    # 改为取最高评分而不是第一个匹配（修复同姓名字顺序依赖）
    candidates_2char = []
    for ref in known_names:
        if len(ref) >= 2 and len(cleaned) >= 2 and cleaned[:2] == ref[:2]:
            candidates_2char.append(ref)
    if candidates_2char:
        # 取最短的（更精确的匹配），例如 "张鹏" vs "张鹏飞"
        return min(candidates_2char, key=len)
    # 字符重叠 + 序列匹配
    best, best_score = None, 0
    nc = set(cleaned)
    for ref in known_names:
        rc = set(ref)
        common = len(nc & rc)
        overlap = common / max(len(nc), len(rc), 1)
        seq = SequenceMatcher(None, cleaned, ref).ratio()
        score = overlap * 10 + seq * 5
        if score > best_score:
            best_score = score
            best = ref
    if best and best_score >= 5:
        return best
    return None  # 无法可信匹配，返回 None 由调用方决定

# ── 业务逻辑 ──
def parse_equipment_sheet(ws):
    persons = []
    for row in range(2, ws.max_row+1):
        name = ws.cell(row=row, column=2).value
        if not name: continue
        name = str(name).strip()
        uni  = str(ws.cell(row=row, column=3).value or '').strip()
        hat  = str(ws.cell(row=row, column=4).value or '').strip()
        boots= str(ws.cell(row=row, column=5).value or '').strip()
        belt = str(ws.cell(row=row, column=6).value or '').strip()
        note = str(ws.cell(row=row, column=7).value or '').strip()
        persons.append(dict(name=name, gender='M' if uni.startswith('M') else 'F',
                            uniform=uni, hat=hat, boots=boots, belt=belt, note=note))
    return persons

def parse_shared_relations(persons):
    rels, seen = [], set()
    for p in persons:
        if not p['note']: continue
        for seg in re.split(r'[，；]', p['note']):
            seg = seg.strip()
            if not seg.startswith('和'): continue
            idx = seg.find('共用')
            if idx < 0: continue
            names = [n.strip() for n in re.split(r'[、，, ]+', seg[1:idx]) if n.strip()]
            after = seg[idx+2:]
            types = [t for t in [('uniform','礼服'),('hat','礼帽'),('boots','马靴'),('belt','腰带')] if t[1] in after]
            for n in names:
                for t, _ in types:
                    code = p[t]
                    key = f"{t}|{code}|{','.join(sorted([p['name'], n]))}"
                    if key not in seen:
                        seen.add(key)
                        rels.append(dict(item_code=code, item_type=t, shared_by=[p['name'], n]))
    return rels

def merge_relations(rels):
    m = {}
    for r in rels:
        k = f"{r['item_type']}|{r['item_code']}"
        m.setdefault(k, set()).update(r['shared_by'])
    return [dict(item_code=k.split('|')[1], item_type=k.split('|')[0], shared_by=list(v)) for k,v in m.items()]

def detect_conflicts(schedule, rels):
    conflicts = []
    for day in schedule:
        daily = set(day['people'])
        for r in rels:
            overlap = [p for p in r['shared_by'] if p in daily]
            if len(overlap) >= 2:
                for i in range(1, len(overlap)):
                    conflicts.append(dict(day=day['day'], item_type=r['item_type'], item_code=r['item_code'],
                                          person_to_move=overlap[i], person_to_keep=overlap[0]))
    return conflicts

def parse_uniform_code(code):
    m = re.match(r'^([FM])(\d+)/(\d+)-(\d+)$', code)
    return dict(gender=m.group(1), height=int(m.group(2)), chest=int(m.group(3))) if m else None

def parse_hat_code(code):
    m = re.match(r'^([FM])(\d{4})$', code)
    return dict(gender=m.group(1), number=int(m.group(2))) if m else None

def find_alternative(person, item_type, used_items, all_items):
    """为冲突人员寻找最接近原尺寸的替换装备。

    礼服评分规则（尺寸档位对齐）：
      - 身高每差 5cm = 一个档位，权重 4 → 跨一档 ≈ 20分
      - 胸围每差 4cm = 一个档位，权重 5 → 跨一档 ≈ 20分
      - 两个维度等权重，确保选到整体最合身的替代品
    礼帽评分：号码差值直接比较。
    """
    own = person[item_type]
    available = [x for x in all_items if x not in used_items and x != own]
    gender = person['gender']
    candidates = []
    for item in available:
        if item_type == 'uniform':
            info = parse_uniform_code(item)
            if not info or info['gender'] != gender: continue
            pi = parse_uniform_code(own) or dict(height=0, chest=0)
            # 身高每cm计4分，胸围每cm计5分 → 一档身高(5cm)≈一档胸围(4cm)≈20分
            score = abs(info['height'] - pi['height']) * 4 + abs(info['chest'] - pi['chest']) * 5
            candidates.append((score, item))
        elif item_type == 'hat':
            info = parse_hat_code(item)
            if not info or info['gender'] != gender: continue
            pi = parse_hat_code(own) or dict(number=0)
            candidates.append((abs(info['number'] - pi['number']), item))
        else:
            # 马靴、腰带等非尺寸装备，任意同性别即可
            if item.startswith(gender):
                candidates.append((0, item))

    # 如果同性别实在没有，放宽到任意性别
    if not candidates:
        for item in available:
            candidates.append((999998, item))
    # 兜底：任意可用
    if not candidates:
        candidates = [(999999, x) for x in all_items if x.startswith(gender)]

    candidates.sort()
    return candidates[0][1] if candidates else None

def build_pool(wb):
    pool = dict(uniform=set(), hat=set())
    for row in range(2, wb['装备分配'].max_row+1):
        for col,key in [(3,'uniform'),(4,'hat')]:
            v = str(wb['装备分配'].cell(row=row, column=col).value or '').strip()
            if v: pool[key].add(v)
    for sn in ['礼服腰带摆放','礼帽摆放']:
        ws = wb[sn]
        for row in range(3, ws.max_row+1):
            for col in range(1, ws.max_column+1):
                v = str(ws.cell(row=row, column=col).value or '').strip()
                if re.match(r'^[FM]\d{2,3}/', v): pool['uniform'].add(v)
                if re.match(r'^[FM]\d{4}$', v): pool['hat'].add(v)
    return pool

def resolve(conflicts, persons, schedule, pool):
    reassigns, changed = [], {}
    by_day = {}
    for c in conflicts: by_day.setdefault(c['day'], []).append(c)
    for day, clist in by_day.items():
        people = next((set(s['people']) for s in schedule if s['day'] == day), set())
        used = {
            'uniform': {p['uniform'] for p in persons if p['name'] in people},
            'hat': {p['hat'] for p in persons if p['name'] in people},
            'boots': {p['boots'] for p in persons if p['name'] in people},
            'belt': {p['belt'] for p in persons if p['name'] in people}
        }
        for c in clist:
            person = next((p for p in persons if p['name'] == c['person_to_move']), None)
            if not person: continue
            t = c['item_type']
            alt = find_alternative(person, t, used[t], pool[t])
            if alt:
                used[t].discard(person[t]); used[t].add(alt)
                reassigns.append(dict(person=person['name'], item_type=t, old_item=person[t], new_item=alt))
                changed.setdefault(person['name'], {})[t] = alt
                person[t] = alt
    return reassigns, changed

def generate_excel(template_path, persons, schedule, changed):
    """按排班生成礼服分配表，匹配参考 Excel 格式。

    输出结构：
      Sheet 1「装备分配」: 按班级分段（周一~周五），每人一行，
        列：班级 | 姓名 | 礼服 | 礼帽 | 马靴 | 腰带
        重分配装备用黄色高亮，不写备注（只看结果）。
      Sheet 2-4: 保持模板原样不动。
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from copy import copy

    wb = openpyxl.load_workbook(template_path)
    ws = wb['装备分配']
    name_map = {p['name']: p for p in persons}

    # 保存原始表头行的样式（第1行）
    header_styles = {}
    for col in range(1, 9):
        cell = ws.cell(row=1, column=col)
        header_styles[col] = {
            'font': copy(cell.font), 'fill': copy(cell.fill),
            'alignment': copy(cell.alignment), 'border': copy(cell.border)
        }

    # 清空所有旧数据（保留第1行表头）
    for row in range(2, ws.max_row + 1):
        for col in range(1, 9):
            ws.cell(row=row, column=col).value = None
            ws.cell(row=row, column=col).fill = PatternFill(fill_type=None)

    # 重设第1行表头（确保被清空后恢复）
    ws.cell(row=1, column=1).value = '班级'
    ws.cell(row=1, column=2).value = '姓名'
    ws.cell(row=1, column=3).value = '礼服'
    ws.cell(row=1, column=4).value = '礼帽'
    ws.cell(row=1, column=5).value = '马靴'
    ws.cell(row=1, column=6).value = '腰带'

    yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    current_row = 2
    written = 0

    for day_entry in schedule:
        day_name = day_entry['day']
        day_people = day_entry['people']

        for name in day_people:
            p = name_map.get(name)
            ch = changed.get(name, {})

            # 班级列
            cell = ws.cell(row=current_row, column=1)
            cell.value = day_name
            cell.font = Font(size=11)
            cell.alignment = center_align
            cell.border = thin_border

            # 姓名列
            cell = ws.cell(row=current_row, column=2)
            cell.value = name
            cell.font = Font(size=11)
            cell.alignment = center_align
            cell.border = thin_border

            if p:
                equip_cols = {3: 'uniform', 4: 'hat', 5: 'boots', 6: 'belt'}
                for col_idx, key in equip_cols.items():
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = p.get(key, '')
                    cell.font = Font(size=11)
                    cell.alignment = center_align
                    cell.border = thin_border
                    if key in ch:
                        cell.fill = yellow_fill
            else:
                ws.cell(row=current_row, column=6).value = '未在库存中找到'

            # 备注列（留空，用户只关心结果）
            cell = ws.cell(row=current_row, column=7)
            cell.border = thin_border

            written += 1
            current_row += 1

        # 班级之间空一行
        current_row += 1

    # Sheet 2-4 保持模板不动（不做任何修改）

    print(f'[GenerateExcel] Wrote {written} people across {len(schedule)} days, {len(changed)} with yellow highlights')
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return base64.b64encode(buf.read()).decode()

def parse_schedule(text):
    schedule = []
    day_patterns = [
        (r'周\s*一', '周一'), (r'周\s*二', '周二'),
        (r'周\s*三', '周三'), (r'周\s*四', '周四'), (r'周\s*五', '周五'),
    ]
    for line in text.strip().split('\n'):
        line = line.strip()
        if not line: continue
        day = None
        for pat, label in day_patterns:
            if re.search(pat, line):
                day = label; break
        if not day: continue
        rest = re.sub(r'周\s*[一二三四五]\s*[:：]?\s*', '', line)
        chunks = re.split(r'[、，,]', rest)
        names = []
        for chunk in chunks:
            n = re.sub(r'\s+', '', chunk)
            if len(n) >= 2:
                names.append(n)
        if names:
            schedule.append(dict(day=day, people=names))
    return schedule

# ── 院系升旗礼服分配 ──
def uniform_sort_key(person):
    """返回排序元组: (性别序, 身高, 胸围, 序号)
    女(F)在前=0, 男(M)在后=1，空装备排最后=2"""
    suit = person.get('uniform', '')
    if not suit:
        return (2, 999, 999, 999)
    gender_code = 0 if suit.startswith('F') else 1
    nums = [int(n) for n in re.findall(r'\d+', suit)]
    while len(nums) < 3:
        nums.append(0)
    return (gender_code, nums[0], nums[1], nums[2])

def sort_people_by_uniform(people):
    """院系升旗排序：女在前男在后，同性别内尺寸从小到大"""
    return sorted(people, key=uniform_sort_key)

def detect_faculty_conflicts(persons):
    """检测队列内的装备冲突：同一件装备被多人使用即冲突"""
    conflicts = []
    for eq_type in ['uniform', 'hat', 'boots', 'belt']:
        eq_map = {}
        for p in persons:
            code = p.get(eq_type, '')
            if not code: continue
            eq_map.setdefault(code, []).append(p['name'])
        for code, names in eq_map.items():
            if len(names) >= 2:
                for i in range(1, len(names)):
                    conflicts.append(dict(
                        item_type=eq_type, item_code=code,
                        person_to_move=names[i], person_to_keep=names[0]
                    ))
    return conflicts

def resolve_faculty_conflicts(conflicts, persons, pool):
    """解决队列内的装备冲突，为冲突者找最接近尺寸的替换品"""
    reassigns, changed = [], {}
    # 当前队列所有人正在用的装备
    used = {}
    for eq_type in ['uniform', 'hat', 'boots', 'belt']:
        used[eq_type] = {p.get(eq_type, '') for p in persons if p.get(eq_type)}

    for c in conflicts:
        person = next((p for p in persons if p['name'] == c['person_to_move']), None)
        if not person: continue
        t = c['item_type']
        alt = find_alternative(person, t, used[t], pool.get(t, set()))
        if alt:
            used[t].discard(person[t]); used[t].add(alt)
            reassigns.append(dict(person=person['name'], item_type=t,
                                  old_item=person[t], new_item=alt))
            changed.setdefault(person['name'], {})[t] = alt
            person[t] = alt
    return reassigns, changed

FACULTY_TEMPLATE = r'C:\Users\夏瑞泽\xwechat_files\wxid_vnl456wm0r9h12_0c22\msg\file\2026-07\26.4.27管理学院院系升旗装备分配表.xlsx'

def generate_faculty_excel(persons, changed):
    """基于模板修改数据，不自己设计样式。

    用 openpyxl 打开 FACULTY_TEMPLATE，只清空 Sheet1 数据行并重写。
    Sheet2/3/4 保持模板原样不动。
    """
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.load_workbook(FACULTY_TEMPLATE)
    yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type='solid')

    sorted_persons = sort_people_by_uniform(persons)

    # ═══════════════════════════════════════════════════════════
    # Sheet 1: 装备分配 — 清空旧数据，写入新数据
    # ═══════════════════════════════════════════════════════════
    ws1 = wb['装备分配']

    # 只清空数据行的值（不重设字体/边框/对齐）
    for row in range(2, ws1.max_row + 1):
        for col in range(1, 7):
            ws1.cell(row=row, column=col).value = None

    # 写入排序后的人员（只 set value）
    for i, p in enumerate(sorted_persons):
        row = i + 2
        name_changed = changed.get(p['name'], {})

        ws1.cell(row=row, column=1).value = p['name']
        ws1.cell(row=row, column=2).value = p.get('uniform', '')
        ws1.cell(row=row, column=3).value = p.get('hat', '')
        ws1.cell(row=row, column=4).value = p.get('boots', '')
        ws1.cell(row=row, column=5).value = p.get('belt', '')

        if name_changed:
            notes = []
            type_names = {'uniform': '礼服', 'hat': '礼帽', 'boots': '马靴', 'belt': '腰带'}
            for t, new in name_changed.items():
                notes.append(f'{type_names.get(t, t)}→{new}')
            ws1.cell(row=row, column=6).value = '；'.join(notes)

            change_map = {'uniform': 2, 'hat': 3, 'boots': 4, 'belt': 5}
            for eq_type, col in change_map.items():
                if eq_type in name_changed:
                    ws1.cell(row=row, column=col).fill = yellow_fill

    # 清空多余行
    for row in range(len(sorted_persons) + 2, ws1.max_row + 1):
        for col in range(1, 7):
            ws1.cell(row=row, column=col).value = None

    # ═══════════════════════════════════════════════════════════
    # Sheet 2-4: 保持模板原样，不做任何修改
    # ═══════════════════════════════════════════════════════════

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return base64.b64encode(buf.read()).decode()

# ── 院系升旗页面 ──
FACULTY_HTML = r'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>院系升旗礼服分配</title>
<style>
:root {
  --flag-red: #B81616; --flag-red-dark: #8E1010; --flag-red-disabled: #D9A4A4;
  --gold: #F5C518; --gold-light: #FDE68A;
  --green-900: #0f2518; --green-700: #1a3a2a; --green-500: #2d5a3f; --green-300: #5a8a6a;
  --cream: #F5F1E6; --paper-warm: #f0ebe3; --paper-card: #fefcf8;
  --ink: #2c1810; --ink-light: #5c4a3a; --ink-faint: #9c8a7a;
  --white: #ffffff; --gray-300: #d4d4cc; --gray-500: #8a8a80;
  --font-heading: 'SimSun','KaiTi','宋体','楷体','Microsoft YaHei',serif;
  --font-body: 'Microsoft YaHei','PingFang SC',system-ui,sans-serif;
  --font-mono: 'SimSun','Consolas','Courier New',monospace;
}
*{box-sizing:border-box;margin:0;padding:0}
body{font:15px/1.8 var(--font-body);color:var(--ink);background:var(--cream);min-height:100vh;-webkit-font-smoothing:antialiased}
/* === MICRO ANIMATIONS === */
@keyframes fadeInUp{from{opacity:0;transform:translateY(20px)}to{opacity:1;transform:translateY(0)}}
@keyframes fadeInDown{from{opacity:0;transform:translateY(-16px)}to{opacity:1;transform:translateY(0)}}
/* === HEADER === */
header{background:var(--flag-red);color:var(--white);padding:20px 24px;text-align:center;position:relative;animation:fadeInDown .8s ease-out;border-bottom:2px solid rgba(245,197,24,.3)}
header a{color:rgba(255,255,255,.8);text-decoration:none;font-size:13px;position:absolute;left:24px;top:50%;transform:translateY(-50%);letter-spacing:.05em;transition:color .15s;display:flex;align-items:center;gap:4px}
header a:hover{color:var(--gold)}
h1{font-family:var(--font-heading);font-size:20px;font-weight:700;letter-spacing:.08em}
header p{font-size:12px;color:rgba(255,255,255,.65);margin-top:2px;letter-spacing:.06em}
header .header-star{color:var(--gold);font-size:11px;margin-left:8px}
/* === MAIN === */
main{max-width:860px;margin:32px auto;padding:0 24px}
/* === CARD === */
.card{background:var(--paper-card);border:1px solid rgba(184,22,22,.08);border-left:3px solid var(--gold);padding:28px;margin-bottom:20px;position:relative;animation:fadeInUp .7s ease-out both;box-shadow:0 2px 8px rgba(0,0,0,.03)}
.card:nth-child(1){animation-delay:.1s}
.card:nth-child(2){animation-delay:.2s}
.card h2{font-family:var(--font-heading);font-size:17px;display:flex;align-items:center;gap:10px;margin-bottom:18px;padding-bottom:12px;border-bottom:2px solid rgba(184,22,22,.15);letter-spacing:.05em;color:var(--green-900)}
/* === BADGE === */
.badge{width:28px;height:28px;background:var(--flag-red);color:var(--white);font-family:var(--font-heading);font-size:14px;font-weight:700;display:inline-flex;align-items:center;justify-content:center;flex-shrink:0}
/* === LABELS === */
label.block{display:block;font-family:var(--font-heading);font-weight:600;margin-bottom:6px;margin-top:14px;font-size:14px;color:var(--green-900);letter-spacing:.04em}
label.block:first-child{margin-top:0}
/* === TEXTAREAS === */
textarea{width:100%;height:130px;border:1px solid var(--ink-faint);padding:12px;font:13px var(--font-mono);resize:vertical;background:var(--cream);color:var(--ink);line-height:1.8;transition:border-color .2s}
textarea:focus{outline:none;border-color:var(--flag-red);box-shadow:0 0 0 2px rgba(184,22,22,.1)}
textarea::placeholder{color:var(--ink-faint);font-style:italic}
/* === UPLOAD ZONE === */
.up{border:2px solid var(--ink-faint);padding:32px;text-align:center;cursor:pointer;transition:all .25s;background:var(--cream);position:relative}
.up::before{content:'';position:absolute;top:8px;left:8px;width:14px;height:14px;border-top:1px solid var(--flag-red);border-left:1px solid var(--flag-red);opacity:.3}
.up::after{content:'';position:absolute;bottom:8px;right:8px;width:14px;height:14px;border-bottom:1px solid var(--flag-red);border-right:1px solid var(--flag-red);opacity:.3}
.up:hover{border-color:var(--flag-red);background:var(--paper-card)}
.up:hover::before,.up:hover::after{border-color:var(--flag-red);opacity:.7}
.up.sel{border-color:var(--green-500);background:#f2f7f3}
.up.sel::before,.up.sel::after{border-color:var(--green-500);opacity:.7}
.up input{display:none}
.up .n{font-family:var(--font-heading);font-weight:600;color:var(--green-500);margin-top:8px;display:none}
/* === OR DIVIDER === */
.or{display:flex;align-items:center;gap:14px;margin:16px 0}
.or hr{flex:1;border:none;border-top:1px solid var(--ink-faint)}
.or span{font-family:var(--font-heading);color:var(--ink-faint);font-size:13px;letter-spacing:.1em;position:relative;padding:0 8px}
.or span::before{content:'\2605';font-size:8px;position:absolute;left:-4px;top:50%;transform:translateY(-50%);color:var(--gold)}
.or span::after{content:'\2605';font-size:8px;position:absolute;right:-4px;top:50%;transform:translateY(-50%);color:var(--gold)}
/* === BUTTONS === */
.btn{width:100%;padding:14px;border:none;font-family:var(--font-heading);font-size:16px;font-weight:700;letter-spacing:.08em;cursor:pointer;transition:all .25s}
.btn-b{background:var(--flag-red);color:var(--white);box-shadow:0 2px 6px rgba(184,22,22,.2)}
.btn-b:hover:not(:disabled){background:var(--flag-red-dark);box-shadow:0 4px 12px rgba(184,22,22,.3);transform:translateY(-1px)}
.btn-b:disabled{background:var(--flag-red-disabled);color:rgba(255,255,255,.6);cursor:not-allowed;box-shadow:none;transform:none}
.btn-g{background:var(--green-700);color:var(--gold);border:1px solid var(--gold)}
.btn-g:hover:not(:disabled){background:var(--green-500);transform:translateY(-1px)}
.btn-g:disabled{background:var(--gray-300);color:var(--gray-500);border-color:var(--gray-300);cursor:not-allowed;transform:none}
/* === SPINNER === */
.spin{width:22px;height:22px;border:2.5px solid rgba(245,197,24,.2);border-top-color:var(--gold);border-radius:50%;animation:s .7s linear infinite}
@keyframes s{to{transform:rotate(360deg)}}
/* === ROLES GRID === */
.roles{display:grid;grid-template-columns:1fr 1fr;gap:14px}
.roles textarea{height:70px}
/* === STATS === */
.stats{display:grid;grid-template-columns:repeat(3,1fr);gap:0;border:1px solid var(--ink-faint);margin:16px 0}
.st{padding:20px 16px;text-align:center;border-right:1px solid var(--ink-faint);background:var(--paper-card);position:relative}
.st:last-child{border-right:none}
.st.c{border-top:3px solid var(--flag-red)}.st.r{border-top:3px solid var(--gold)}.st.p{border-top:3px solid var(--green-500)}
.st b{font-family:var(--font-heading);font-size:30px;font-weight:700;display:block;line-height:1.1}
.st.c b{color:var(--flag-red)}.st.r b{color:var(--gold)}.st.p b{color:var(--green-500)}
.st span{color:var(--ink-light);font-size:12px;letter-spacing:.08em;text-transform:uppercase}
/* === TABLE === */
table{width:100%;border-collapse:collapse;font-size:13px;margin-top:8px;border:1px solid var(--ink-faint)}
th{background:var(--green-900);color:var(--gold-light);padding:8px 10px;text-align:center;font-family:var(--font-heading);font-weight:600;font-size:12px;letter-spacing:.05em;text-transform:uppercase;border:1px solid rgba(245,197,24,.15)}
td{padding:8px 10px;border:1px solid var(--ink-faint);text-align:center}
tr:nth-child(even) td{background:var(--paper-warm)}
.old{text-decoration:line-through;color:var(--ink-faint);font-size:12px}
.hl{background:#fffde7;font-weight:600}
/* === HINT / ERROR === */
.hint{display:none;background:#fffde7;border-left:3px solid var(--gold);padding:10px 14px;margin-top:10px;font-size:13px;color:#6b5a20}
.err{display:none;background:#fdf0f2;border-left:3px solid var(--flag-red);padding:12px 14px;color:var(--flag-red);margin-top:12px;font-size:13px}
/* === LOADING === */
#load{text-align:center;padding:40px;display:none;animation:fadeInUp .4s ease-out}
#load p{color:var(--ink-light);margin-top:12px;font-size:13px}
/* === FOOTER === */
footer{text-align:center;padding:28px;font-family:var(--font-heading);font-size:12px;color:var(--green-700);letter-spacing:.06em;border-top:1px solid rgba(184,22,22,.08)}
footer .footer-star{color:var(--gold);font-size:10px;margin:0 6px}
/* === SECTION TITLES === */
#pt h3,#rt h3{font-family:var(--font-heading);font-size:15px;font-weight:700;margin:16px 0 6px;color:var(--green-900);letter-spacing:.05em}
</style>
</head>
<body>
<header>
<a href="/"><span>&#8592;</span> 返回首页</a>
<h1>&#127891; 院系升旗礼服分配 <span class="header-star">&#9733;</span></h1>
<p>上传库存表 &#8594; 拍照识别人名 &#8594; 生成分配表</p>
</header>
<main>

<div class="card">
<h2><span class="badge">1</span> 上传礼服库存表 (.xlsx)</h2>
<div class="up" id="ea"><div style="font-size:36px;margin-bottom:8px">&#128206;</div><div>上传包含所有人员装备信息的Excel</div><div style="color:var(--ink-faint);font-size:13px">系统将从中提取装备库存</div>
<input type="file" id="ei" accept=".xlsx"/><div class="n" id="en"></div></div>
</div>

<div class="card">
<h2><span class="badge">2</span> 拍照上传人员安排表</h2>
<div class="up" id="ia"><div style="font-size:36px;margin-bottom:8px">&#128247;</div><div>点击上传人员安排表照片（手写也可以）</div><div style="color:var(--ink-faint);font-size:13px">AI自动识别队列人员姓名</div>
<input type="file" id="ii" accept="image/*"/><div class="n" id="inm"></div></div>
<div class="hint" id="ohint">&#9888; OCR识别结果已填入下方，请核对修正后再点生成</div>
<div class="or"><hr><span>或手动输入 / 修正队列人员</span><hr></div>
<p style="color:var(--ink-light);font-size:13px;margin-bottom:8px">每行一人或顿号分隔。总负责、场控、后勤、摄影不参与分配。</p>
<textarea id="rp" placeholder="林珩&#10;韩雅丽&#10;艾克达&#10;张鹏&#10;夏瑞泽&#10;戴傲&#10;叶宇轩"></textarea>
<div class="roles">
<div><label class="block">总负责</label><textarea id="r1" rows="2" placeholder="（可选，不参与分配）"></textarea></div>
<div><label class="block">场控</label><textarea id="r2" rows="2" placeholder="（可选，不参与分配）"></textarea></div>
<div><label class="block">后勤</label><textarea id="r3" rows="2" placeholder="（可选，不参与分配）"></textarea></div>
<div><label class="block">摄影</label><textarea id="r4" rows="2" placeholder="（可选，不参与分配）"></textarea></div>
</div>
</div>

<button class="btn btn-b" id="gb" disabled>&#128269; 预览排序 &amp; 生成分配表</button>

<div class="err" id="er"></div>
<div id="load"><div class="spin" style="display:inline-block;border-color:rgba(245,197,24,.2);border-top-color:var(--gold);width:36px;height:36px"></div><p>正在识别并生成分配方案...</p></div>

<div id="res" style="display:none"><div class="card">
<h2><span class="badge">3</span> 生成结果</h2>
<p style="color:var(--ink-light);margin-bottom:16px" id="rm"></p>
<div class="stats" id="sr" style="display:none">
<div class="st c"><b id="cc">0</b><span>冲突数</span></div>
<div class="st r"><b id="rc">0</b><span>重分配</span></div>
<div class="st p"><b id="ac">0</b><span>受影响人数</span></div>
</div>
<div id="pt"></div>
<div id="rt"></div>
<div style="margin-top:20px"><button class="btn btn-g" id="db" disabled>&#128229; 下载礼服分配表 (.xlsx)</button></div>
</div></div>
</main>
<footer><span class="footer-star">&#9733;</span> 浙江大学国旗仪仗队 &copy; 2026 <span class="footer-star">&#9733;</span></footer>
<script>
var ef=null,imgf=null,xb64='';
function ua(id,inpId,nmId,cb){
  var a=document.getElementById(id),inp=document.getElementById(inpId),nm=document.getElementById(nmId);
  a.addEventListener('click',function(){inp.click()});
  a.addEventListener('dragover',function(e){e.preventDefault()});
  a.addEventListener('drop',function(e){e.preventDefault();var f=e.dataTransfer.files[0];if(f)cb(f,a,nm)});
  inp.addEventListener('change',function(){var f=inp.files[0];if(f)cb(f,a,nm)});
}
function sf(f,a,nm){a.classList.add('sel');nm.style.display='block';nm.textContent=f.name;chk()}
ua('ea','ei','en',function(f,a,nm){ef=f;sf(f,a,nm)});
ua('ia','ii','inm',function(f,a,nm){imgf=f;sf(f,a,nm);document.getElementById('ohint').style.display='block'});
document.getElementById('rp').addEventListener('input',chk);
function chk(){
  var roster = document.getElementById('rp').value.trim();
  document.getElementById('gb').disabled=!(ef && (roster || imgf));
}
document.getElementById('gb').addEventListener('click',async function(){
  if(!ef)return;
  document.getElementById('er').style.display='none';
  document.getElementById('res').style.display='none';
  document.getElementById('load').style.display='block';
  document.getElementById('gb').disabled=true;
  try{
    var fd=new FormData();
    fd.append('excel',ef);
    fd.append('roster',document.getElementById('rp').value.trim());
    if(imgf)fd.append('image',imgf);
    var meta = ['总负责','场控','后勤','摄影'];
    for(var i=1;i<=4;i++){
      var v = document.getElementById('r'+i).value.trim();
      if(v) fd.append('meta_'+i, v);
    }
    var r=await fetch('/generate-faculty',{method:'POST',body:fd});
    if(!r.ok){var t=await r.text();throw new Error(t.substring(0,500))}
    show(await r.json());
  }catch(e){
    document.getElementById('er').textContent='生成失败: '+e.message;
    document.getElementById('er').style.display='block';
  }
  document.getElementById('load').style.display='none';chk();
});
function show(d){
  document.getElementById('res').style.display='block';
  document.getElementById('rm').textContent=d.message;
  document.getElementById('sr').style.display='grid';
  document.getElementById('cc').textContent=d.conflicts_count;
  document.getElementById('rc').textContent=d.reassignments_count;
  document.getElementById('ac').textContent=new Set(d.reassignments.map(function(r){return r.person})).size;
  if(d.ocr_text){
    document.getElementById('rp').value=d.ocr_text;
    document.getElementById('ohint').style.display='block';
  }
  if(d.meta_people){
    if(d.meta_people['总负责']) document.getElementById('r1').value=d.meta_people['总负责'].join('\n');
    if(d.meta_people['场控']) document.getElementById('r2').value=d.meta_people['场控'].join('\n');
    if(d.meta_people['后勤']) document.getElementById('r3').value=d.meta_people['后勤'].join('\n');
    if(d.meta_people['摄影']) document.getElementById('r4').value=d.meta_people['摄影'].join('\n');
  }
  xb64=d.excel_base64;
  document.getElementById('db').disabled=false;
  document.getElementById('db').onclick=function(){
    var b=atob(xb64),u8=new Uint8Array(b.length);
    for(var i=0;i<b.length;i++)u8[i]=b.charCodeAt(i);
    var blob=new Blob([u8],{type:'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'});
    var a=document.createElement('a');a.href=URL.createObjectURL(blob);a.download='院系升旗礼服分配.xlsx';a.click();
  };
  var h='<h3>排序后人员列表</h3><table><tr><th>序号</th><th>姓名</th><th>性别</th><th>礼服</th><th>礼帽</th><th>马靴</th><th>腰带</th></tr>';
  d.sorted_persons.forEach(function(p,i){
    var ch = d.changes[p.name] || {};
    h+='<tr>';
    h+='<td>'+(i+1)+'</td>';
    h+='<td><b>'+p.name+'</b></td>';
    h+='<td>'+(p.gender==='F'?'女':'男')+'</td>';
    ['uniform','hat','boots','belt'].forEach(function(k){
      var v = p[k]||'';
      var cls = ch[k] ? 'hl' : '';
      var oldv = ch[k] ? '<br><span class=old>'+d.original_equip[p.name+'|'+k]+'→</span>' : '';
      h+='<td class='+cls+'>'+v+oldv+'</td>';
    });
    h+='</tr>';
  });
  h+='</table>';
  document.getElementById('pt').innerHTML=h;
  if(d.conflicts.length){
    h='<h3>冲突详情</h3><table><tr><th>装备类型</th><th>编号</th><th>需变动</th><th>保留</th></tr>';
    d.conflicts.forEach(function(c){h+='<tr><td>'+c.item_type+'</td><td>'+c.item_code+'</td><td>'+c.person_to_move+'</td><td>'+c.person_to_keep+'</td></tr>'});
    h+='</table>';
  }
  if(d.reassignments.length){
    h='<h3>重分配方案</h3><table><tr><th>姓名</th><th>装备</th><th>旧编号</th><th>新编号</th></tr>';
    d.reassignments.forEach(function(r){h+='<tr><td>'+r.person+'</td><td>'+r.item_type+'</td><td class=old>'+r.old_item+'</td><td class=hl>'+r.new_item+'</td></tr>'});
    h+='</table>';
  }
  document.getElementById('rt').innerHTML=h;
  document.getElementById('res').scrollIntoView({behavior:'smooth'});
}
</script>
</body></html>'''
# ── 日常升旗页面（从旧首页移过来）──
HOME_HTML = r'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>礼服自动分配系统</title>
<style>
:root {
  --flag-red: #B81616; --flag-red-dark: #8E1010; --flag-red-disabled: #D9A4A4;
  --gold: #F5C518; --gold-light: #FDE68A;
  --green-900: #0f2518; --green-700: #1a3a2a; --green-500: #2d5a3f;
  --cream: #F5F1E6; --white: #FFFFFF;
  --ink: #2c1810; --ink-light: #5c4a3a; --ink-faint: #9c8a7a;
  --font-heading: 'SimSun','KaiTi','宋体','楷体','Microsoft YaHei',serif;
  --font-body: 'Microsoft YaHei','PingFang SC',system-ui,sans-serif;
}
*{box-sizing:border-box;margin:0;padding:0}
/* === PAGE === */
body{font:15px/1.8 var(--font-body);color:var(--ink);background:var(--cream);min-height:100vh;-webkit-font-smoothing:antialiased}
/* Subtle star watermark - very faint */
body::before{content:'';position:fixed;inset:0;pointer-events:none;z-index:0;
  background:radial-gradient(circle at 15% 20%, rgba(245,197,24,0.03) 0%, transparent 40%),
             radial-gradient(circle at 85% 60%, rgba(184,22,22,0.02) 0%, transparent 40%),
             radial-gradient(circle at 50% 90%, rgba(245,197,24,0.02) 0%, transparent 30%);
}
/* Top gradient transition */
body::after{content:'';position:absolute;top:0;left:0;right:0;height:120px;pointer-events:none;z-index:0;
  background:linear-gradient(180deg, rgba(184,22,22,0.06) 0%, transparent 100%);}

/* === FADE IN UP ANIMATION === */
@keyframes fadeInUp{from{opacity:0;transform:translateY(20px)}to{opacity:1;transform:translateY(0)}}
@keyframes fadeInDown{from{opacity:0;transform:translateY(-16px)}to{opacity:1;transform:translateY(0)}}

/* === HEADER === */
header{background:var(--flag-red);color:var(--white);padding:20px 32px;display:flex;align-items:center;justify-content:space-between;position:relative;z-index:1;animation:fadeInDown .8s ease-out;border-bottom:2px solid rgba(245,197,24,.3)}
header .brand{display:flex;align-items:center;gap:12px}
header h1{font-family:var(--font-heading);font-size:18px;font-weight:700;letter-spacing:.08em;color:var(--white)}
header .stars{display:flex;gap:4px;font-size:14px;color:var(--gold);line-height:1}
/* === MAIN === */
main{max-width:760px;margin:56px auto 48px;padding:0 24px;position:relative;z-index:1}
/* === ENTRIES GRID === */
.entries{display:grid;grid-template-columns:1fr 1fr;gap:20px;animation:fadeInUp .8s ease-out .15s both}
@media(max-width:640px){.entries{grid-template-columns:1fr}}
/* === CARD === */
.card{display:block;background:var(--white);border:1px solid rgba(184,22,22,.08);border-left:3px solid var(--gold);border-radius:8px;padding:40px 28px;text-align:center;text-decoration:none;color:inherit;cursor:pointer;transition:all .3s ease;box-shadow:0 2px 12px rgba(0,0,0,.04);position:relative;overflow:hidden}
.card:hover{border-left-color:var(--flag-red);transform:translateY(-4px);box-shadow:0 8px 28px rgba(184,22,22,.1)}
.card .icon{font-size:42px;margin-bottom:14px;display:block;transition:transform .3s}
.card:hover .icon{transform:scale(1.05)}
.card .star-deco{display:inline-block;color:var(--gold);font-size:12px;margin:0 6px}
.card h2{font-family:var(--font-heading);font-size:18px;font-weight:700;color:var(--green-900);margin-bottom:10px;letter-spacing:.06em}
.card p{font-size:13px;color:var(--ink-light);line-height:1.7}
/* Gold line accent at top of card */
.card::before{content:'';display:block;width:40px;height:2px;background:var(--gold);margin:0 auto 16px;transition:width .3s,background .3s}
.card:hover::before{width:56px;background:var(--flag-red)}
/* === FOOTER === */
footer{text-align:center;padding:32px;font-family:var(--font-heading);font-size:12px;color:var(--green-700);letter-spacing:.06em;position:relative;z-index:1;border-top:1px solid rgba(184,22,22,.08)}
footer .footer-star{color:var(--gold);font-size:10px;margin:0 6px}
</style>
</head>
<body>
<header>
<div class="brand"><span class="stars">&#9733;</span><h1>浙江大学国旗仪仗队</h1></div>
<div class="stars">&#9733;</div>
</header>
<main>
<div class="entries">
<a href="/faculty" class="card">
  <div class="icon">&#127891;</div>
  <h2>院系升旗礼服分配 <span class="star-deco">&#9733;</span></h2>
  <p>上传礼服库存表 + 拍照上传人员安排表<br>自动识别 &#8594; 按尺寸排序 &#8594; 冲突检测 &#8594; 下载分配表</p>
</a>
<a href="/warehouse" class="card">
  <div class="icon">&#128230;</div>
  <h2>物资仓库 <span class="star-deco">&#9733;</span></h2>
  <p>马靴 · 腰带 · 礼服 存放位置<br>搜索库位 / 姓名 / 尺码</p>
</a>
</div>
</main>
<footer><span class="footer-star">&#9733;</span> 浙江大学国旗仪仗队 &copy; 2026 <span class="footer-star">&#9733;</span></footer>
</body></html>'''


# ── Routes ──
@app.route('/')
def index():
    return HOME_HTML

@app.route('/faculty')
def faculty():
    return FACULTY_HTML

@app.route('/warehouse')
def warehouse():
    with open('warehouse.html', 'r', encoding='utf-8') as f:
        return f.read()

@app.route('/generate', methods=['POST'])
def generate():
    import openpyxl
    from openpyxl.styles import PatternFill

    excel_file = request.files.get('excel')
    if not excel_file:
        return '请上传 Excel 文件', 400

    excel_bytes = excel_file.read()
    schedule_text = request.form.get('schedule', '').strip()
    image_file = request.files.get('image')
    ocr_parsed_text = ''

    # ── OCR 图片（EasyOCR 手写中文识别）──
    if image_file and image_file.filename:
        img_bytes = image_file.read()
        if len(img_bytes) > 0:
            from PIL import Image
            with tempfile.NamedTemporaryFile(suffix='.jpg', delete=False) as t:
                t.write(img_bytes); ip = t.name
            try:
                img = Image.open(ip).convert('RGB')
                w, h = img.size

                # 从 Excel 加载已知人名
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as et:
                    et.write(excel_bytes); etmp = et.name
                try:
                    wb_tmp = openpyxl.load_workbook(etmp)
                    known_names = []
                    for row in range(2, wb_tmp['装备分配'].max_row + 1):
                        n = wb_tmp['装备分配'].cell(row=row, column=2).value
                        if n and str(n).strip():
                            known_names.append(str(n).strip())
                    wb_tmp.close()
                finally:
                    os.unlink(etmp)

                # 只降不升（避免小图放大失真），2000px + low_text=0.2
                TARGET_W = min(w, 2000)
                scale = TARGET_W / w
                new_w, new_h = TARGET_W, int(h * scale)
                img = img.resize((new_w, new_h), Image.LANCZOS)
                pre = ip + '_pre.png'; img.save(pre)

                import easyocr, numpy as np, math
                reader = easyocr.Reader(['ch_sim'], gpu=False)

                print(f'[Daily-OCR] Image: {w}x{h}', flush=True)

                DAY_LABELS = ['周一', '周二', '周三', '周四', '周五']
                day_schedule = {}

                # 全图粗扫（宽松参数，抓取更多文本块）
                img_np = np.array(img)
                if len(img_np.shape) == 3:
                    gray = np.mean(img_np, axis=2).astype(np.uint8)
                else:
                    gray = img_np

                results = reader.readtext(img_np, detail=1, low_text=0.2, text_threshold=0.3)

                # 过滤边缘噪点
                filtered = []
                for bbox, text, conf in results:
                    x1, y1 = bbox[0]; x3, y3 = bbox[2]
                    t = text.strip()
                    if not t: continue
                    if x1 <= 5 and y1 <= 5: continue
                    if x3 >= new_w - 5 and y3 >= new_h - 5: continue
                    filtered.append((bbox, t, conf))
                results = filtered

                # 构建文本块
                blocks = []
                for bbox, text, conf in results:
                    x1, y1 = bbox[0]; x3, y3 = bbox[2]
                    blocks.append({'text': text, 'x1': x1, 'x3': x3, 'cx': (x1 + x3) / 2,
                                   'y1': y1, 'y3': y3, 'conf': conf})

                # X 聚类分列（间隔 > 40px 视为不同列）
                blocks.sort(key=lambda b: b['cx'])
                x_clusters = []
                if blocks:
                    x_clusters = [[blocks[0]]]
                    for b in blocks[1:]:
                        if b['cx'] - x_clusters[-1][-1]['cx'] < 40:
                            x_clusters[-1].append(b)
                        else:
                            x_clusters.append([b])
                print(f'[Daily-OCR] X-clusters: {len(x_clusters)}, sizes: {[len(c) for c in x_clusters]}')

                if len(x_clusters) >= 5:
                    # 多于 5 列：取最右 5 列
                    if len(x_clusters) > 5:
                        x_clusters = x_clusters[-5:]

                    # 黄色背景检测 — 关闭，全图色偏导致整列误判为黄色
                    # TODO: 当有真实黄色高亮排班表时再启用
                    # img_hsv = cv2.cvtColor(img_np, cv2.COLOR_RGB2HSV)
                    # yellow_low = np.array([20, 80, 150])
                    # yellow_high = np.array([35, 255, 255])
                    # yellow_mask = cv2.inRange(img_hsv, yellow_low, yellow_high)

                    for ci, cluster in enumerate(x_clusters):
                        day = DAY_LABELS[ci] if ci < 5 else ''
                        cluster.sort(key=lambda b: b['y1'])

                        # 暴力匹配所有文本块
                        seen = set()
                        for b in cluster:
                            t = b['text'].strip()
                            if len(t) < 2: continue
                            corrected = force_correct_name(t, known_names)
                            if not corrected:
                                corrected = fuzzy_match_aggressive(t, known_names)
                            if corrected and corrected not in seen:
                                seen.add(corrected)

                        if seen:
                            day_schedule.setdefault(day, []).extend(seen)
                            print(f'[Daily-OCR] {day}: {len(seen)} names from {len(cluster)} blocks', flush=True)

                    # 去重

                # 全图 fallback（X 聚类无结果时用全图按日期标题分列）
                if not day_schedule:
                    print('[Daily-OCR] X-cluster failed, trying full-image fallback')
                    results = reader.readtext(np.array(img), detail=1, low_text=0.2, text_threshold=0.4)
                    filtered = []
                    for bbox, text, conf in results:
                        x1, y1 = bbox[0]; x3, y3 = bbox[2]
                        t = text.strip()
                        if not t: continue
                        if x1 <= 5 and y1 <= 5: continue
                        if x3 >= new_w - 5 and y3 >= new_h - 5: continue
                        filtered.append((bbox, t, conf))
                    results = filtered

                    blocks = []
                    for bbox, text, conf in results:
                        x1, y1 = bbox[0]; x3, y3 = bbox[2]
                        blocks.append({'text': text, 'x1': x1, 'x3': x3, 'cx': (x1 + x3) / 2, 'y1': y1, 'y3': y3})

                    days_info = []
                    name_entries = []
                    for b in blocks:
                        t = b['text'].strip()
                        if any(d in t for d in DAY_LABELS):
                            days_info.append({'day': t, 'x1': b['x1'], 'x2': b['x3'], 'cx': b['cx']})
                        elif len(t) >= 2:
                            name_entries.append({'name': t, 'x': b['cx'], 'y': b['y1']})

                    days_info.sort(key=lambda d: d['x1'])
                    grouped = {d['day']: [] for d in days_info}
                    if days_info:
                        col_width = (days_info[-1]['x2'] - days_info[0]['x1']) / len(days_info)
                        for n in name_entries:
                            best_d, best_dist = None, 99999
                            for d in days_info:
                                dist = abs(n['x'] - d['cx'])
                                if dist < best_dist: best_dist, best_d = dist, d
                            if best_d and best_dist < col_width * 1.3:
                                grouped[best_d['day']].append((n['y'], n['name']))

                    for day_label, nl in grouped.items():
                        seen = set()
                        unique = []
                        for _, raw_name in sorted(nl):
                            corrected = force_correct_name(raw_name, known_names)
                            if not corrected:
                                corrected = fuzzy_match_aggressive(raw_name, known_names)
                                if not corrected:
                                    continue
                            if corrected not in seen:
                                seen.add(corrected); unique.append(corrected)
                        if unique:
                            day_schedule[day_label] = unique

                os.unlink(pre)

                # 输出结果
                ocr_schedule = []
                for day in DAY_LABELS:
                    names = day_schedule.get(day, [])
                    seen = set()
                    unique = [n for n in names if not (n in seen or seen.add(n))]
                    if unique:
                        ocr_schedule.append(dict(day=day, people=unique))

                print(f'[Daily-OCR] Schedule: {[(d["day"], d["people"]) for d in ocr_schedule]}')

                if ocr_schedule:
                    ocr_parsed_text = '\n'.join(
                        d['day'] + '：' + '、'.join(d['people']) for d in ocr_schedule
                    )
                    schedule_text = ocr_parsed_text

            except Exception as e:
                import traceback
                print(f'[OCR ERROR] {e}')
                traceback.print_exc()
            finally:
                try: os.unlink(ip)
                except: pass
    if not schedule_text:
        return '请提供排班文本或上传排班照片（照片需要能看清文字）', 400

    schedule = parse_schedule(schedule_text)
    if not schedule:
        return f'排班格式无法解析，请检查格式。\n当前内容:\n{schedule_text[:300]}', 400

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as t:
        t.write(excel_bytes); tmp = t.name

    try:
        wb = openpyxl.load_workbook(tmp)
        persons = parse_equipment_sheet(wb['装备分配'])
        print(f'[Generate] Parsed {len(persons)} persons from Excel')
        rels = parse_shared_relations(persons)
        mrels = merge_relations(rels)
        print(f'[Generate] Shared relations: {len(mrels)} groups')
        for r in mrels:
            print(f'  {r["item_type"]} {r["item_code"]} shared by {r["shared_by"]}')
        conflicts = detect_conflicts(schedule, mrels)
        print(f'[Generate] Schedule days: {[s["day"] for s in schedule]}')
        for s in schedule:
            print(f'  {s["day"]}: {s["people"][:5]}... total {len(s["people"])}')
        print(f'[Generate] Conflicts detected: {len(conflicts)}')
        for c in conflicts:
            print(f'  {c["day"]} {c["item_type"]} {c["item_code"]}: move {c["person_to_move"]}, keep {c["person_to_keep"]}')
        pool = build_pool(wb)
        # 为 boots/belt 提供空池（当前只管理 uniform/hat）
        for key in ['boots', 'belt']:
            if key not in pool:
                pool[key] = set()
        print(f'[Generate] Pool: {len(pool["uniform"])} uniforms, {len(pool["hat"])} hats')
        reassigns, changed = resolve(conflicts, persons, schedule, pool)
        print(f'[Generate] Reassignments: {len(reassigns)}, changed persons: {list(changed.keys())}')
        for r in reassigns:
            print(f'  {r["person"]}: {r["item_type"]} {r["old_item"]} -> {r["new_item"]}')
        b64 = generate_excel(tmp, persons, schedule, changed)

        return jsonify(
            success=True,
            message=f'生成完成: {len(conflicts)} 个冲突, {len(reassigns)} 个重分配',
            conflicts_count=len(conflicts),
            reassignments_count=len(reassigns),
            conflicts=[dict(day=c['day'],
                item_type='礼服' if c['item_type']=='uniform' else '礼帽',
                item_code=c['item_code'],
                person_to_move=c['person_to_move'],
                person_to_keep=c['person_to_keep']) for c in conflicts],
            reassignments=[dict(person=r['person'],
                item_type='礼服' if r['item_type']=='uniform' else '礼帽',
                old_item=r['old_item'], new_item=r['new_item']) for r in reassigns],
            excel_base64=b64,
            ocr_text=(ocr_parsed_text if image_file and image_file.filename else ''),
        )
    finally:
        os.unlink(tmp)


@app.route('/generate-faculty', methods=['POST'])
def generate_faculty():
    """院系升旗礼服分配"""
    import openpyxl

    excel_file = request.files.get('excel')
    if not excel_file:
        return '请上传礼服库存 Excel 文件', 400

    excel_bytes = excel_file.read()
    roster_text = request.form.get('roster', '').strip()
    image_file = request.files.get('image')
    ocr_parsed_text = ''
    ocr_roles = {}

    # ── OCR 识别（5步列聚类方案，与 app.py 完全一致）──
    if image_file and image_file.filename:
        img_bytes = image_file.read()
        if len(img_bytes) > 0:
            from PIL import Image
            with tempfile.NamedTemporaryFile(suffix='.jpg', delete=False) as t:
                t.write(img_bytes); ip = t.name
            try:
                img = Image.open(ip).convert('RGB')
                w, h = img.size

                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as et:
                    et.write(excel_bytes); etmp = et.name
                try:
                    wb_tmp = openpyxl.load_workbook(etmp)
                    known_names = []
                    for row in range(2, wb_tmp['装备分配'].max_row + 1):
                        n = wb_tmp['装备分配'].cell(row=row, column=2).value
                        if n and str(n).strip():
                            known_names.append(str(n).strip())
                    wb_tmp.close()
                finally:
                    os.unlink(etmp)

                # Step 1: 只降不升（避免小图放大失真），2000px + low_text=0.2
                TARGET_W = min(w, 2000)
                scale = TARGET_W / w
                new_w, new_h = TARGET_W, int(h * scale)
                img = img.resize((new_w, new_h), Image.LANCZOS)
                pre = ip + '_pre.png'; img.save(pre)

                import easyocr, numpy as np, math
                reader = easyocr.Reader(['ch_sim'], gpu=False)
                results = reader.readtext(np.array(img), detail=1, low_text=0.2, text_threshold=0.4)

                print(f'[Faculty-OCR] Image: {w}x{h} -> {new_w}x{new_h}, {len(results)} text regions')

                # 过滤真正贴边的噪点（绝对5px，不用百分比，避免误删小图片中的文字）
                filtered = []
                for bbox, text, conf in results:
                    x1, y1 = bbox[0]; x3, y3 = bbox[2]
                    t = text.strip()
                    if not t: continue
                    if x1 <= 5 and y1 <= 5: continue
                    if x3 >= new_w - 5 and y3 >= new_h - 5: continue
                    filtered.append((bbox, t, conf))
                results = filtered

                blocks = []
                for bbox, text, conf in results:
                    x1, y1 = bbox[0]; x3, y3 = bbox[2]
                    blocks.append({
                        'text': text, 'x1': x1, 'x3': x3,
                        'cx': (x1 + x3) / 2, 'y1': y1, 'y3': y3
                    })

                # Step 2: 固定列宽 = 图片宽度 / 8
                COL_W = new_w / 8
                blocks.sort(key=lambda b: b['cx'])
                col_buckets = {}
                for b in blocks:
                    col_idx = int(b['cx'] // COL_W)
                    if col_idx not in col_buckets:
                        col_buckets[col_idx] = []
                    col_buckets[col_idx].append(b)

                sorted_cols = sorted(col_buckets.items())
                merged_cols = []
                for col_idx, col_blocks in sorted_cols:
                    if merged_cols and col_idx - merged_cols[-1][0] <= 1:
                        prev_max_x = max(b['x3'] for b in merged_cols[-1][1])
                        cur_min_x = min(b['x1'] for b in col_blocks)
                        if cur_min_x - prev_max_x < COL_W * 0.3:
                            merged_cols[-1] = (merged_cols[-1][0], merged_cols[-1][1] + col_blocks)
                            continue
                    merged_cols.append((col_idx, col_blocks))

                # Step 3: 标记角色列，左边所有列全算队列
                ROLE_KW = ['总负责', '场控', '后勤', '摄影']
                QUEUE_TITLE_KW = ['擎护旗', '队列']

                first_role_col_idx = None
                for i, (_, col_blocks) in enumerate(merged_cols):
                    col_text = ' '.join(b['text'] for b in col_blocks)
                    if any(kw in col_text for kw in ROLE_KW):
                        first_role_col_idx = i
                        break

                if first_role_col_idx is not None:
                    col_types = ['role' if i >= first_role_col_idx else 'queue' for i in range(len(merged_cols))]
                else:
                    col_types = ['queue' if i < 4 else 'role' for i in range(len(merged_cols))]

                # Step 4: 人名清洗 + 单字合并 + 模糊匹配 + 全图碎片兜底
                def server_fuzzy_match(name, name_list):
                    """字符重合度 >= 50% 就尝试还原"""
                    if not name or len(name) < 1: return name
                    if name in name_list: return name
                    nc = set(name)
                    best, best_score = name, 0
                    for ref in name_list:
                        rc = set(ref)
                        common = len(nc & rc)
                        overlap = common / max(len(nc), len(rc))
                        seq = SequenceMatcher(None, name, ref).ratio()
                        score = overlap * 10 + seq * 5
                        if score > best_score:
                            best_score = score
                            best = ref
                    if best_score >= 5 and best != name:
                        return best
                    return name

                # Step 4: 提取队列人名 + 角色人名
                def clean_name_block(blocks_in_col):
                    """单字合并 + 过滤关键词，返回清洗后的文本列表"""
                    blocks_in_col.sort(key=lambda b: b['y1'])
                    merged = []
                    i = 0
                    while i < len(blocks_in_col):
                        b = blocks_in_col[i]
                        t = b['text'].strip()
                        if len(t) == 1 and i + 1 < len(blocks_in_col):
                            nb = blocks_in_col[i + 1]
                            char_h = b['y3'] - b['y1']
                            if char_h > 0 and nb['y1'] - b['y3'] < char_h * 3.0:
                                t = t + nb['text'].strip()
                                i += 1
                        merged.append(t)
                        i += 1
                    cleaned = []
                    for t in merged:
                        if not t: continue
                        if t in QUEUE_TITLE_KW: continue
                        if t in ROLE_KW: continue
                        for kw in QUEUE_TITLE_KW + ROLE_KW:
                            if kw in t and t != kw:
                                t = t.replace(kw, '').strip()
                        if t:
                            cleaned.append(t)
                    return cleaned

                queue_raw = []
                ocr_roles = {}  # { '总负责': [names...], '场控': [...] }

                # 先收集所有块（按列类型归类）
                all_role_blocks = []
                for idx, (_, col_blocks) in enumerate(merged_cols):
                    if col_types[idx] == 'queue':
                        for t in clean_name_block(col_blocks):
                            corrected = server_fuzzy_match(t, known_names)
                            if corrected not in queue_raw:
                                queue_raw.append(corrected)
                    else:
                        all_role_blocks.extend(col_blocks)

                # 所有角色块统一按 y 分段，每个角色关键词下方的人名归该角色
                if all_role_blocks:
                    all_role_blocks.sort(key=lambda b: b['y1'])
                    role_markers = []  # [(y, keyword), ...]
                    name_blocks = []
                    for b in all_role_blocks:
                        t = b['text'].strip()
                        if t in ROLE_KW:
                            role_markers.append((b['y1'], t))
                        else:
                            name_blocks.append(b)
                    if role_markers:
                        role_markers.sort(key=lambda x: x[0])
                        for _, rk in role_markers:
                            ocr_roles.setdefault(rk, [])
                        # 每个人名匹配最近的（同行 y 有 OCR 抖动，不用区间）
                        for nb in name_blocks:
                            best_role, best_dy = None, float('inf')
                            for ry, rk in role_markers:
                                dy = abs(nb['y1'] - ry)
                                if dy < best_dy:
                                    best_dy, best_role = dy, rk
                            if best_role:
                                for ct in clean_name_block([nb]):
                                    corrected = server_fuzzy_match(ct, known_names)
                                    if corrected not in ocr_roles[best_role]:
                                        ocr_roles[best_role].append(corrected)

                # 从队列中移除角色人员
                role_people = set()
                for names in ocr_roles.values():
                    role_people.update(names)
                queue_raw = [n for n in queue_raw if n not in role_people]

                # 兜底：全图碎片字符级匹配
                all_text = ''.join(b['text'] for b in blocks)
                all_chars = set(all_text)
                for ref_name in known_names:
                    if ref_name in queue_raw or ref_name in role_people: continue
                    rc = set(ref_name)
                    hit = all_chars & rc
                    if len(hit) >= max(2, len(rc) * 0.5):
                        for b in blocks:
                            matched = server_fuzzy_match(b['text'], known_names)
                            if matched == ref_name:
                                if matched not in queue_raw and matched not in role_people:
                                    # 判断是否属于角色列
                                    col_i = int(b['cx'] // COL_W)
                                    is_role = False
                                    if first_role_col_idx is not None:
                                        for mi, (mci, _) in enumerate(merged_cols):
                                            if col_i == mci and col_types[mi] == 'role':
                                                is_role = True
                                                break
                                    if is_role:
                                        if ref_name not in role_people:
                                            queue_raw.append(ref_name)
                                    else:
                                        if ref_name not in queue_raw:
                                            queue_raw.append(ref_name)
                                break

                print(f'[Faculty-OCR] {len(merged_cols)} columns, types={col_types}')
                print(f'[Faculty-OCR] Queue names ({len(queue_raw)}): {queue_raw}')
                print(f'[Faculty-OCR] Role people: {ocr_roles}')

                if queue_raw:
                    ocr_parsed_text = '\n'.join(queue_raw)
                    roster_text = ocr_parsed_text

                os.unlink(pre)
            except Exception as e:
                import traceback
                print(f'[Faculty-OCR ERROR] {e}')
                traceback.print_exc()
            finally:
                try: os.unlink(ip)
                except: pass
    if not roster_text:
        return '请提供队列人员名单或上传照片', 400

    # 解析队列人员姓名
    queue_names = []
    for chunk in re.split(r'[、，,\n\s]+', roster_text):
        n = chunk.strip()
        if len(n) >= 2:
            queue_names.append(n)

    # 其他角色（手动输入优先，OCR 自动填充兜底）
    meta_people = {}
    role_labels = {1: '总负责', 2: '场控', 3: '后勤', 4: '摄影'}
    # 手动输入优先
    for i, label in role_labels.items():
        v = request.form.get(f'meta_{i}', '').strip()
        if v:
            names = [n.strip() for n in re.split(r'[、，,\n\s]+', v) if len(n.strip()) >= 2]
            if names:
                meta_people[label] = names
    # OCR 角色兜底（只补充手动输入为空的角色）
    for label in role_labels.values():
        if label not in meta_people and label in ocr_roles:
            meta_people[label] = ocr_roles[label]

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as t:
        t.write(excel_bytes); tmp = t.name

    try:
        wb = openpyxl.load_workbook(tmp)
        all_persons = parse_equipment_sheet(wb['装备分配'])
        name_map = {p['name']: p for p in all_persons}
        pool = build_pool(wb)
        # 为 boots/belt 提供空池（当前只管理 uniform/hat）
        for key in ['boots', 'belt']:
            if key not in pool:
                pool[key] = set()

        # 从库存中找到队列人员对应的装备信息
        faculty_persons = []
        missing = []
        for name in queue_names:
            p = name_map.get(name)
            if p:
                faculty_persons.append(dict(p))  # copy
            else:
                print(f'[Faculty] Unknown person: {name}')
                missing.append(name)

        # 保存初始装备状态（用于前端显示新旧对比）
        original_equip = {}
        for p in faculty_persons:
            for k in ['uniform', 'hat', 'boots', 'belt']:
                original_equip[f'{p["name"]}|{k}'] = p.get(k, '')

        # 冲突检测与解决
        conflicts = detect_faculty_conflicts(faculty_persons)
        reassigns, changed = resolve_faculty_conflicts(conflicts, faculty_persons, pool)
        b64 = generate_faculty_excel(faculty_persons, changed)

        # 按排序后的顺序返回预览数据
        sorted_persons = sort_people_by_uniform(faculty_persons)

        return jsonify(
            success=True,
            message=f'生成完成: {len(faculty_persons)} 人, {len(conflicts)} 个冲突, {len(reassigns)} 个重分配' + (f'（{len(missing)} 人未找到）' if missing else ''),
            conflicts_count=len(conflicts),
            reassignments_count=len(reassigns),
            sorted_persons=[dict(name=p['name'], gender=p['gender'],
                uniform=p.get('uniform',''), hat=p.get('hat',''),
                boots=p.get('boots',''), belt=p.get('belt','')) for p in sorted_persons],
            changes={k: v for k, v in changed.items()},
            original_equip=original_equip,
            missing=missing,
            conflicts=[dict(
                item_type={'uniform':'礼服','hat':'礼帽','boots':'马靴','belt':'腰带'}.get(c['item_type'], c['item_type']),
                item_code=c['item_code'],
                person_to_move=c['person_to_move'],
                person_to_keep=c['person_to_keep']) for c in conflicts],
            reassignments=[dict(person=r['person'],
                item_type={'uniform':'礼服','hat':'礼帽','boots':'马靴','belt':'腰带'}.get(r['item_type'], r['item_type']),
                old_item=r['old_item'], new_item=r['new_item']) for r in reassigns],
            excel_base64=b64,
            ocr_text=(ocr_parsed_text if image_file and image_file.filename else ''),
            meta_people=meta_people,
        )
    finally:
        os.unlink(tmp)


if __name__ == '__main__':
    from waitress import serve
    import os
    # Hugging Face Spaces requires port 7860
    port = int(os.environ.get('PORT', PORT))
    print(f'\n   礼服自动分配系统启动!')
    print(f'   http://0.0.0.0:{port}\n')
    serve(app, host='0.0.0.0', port=port, threads=8, channel_timeout=300)
