"""
直接生成！输入上月Excel + 排班文本 → 输出新Excel。
"""
import base64, io, os, re, tempfile
import openpyxl
from openpyxl.styles import PatternFill

YELLOW = 'FFFFFF00'
IN_EXCEL = r"c:\Users\夏瑞泽\xwechat_files\wxid_vnl456wm0r9h12_0c22\msg\file\2026-07\2025-2026学年夏五周到夏八周礼服分配(1)(1).xlsx"
OUT_EXCEL = r"c:\Users\夏瑞泽\Desktop\firct cc\uniform-assigner\本月礼服分配.xlsx"

# ====== 排班数据 ======
# 手动OCR图片，直接写入排班表
SCHEDULE = [
    ("周二", ["李诗诗","郭婷心","岳佳凝","方嫄","韩雅丽","吴昕桐","张鹏","夏瑞泽","胡思源","沈昊","孜英","顾铭文","陈涵予"]),
    ("周三", ["柯天翊","章芮容","余佳卉","马欣雅","吕凯进","姚爽","柳洋","章晏华","王梓丞","姚忻成","张子恒","张博宣","江文欣"]),
    ("周四", ["林珩","王雨梦","施东隅","艾克达","俞恩祺","潘睿","张嘉靖","肖玉茂","朱家兴","李承哲","韩钧宇","戴傲","韦景浩"]),
    ("周五", ["吴桐","段茗萱","张艺","许诺","楼震霆","叶宇轩","文科","李思齐","钱俊宇","陆荪桐","林育臣","陈思骆"]),
    ("周一", ["陈思骆","方佳瑶","呼岳洋","冯敬栩","叶栩浩","项烨","孙鸣","努尔加娜特","张煜琦","张子昂"]),
]
# 周一班 = 其他班抽调的人，他们只出现在周一

# ====== 解析 ======
wb = openpyxl.load_workbook(IN_EXCEL)
ws = wb['装备分配']
persons = []
for row in range(2, ws.max_row+1):
    name = ws.cell(row=row, column=2).value
    if not name: continue
    name = str(name).strip()
    uni  = str(ws.cell(row=row, column=3).value or '').strip()
    hat  = str(ws.cell(row=row, column=4).value or '').strip()
    boots= str(ws.cell(row=row, column=5).value or '').strip()
    belt = str(ws.cell(row=row, column=6).value or '').strip()
    note = str(ws.cell(row=row, column=7).value or '').strip()
    persons.append(dict(name=name, gender='M' if uni.startswith('M') else 'F', uniform=uni, hat=hat, boots=boots, belt=belt, note=note))
name_map = {p['name']:p for p in persons}
print(f"解析 {len(persons)} 人")

# ====== 共享关系 ======
rels, seen = [], set()
for p in persons:
    if not p['note']: continue
    for seg in re.split(r'[，；]', p['note']):
        seg = seg.strip()
        if not seg.startswith('和'): continue
        idx = seg.find('共用');
        if idx < 0: continue
        names = [n.strip() for n in re.split(r'[、，, ]+', seg[1:idx]) if n.strip()]
        after = seg[idx+2:]
        types = [t for t in [('uniform','礼服'),('hat','礼帽'),('boots','马靴'),('belt','腰带')] if t[1] in after]
        for n in names:
            for t,_ in types:
                key = f"{t}|{p[t]}|{','.join(sorted([p['name'],n]))}"
                if key not in seen:
                    seen.add(key)
                    rels.append(dict(item_code=p[t], item_type=t, shared_by=[p['name'],n]))

# 合并同item_code
m = {}
for r in rels:
    k = f"{r['item_type']}|{r['item_code']}"
    m.setdefault(k, set()).update(r['shared_by'])
mrels = [dict(item_code=k.split('|')[1], item_type=k.split('|')[0], shared_by=list(v)) for k,v in m.items()]
print(f"共享关系 {len(mrels)} 组")

# ====== 冲突检测 ======
conflicts = []
for day, people in SCHEDULE:
    daily = set(people)
    for r in mrels:
        overlap = [p for p in r['shared_by'] if p in daily]
        if len(overlap) >= 2:
            for i in range(1, len(overlap)):
                conflicts.append(dict(day=day, item_type=r['item_type'], item_code=r['item_code'], person_to_move=overlap[i], person_to_keep=overlap[0]))
print(f"冲突 {len(conflicts)} 个")
for c in conflicts:
    print(f"  {c['day']}: {c['person_to_move']} vs {c['person_to_keep']} ({c['item_type']}/{c['item_code']})")

# ====== 装备池 ======
pool = dict(uniform=set(), hat=set())
for p in persons:
    if p['uniform']: pool['uniform'].add(p['uniform'])
    if p['hat']: pool['hat'].add(p['hat'])
pool['uniform'].update({"M180/96-08","M180/96-09","M180/92-07","M180/92-08","M180/92-09","M180/92-10","M180/92-11","M180/92-12","F170/84-05","F170/88-03","F175/88-02","F175/92-02","M175/96-01","M180/92-01","M180/100-02","M180/100-03","M180/100-04","M180/100-05","M185/100-03","F165/84-04","F165/84-05","M175/88-02","M180/92-02","F165/80-01","F165/80-02"})
pool['hat'].update({"F5504","F5609","M5708","M5709","M5812","M5813","M5814","M5815","M5905","M6001","F5801","F5802","F5803","M5809"})
print(f"装备池 uniform={len(pool['uniform'])} hat={len(pool['hat'])}")

# ====== 重分配 ======
def parse_uniform(code):
    m = re.match(r'^([FM])(\d+)/(\d+)-(\d+)$', code)
    return dict(g=m.group(1),h=int(m.group(2)),c=int(m.group(3))) if m else None
def parse_hat(code):
    m = re.match(r'^([FM])(\d+)$', code)
    return dict(g=m.group(1),n=int(m.group(2))) if m else None
def find_alt(person, item_type, used):
    own = person[item_type]; gender = person['gender']
    av = [x for x in pool[item_type] if x not in used and x != own]
    cand = []
    for item in av:
        if item_type == 'uniform':
            info = parse_uniform(item)
            if not info or info['g'] != gender: continue
            pi = parse_uniform(own) or dict(h=0,c=0)
            cand.append((abs(info['h']-pi['h'])*100+abs(info['c']-pi['c']), item))
        else:
            info = parse_hat(item)
            if not info or info['g'] != gender: continue
            pi = parse_hat(own) or dict(n=0)
            cand.append((abs(info['n']-pi['n']), item))
    if not cand: cand = [(999999,x) for x in av if x.startswith(gender)]
    cand.sort(); return cand[0][1] if cand else None

reassigns, changed = [], {}
by_day = {}
for c in conflicts: by_day.setdefault(c['day'], []).append(c)
for day, clist in by_day.items():
    day_people = set(next(s[1] for s in SCHEDULE if s[0]==day))
    used_u = {p['uniform'] for p in persons if p['name'] in day_people}
    used_h = {p['hat'] for p in persons if p['name'] in day_people}
    for c in clist:
        person = name_map.get(c['person_to_move'])
        if not person: continue
        t = c['item_type']
        alt = find_alt(person, t, used_u if t=='uniform' else used_h)
        if alt:
            (used_u if t=='uniform' else used_h).discard(person[t])
            (used_u if t=='uniform' else used_h).add(alt)
            reassigns.append(dict(person=person['name'], item_type=t, old_item=person[t], new_item=alt))
            changed.setdefault(person['name'], {})[t] = alt
            person[t] = alt

print(f"\n重分配 {len(reassigns)} 次:")
for r in reassigns:
    print(f"  {r['person']}: {r['item_type']} {r['old_item']} → {r['new_item']}")

# ====== 更新备注 ======
for r in reassigns:
    p = name_map[r['person']]
    label = '礼服' if r['item_type']=='uniform' else '礼帽'
    sharers = [x['name'] for x in persons if x['name']!=p['name'] and x[r['item_type']]==r['new_item']]
    if sharers:
        p['note'] = '和' + '、'.join(sharers) + '共用' + label
        changed.setdefault(p['name'], {})['note'] = p['note']

# ====== 生成 Excel ======
ws = wb['装备分配']
for row in range(2, ws.max_row+1):
    name = ws.cell(row=row, column=2).value
    if not name: continue
    name = str(name).strip()
    p = name_map.get(name)
    if not p: continue
    ch = changed.get(name, {})
    for col,key in [(3,'uniform'),(4,'hat'),(5,'boots'),(6,'belt'),(7,'note')]:
        cell = ws.cell(row=row, column=col)
        cell.value = p.get(key, '') or None
        if key in ch:
            cell.fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type='solid')

wb.save(OUT_EXCEL)
print(f"\n✅ 已生成: {OUT_EXCEL}")
