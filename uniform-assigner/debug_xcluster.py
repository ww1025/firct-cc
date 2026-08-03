"""离线测试 X 聚类 OCR"""
import cv2, numpy as np, easyocr, re, sys
from PIL import Image
from difflib import SequenceMatcher

sys.path.insert(0, r'C:\Users\夏瑞泽\Desktop\firct cc\uniform-assigner')
from server import force_correct_name, fuzzy_match

image_path = r'c:\Users\夏瑞泽\xwechat_files\wxid_vnl456wm0r9h12_0c22\temp\RWTemp\2026-07\9e20f478899dc29eb19741386f9343c8\cfd5ca841af9d1bf15e098f16c5f3147.png'
img = Image.open(image_path).convert('RGB')
w, h = img.size
print(f'Image: {w}x{h}')

TARGET_W = min(w, 2000)
scale = TARGET_W / w
new_w, new_h = TARGET_W, int(h * scale)
img = img.resize((new_w, new_h), Image.LANCZOS)

reader = easyocr.Reader(['ch_sim'], gpu=False)
img_np = np.array(img)
gray = np.mean(img_np, axis=2).astype(np.uint8)

results = reader.readtext(img_np, detail=1, low_text=0.2, text_threshold=0.3)
print(f'Raw OCR: {len(results)} results')

# Filter edges
filtered = []
for bbox, text, conf in results:
    x1, y1 = bbox[0]; x3, y3 = bbox[2]
    t = text.strip()
    if not t: continue
    if x1 <= 5 and y1 <= 5: continue
    if x3 >= new_w - 5 and y3 >= new_h - 5: continue
    filtered.append((bbox, t, conf))
print(f'Filtered: {len(filtered)}')

blocks = []
for bbox, text, conf in filtered:
    x1, y1 = bbox[0]; x3, y3 = bbox[2]
    blocks.append({'text': text, 'x1': x1, 'x3': x3, 'cx': (x1+x3)/2, 'y1': y1, 'y3': y3})

# X clustering with different gap thresholds
for gap in [30, 40, 50, 60, 80]:
    blocks.sort(key=lambda b: b['cx'])
    x_clusters = []
    if blocks:
        x_clusters = [[blocks[0]]]
        for b in blocks[1:]:
            if b['cx'] - x_clusters[-1][-1]['cx'] < gap:
                x_clusters[-1].append(b)
            else:
                x_clusters.append([b])
    print(f'\nGap={gap}: {len(x_clusters)} clusters, sizes={[len(c) for c in x_clusters]}')
    for ci, cl in enumerate(x_clusters):
        xs = [b['x1'] for b in cl]
        print(f'  C{ci}: x={min(xs):.0f}-{max(xs):.0f}, n={len(cl)}')

# Test with actual known names
import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\夏瑞泽\Desktop\firct cc\uniform-assigner\本月礼服分配.xlsx')
known_names = []
for row in range(2, wb['装备分配'].max_row + 1):
    n = wb['装备分配'].cell(row=row, column=2).value
    if n and str(n).strip():
        known_names.append(str(n).strip())
wb.close()
print(f'\nKnown names: {len(known_names)}')

# Test force_correct_name on a few OCR outputs
tests = ['圊二', '匪_', '亏螈_', '楼霞霆', '夏毽', '张遽宣', '童曼垄', '柯丢衄']
for t in tests:
    c = force_correct_name(t, known_names)
    if not c:
        c = fuzzy_match(t, known_names)
    print(f'  "{t}" -> "{c}" (in known: {c in known_names})')
