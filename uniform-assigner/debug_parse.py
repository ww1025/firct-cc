"""Final debug: trace the actual OCR output on server"""
import cv2, numpy as np, easyocr, sys
from PIL import Image
sys.path.insert(0, r'C:\Users\夏瑞泽\Desktop\firct cc\uniform-assigner')
from server import force_correct_name, fuzzy_match, fuzzy_match_aggressive

image_path = r'c:\Users\夏瑞泽\xwechat_files\wxid_vnl456wm0r9h12_0c22\temp\RWTemp\2026-07\9e20f478899dc29eb19741386f9343c8\cfd5ca841af9d1bf15e098f16c5f3147.png'
img = Image.open(image_path).convert('RGB')
w, h = img.size
img_np = np.array(img)
gray = np.mean(img_np, axis=2).astype(np.uint8)

import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\夏瑞泽\Desktop\firct cc\uniform-assigner\本月礼服分配.xlsx')
known_names = [str(wb['装备分配'].cell(row=r, column=2).value).strip() for r in range(2, wb['装备分配'].max_row+1) if wb['装备分配'].cell(row=r, column=2).value]
wb.close()

DAY_LABELS = ['周一', '周二', '周三', '周四', '周五']

reader = easyocr.Reader(['ch_sim'], gpu=False)
results = reader.readtext(img_np, detail=1, low_text=0.2, text_threshold=0.3)

filtered = []
for bbox, text, conf in results:
    x1, y1 = bbox[0]; x3, y3 = bbox[2]
    t = text.strip()
    if not t: continue
    if x1 <= 5 and y1 <= 5: continue
    if x3 >= w - 5 and y3 >= h - 5: continue
    filtered.append((bbox, t, conf))

blocks = []
for bbox, text, conf in filtered:
    x1, y1 = bbox[0]; x3, y3 = bbox[2]
    blocks.append({'text': text, 'x1': x1, 'x3': x3, 'cx': (x1+x3)/2, 'y1': y1, 'y3': y3, 'conf': conf})

blocks.sort(key=lambda b: b['cx'])
x_clusters = [[blocks[0]]]
for b in blocks[1:]:
    if b['cx'] - x_clusters[-1][-1]['cx'] < 40:
        x_clusters[-1].append(b)
    else:
        x_clusters.append([b])

print(f'Clusters: {len(x_clusters)}')

day_schedule = {}
for ci, cluster in enumerate(x_clusters):
    day = DAY_LABELS[ci] if ci < 5 else ''
    cluster.sort(key=lambda b: b['y1'])
    seen = set()
    for b in cluster:
        t = b['text'].strip()
        if len(t) < 2: continue
        corrected = force_correct_name(t, known_names)
        if not corrected:
            corrected = fuzzy_match_aggressive(t, known_names)
        if corrected and corrected not in seen:
            seen.add(corrected)

    if seen:
        day_schedule.setdefault(day, []).extend(seen)
        print(f'{day}: {len(seen)} names')

# Now build the ocr_parsed_text - THIS IS THE KEY OUTPUT
lines = []
for day in DAY_LABELS:
    names = day_schedule.get(day, [])
    if names:
        lines.append(f'{day}：{"、".join(names)}')
ocr_parsed_text = '\n'.join(lines)
print(f'\nOCR output ({len(lines)} lines):')
print(ocr_parsed_text)

# Now test parse_schedule on this
sys.path.insert(0, r'C:\Users\夏瑞泽\Desktop\firct cc\uniform-assigner')
from server import parse_schedule
s = parse_schedule(ocr_parsed_text)
print(f'\nSchedule parsed: {len(s)} days')
for d in s:
    print(f'  {d["day"]}: {len(d["people"])}')
