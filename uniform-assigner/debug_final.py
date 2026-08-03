"""Final offline column OCR test - use original blocks approach instead of column re-OCR"""
import cv2, numpy as np, easyocr, sys
from PIL import Image
from difflib import SequenceMatcher

sys.path.insert(0, r'C:\Users\夏瑞泽\Desktop\firct cc\uniform-assigner')
from server import force_correct_name, fuzzy_match, fuzzy_match_aggressive

image_path = r'c:\Users\夏瑞泽\xwechat_files\wxid_vnl456wm0r9h12_0c22\temp\RWTemp\2026-07\9e20f478899dc29eb19741386f9343c8\cfd5ca841af9d1bf15e098f16c5f3147.png'
img = Image.open(image_path).convert('RGB')
w, h = img.size
new_w, new_h = w, h
img_np = np.array(img)
gray = np.mean(img_np, axis=2).astype(np.uint8)

reader = easyocr.Reader(['ch_sim'], gpu=False)

# Load known names
import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\夏瑞泽\Desktop\firct cc\uniform-assigner\本月礼服分配.xlsx')
known_names = []
for row in range(2, wb['装备分配'].max_row + 1):
    n = wb['装备分配'].cell(row=row, column=2).value
    if n and str(n).strip():
        known_names.append(str(n).strip())
wb.close()

DAY_LABELS = ['周一', '周二', '周三', '周四', '周五']

# DO NOT use column re-OCR. Use original full-image OCR blocks + X-clustering directly.
# The blocks from the first OCR pass already contain all the text.
print("=== Using original OCR blocks + X-clustering (no re-OCR) ===")

# Get blocks from full-image OCR
results = reader.readtext(img_np, detail=1, low_text=0.2, text_threshold=0.3)
filtered = []
for bbox, text, conf in results:
    x1, y1 = bbox[0]; x3, y3 = bbox[2]
    t = text.strip()
    if not t: continue
    if x1 <= 5 and y1 <= 5: continue
    if x3 >= new_w - 5 and y3 >= new_h - 5: continue
    filtered.append((bbox, t, conf))

blocks = []
for bbox, text, conf in filtered:
    x1, y1 = bbox[0]; x3, y3 = bbox[2]
    blocks.append({'text': text, 'x1': x1, 'x3': x3, 'cx': (x1+x3)/2, 'y1': y1, 'y3': y3, 'conf': conf})

# X clustering
blocks.sort(key=lambda b: b['cx'])
x_clusters = [[blocks[0]]]
for b in blocks[1:]:
    if b['cx'] - x_clusters[-1][-1]['cx'] < 40:
        x_clusters[-1].append(b)
    else:
        x_clusters.append([b])
print(f'X-clusters: {len(x_clusters)}')

day_schedule = {}
for ci, cluster in enumerate(x_clusters):
    cluster.sort(key=lambda b: b['y1'])
    day = DAY_LABELS[ci] if ci < 5 else f'Unknown{ci}'

    # Print raw OCR results BEFORE name correction
    raw_names = [b['text'] for b in cluster if b['y1'] > new_h * 0.1]  # skip header
    print(f'\n{day} ({len(cluster)} blocks): raw={raw_names[:5]}...')

    seen = set()
    for b in cluster:
        t = b['text'].strip()
        if len(t) < 2: continue
        corrected = force_correct_name(t, known_names)
        if not corrected:
            corrected = fuzzy_match_aggressive(t, known_names)
        if corrected and corrected not in seen:
            seen.add(corrected)
            print(f'  {t} ({b["conf"]:.4f}) -> {corrected}')

    if seen:
        day_schedule[day] = list(seen)

print(f'\n=== Final schedule ===')
for d in DAY_LABELS:
    names = day_schedule.get(d, [])
    print(f'{d}: {len(names)} people - {names}')
