"""
浙江大学国旗仪仗队 —— Streamlit Cloud 部署版
首页 + 院系礼服分配 + 物资仓库
"""
import streamlit as st
import base64, io, os, re, tempfile, math
from difflib import SequenceMatcher
import openpyxl
from openpyxl.styles import PatternFill
from PIL import Image
import numpy as np

st.set_page_config(page_title="浙江大学国旗仪仗队", page_icon="🇨🇳", layout="wide", initial_sidebar_state="collapsed")

# ── CSS 注入 ──
def inject_css():
    pass  # 所有 CSS 已合并到主 style 块中

# ── 国旗红样式 ──
FLAG_STYLE = """
<style>
.flag-header {
    background: #B81616; color: #fff; padding: 20px 32px;
    display: flex; align-items: center; justify-content: space-between;
    border-bottom: 2px solid rgba(245,197,24,.3);
    margin: -1rem -1rem 2rem -1rem;
}
.flag-header h1 {font-family: 'SimSun','KaiTi',serif; font-size: 18px; font-weight: 700; letter-spacing: .08em; color: #fff; margin: 0}
.flag-header .star {color: #F5C518; font-size: 16px}
.flag-footer {
    text-align: center; padding: 28px; color: #1a3a2a;
    font-family: 'SimSun','KaiTi',serif; font-size: 12px;
    letter-spacing: .06em; border-top: 1px solid rgba(184,22,22,.08);
    margin: 3rem 0 0 0;
}
.flag-footer .star {color: #F5C518; font-size: 10px; margin: 0 6px}
</style>
"""

# ── 业务逻辑 ──
YELLOW = 'FFFFFF00'

@st.cache_resource
def get_ocr_reader():
    import easyocr
    return easyocr.Reader(['ch_sim'], gpu=False)

def fuzzy_match(ocr_name, known_names):
    if not known_names: return ocr_name
    if ocr_name in known_names: return ocr_name
    cleaned = re.sub(r'[_|^~\s\d]+', '', ocr_name)
    if cleaned in known_names: return cleaned
    for ref in known_names:
        if len(ref) >= 2 and len(cleaned) >= 2 and cleaned[:2] == ref[:2]: return ref
    best, best_score = None, 0
    for ref in known_names:
        common = len(set(cleaned) & set(ref))
        ratio = SequenceMatcher(None, cleaned, ref).ratio()
        score = common * 3 + ratio * 5
        if score > best_score: best_score, best = score, ref
    return best if best and best_score >= 6 else re.sub(r'[_|^~\s\d]+', '', ocr_name)

def fuzzy_match_aggressive(ocr_name, known_names):
    if not ocr_name or not known_names: return None
    if ocr_name in known_names: return ocr_name
    cleaned = re.sub(r'[_|^~\s\d\s]+', '', ocr_name)
    if not cleaned: return None
    if cleaned in known_names: return cleaned
    candidates = [r for r in known_names if len(r) >= 2 and len(cleaned) >= 2 and cleaned[:2] == r[:2]]
    if candidates: return min(candidates, key=len)
    best, best_score = None, 0
    cc = set(cleaned)
    for ref in known_names:
        rc = set(ref); common = len(cc & rc)
        if common == 0: continue
        overlap = common / max(len(cc), len(rc))
        score = overlap * 10 - abs(len(ref) - len(cleaned)) * 1.5
        if score > best_score: best_score, best = score, ref
    return best if best and best_score >= 2.0 else None

def force_correct_name(ocr_text, known_names):
    if not ocr_text or len(ocr_text) < 2: return None
    if ocr_text in known_names: return ocr_text
    cleaned = re.sub(r'[_|^~\s\d]+', '', ocr_text)
    if cleaned in known_names: return cleaned
    candidates = [r for r in known_names if len(r) >= 2 and len(cleaned) >= 2 and cleaned[:2] == r[:2]]
    if candidates: return min(candidates, key=len)
    best, best_score = None, 0
    nc = set(cleaned)
    for ref in known_names:
        rc = set(ref); common = len(nc & rc)
        overlap = common / max(len(nc), len(rc), 1)
        seq = SequenceMatcher(None, cleaned, ref).ratio()
        score = overlap * 10 + seq * 5
        if score > best_score: best_score, best = score, ref
    return best if best and best_score >= 5 else None

def parse_equipment_sheet(ws):
    persons = []
    for row in range(2, ws.max_row+1):
        name = ws.cell(row=row, column=2).value
        if not name: continue
        name = str(name).strip()
        uni = str(ws.cell(row=row, column=3).value or '').strip()
        hat = str(ws.cell(row=row, column=4).value or '').strip()
        boots = str(ws.cell(row=row, column=5).value or '').strip()
        belt = str(ws.cell(row=row, column=6).value or '').strip()
        note = str(ws.cell(row=row, column=7).value or '').strip()
        persons.append(dict(name=name, gender='M' if uni.startswith('M') else 'F',
                            uniform=uni, hat=hat, boots=boots, belt=belt, note=note))
    return persons

def parse_shared_relations(persons):
    rels, seen = [], set()
    for p in persons:
        if not p['note']: continue
        for seg in re.split(r'[，；]', p['note']):
            seg = seg.strip()
            if not seg.startswith('和'): continue
            idx = seg.find('共用')
            if idx < 0: continue
            names = [n.strip() for n in re.split(r'[、，, ]+', seg[1:idx]) if n.strip()]
            after = seg[idx+2:]
            types = [t for t in [('uniform','礼服'),('hat','礼帽'),('boots','马靴'),('belt','腰带')] if t[1] in after]
            for n in names:
                for t, _ in types:
                    code = p[t]
                    key = f"{t}|{code}|{','.join(sorted([p['name'], n]))}"
                    if key not in seen:
                        seen.add(key)
                        rels.append(dict(item_code=code, item_type=t, shared_by=[p['name'], n]))
    return rels

def merge_relations(rels):
    m = {}
    for r in rels:
        k = f"{r['item_type']}|{r['item_code']}"
        m.setdefault(k, set()).update(r['shared_by'])
    return [dict(item_code=k.split('|')[1], item_type=k.split('|')[0], shared_by=list(v)) for k, v in m.items()]

def detect_conflicts(schedule, rels):
    conflicts = []
    for day in schedule:
        daily = set(day['people'])
        for r in rels:
            overlap = [p for p in r['shared_by'] if p in daily]
            if len(overlap) >= 2:
                for i in range(1, len(overlap)):
                    conflicts.append(dict(day=day['day'], item_type=r['item_type'], item_code=r['item_code'],
                                          person_to_move=overlap[i], person_to_keep=overlap[0]))
    return conflicts

def parse_uniform_code(code):
    m = re.match(r'^([FM])(\d+)/(\d+)-(\d+)$', code)
    return dict(gender=m.group(1), height=int(m.group(2)), chest=int(m.group(3))) if m else None

def parse_hat_code(code):
    m = re.match(r'^([FM])(\d{4})$', code)
    return dict(gender=m.group(1), number=int(m.group(2))) if m else None

def find_alternative(person, item_type, used_items, all_items):
    own = person[item_type]
    available = [x for x in all_items if x not in used_items and x != own]
    gender = person['gender']
    candidates = []
    for item in available:
        if item_type == 'uniform':
            info = parse_uniform_code(item)
            if not info or info['gender'] != gender: continue
            pi = parse_uniform_code(own) or dict(height=0, chest=0)
            score = abs(info['height'] - pi['height']) * 4 + abs(info['chest'] - pi['chest']) * 5
            candidates.append((score, item))
        elif item_type == 'hat':
            info = parse_hat_code(item)
            if not info or info['gender'] != gender: continue
            pi = parse_hat_code(own) or dict(number=0)
            candidates.append((abs(info['number'] - pi['number']), item))
        else:
            if item.startswith(gender): candidates.append((0, item))
    if not candidates:
        for item in available: candidates.append((999998, item))
    if not candidates:
        candidates = [(999999, x) for x in all_items if x.startswith(gender)]
    candidates.sort()
    return candidates[0][1] if candidates else None

def build_pool(wb):
    pool = dict(uniform=set(), hat=set())
    for row in range(2, wb['装备分配'].max_row+1):
        for col, key in [(3,'uniform'),(4,'hat')]:
            v = str(wb['装备分配'].cell(row=row, column=col).value or '').strip()
            if v: pool[key].add(v)
    for sn in ['礼服腰带摆放','礼帽摆放']:
        if sn in wb.sheetnames:
            ws = wb[sn]
            for row in range(3, ws.max_row+1):
                for col in range(1, ws.max_column+1):
                    v = str(ws.cell(row=row, column=col).value or '').strip()
                    if re.match(r'^[FM]\d{2,3}/', v): pool['uniform'].add(v)
                    if re.match(r'^[FM]\d{4}$', v): pool['hat'].add(v)
    return pool

def resolve(conflicts, persons, schedule, pool):
    reassigns, changed = [], {}
    by_day = {}
    for c in conflicts: by_day.setdefault(c['day'], []).append(c)
    for day, clist in by_day.items():
        people = next((set(s['people']) for s in schedule if s['day'] == day), set())
        used = {
            'uniform': {p['uniform'] for p in persons if p['name'] in people},
            'hat': {p['hat'] for p in persons if p['name'] in people}
        }
        for c in clist:
            person = next((p for p in persons if p['name'] == c['person_to_move']), None)
            if not person: continue
            t = c['item_type']
            alt = find_alternative(person, t, used[t], pool.get(t, set()))
            if alt:
                used[t].discard(person[t]); used[t].add(alt)
                reassigns.append(dict(person=person['name'], item_type=t, old_item=person[t], new_item=alt))
                changed.setdefault(person['name'], {})[t] = alt
                person[t] = alt
    return reassigns, changed

def generate_excel(template_path, persons, changed):
    wb = openpyxl.load_workbook(template_path)
    ws = wb['装备分配']
    name_map = {p['name']: p for p in persons}
    for row in range(2, ws.max_row+1):
        name = ws.cell(row=row, column=2).value
        if not name: continue
        name = str(name).strip()
        p = name_map.get(name)
        if not p: continue
        ch = changed.get(name, {})
        for col, key in [(3,'uniform'),(4,'hat'),(5,'boots'),(6,'belt'),(7,'note')]:
            cell = ws.cell(row=row, column=col)
            cell.value = p.get(key, '') or None
            if key in ch:
                cell.fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type='solid')
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return base64.b64encode(buf.read()).decode()

def parse_schedule(text):
    schedule = []
    day_patterns = [
        (r'周\s*一', '周一'), (r'周\s*二', '周二'),
        (r'周\s*三', '周三'), (r'周\s*四', '周四'), (r'周\s*五', '周五'),
    ]
    for line in text.strip().split('\n'):
        line = line.strip()
        if not line: continue
        day = None
        for pat, label in day_patterns:
            if re.search(pat, line): day = label; break
        if not day: continue
        rest = re.sub(r'周\s*[一二三四五]\s*[:：]?\s*', '', line)
        chunks = re.split(r'[、，,]', rest)
        names = []
        for chunk in chunks:
            n = re.sub(r'\s+', '', chunk)
            if len(n) >= 2: names.append(n)
        if names: schedule.append(dict(day=day, people=names))
    return schedule

# ── 院系相关 ──
def uniform_sort_key(person):
    suit = person.get('uniform', '')
    if not suit: return (2, 999, 999, 999)
    gender_code = 0 if suit.startswith('F') else 1
    nums = [int(n) for n in re.findall(r'\d+', suit)]
    while len(nums) < 3: nums.append(0)
    return (gender_code, nums[0], nums[1], nums[2])

def sort_people_by_uniform(people):
    return sorted(people, key=uniform_sort_key)

def detect_faculty_conflicts(persons):
    conflicts = []
    for eq_type in ['uniform', 'hat', 'boots', 'belt']:
        eq_map = {}
        for p in persons:
            code = p.get(eq_type, '')
            if not code: continue
            eq_map.setdefault(code, []).append(p['name'])
        for code, names in eq_map.items():
            if len(names) >= 2:
                for i in range(1, len(names)):
                    conflicts.append(dict(item_type=eq_type, item_code=code, person_to_move=names[i], person_to_keep=names[0]))
    return conflicts

def resolve_faculty_conflicts(conflicts, persons, pool):
    reassigns, changed = [], {}
    used = {}
    for eq_type in ['uniform', 'hat', 'boots', 'belt']:
        used[eq_type] = {p.get(eq_type, '') for p in persons if p.get(eq_type)}
    for c in conflicts:
        person = next((p for p in persons if p['name'] == c['person_to_move']), None)
        if not person: continue
        t = c['item_type']
        alt = find_alternative(person, t, used[t], pool.get(t, set()))
        if alt:
            used[t].discard(person[t]); used[t].add(alt)
            reassigns.append(dict(person=person['name'], item_type=t, old_item=person[t], new_item=alt))
            changed.setdefault(person['name'], {})[t] = alt
            person[t] = alt
    return reassigns, changed

FACULTY_TEMPLATE = os.path.join(os.path.dirname(__file__), '院系升旗装备分配表模板.xlsx')

def generate_faculty_excel(persons, changed):
    wb = openpyxl.load_workbook(FACULTY_TEMPLATE)
    yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type='solid')
    sorted_persons = sort_people_by_uniform(persons)
    ws1 = wb['装备分配']
    for row in range(2, ws1.max_row + 1):
        for col in range(1, 7): ws1.cell(row=row, column=col).value = None
    for i, p in enumerate(sorted_persons):
        row = i + 2
        name_changed = changed.get(p['name'], {})
        ws1.cell(row=row, column=1).value = p['name']
        ws1.cell(row=row, column=2).value = p.get('uniform', '')
        ws1.cell(row=row, column=3).value = p.get('hat', '')
        ws1.cell(row=row, column=4).value = p.get('boots', '')
        ws1.cell(row=row, column=5).value = p.get('belt', '')
        if name_changed:
            change_map = {'uniform': 2, 'hat': 3, 'boots': 4, 'belt': 5}
            for eq_type, col in change_map.items():
                if eq_type in name_changed:
                    ws1.cell(row=row, column=col).fill = yellow_fill
    for row in range(len(sorted_persons) + 2, ws1.max_row + 1):
        for col in range(1, 7): ws1.cell(row=row, column=col).value = None
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return base64.b64encode(buf.read()).decode()

# ── OCR ──
def ocr_faculty_roster(img_bytes, excel_bytes):
    """返回 (queue_text, ocr_roles)，与 server.py 完全一致"""
    from PIL import Image
    import numpy as np
    import math

    with tempfile.NamedTemporaryFile(suffix='.jpg', delete=False) as t:
        t.write(img_bytes); ip = t.name
    try:
        img = Image.open(ip).convert('RGB')
        w, h = img.size

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as et:
            et.write(excel_bytes); etmp = et.name
        try:
            wb_tmp = openpyxl.load_workbook(etmp)
            known_names = []
            for row in range(2, wb_tmp['装备分配'].max_row + 1):
                n = wb_tmp['装备分配'].cell(row=row, column=2).value
                if n and str(n).strip(): known_names.append(str(n).strip())
            wb_tmp.close()
        finally:
            os.unlink(etmp)

        TARGET_W = min(w, 2000)
        scale = TARGET_W / w
        new_w, new_h = TARGET_W, int(h * scale)
        img = img.resize((new_w, new_h), Image.LANCZOS)
        pre = ip + '_pre.png'; img.save(pre)

        reader = get_ocr_reader()
        results = reader.readtext(np.array(img), detail=1, low_text=0.2, text_threshold=0.4)

        filtered = []
        for bbox, text, conf in results:
            x1, y1 = bbox[0]; x3, y3 = bbox[2]
            t = text.strip()
            if not t: continue
            if x1 <= 5 and y1 <= 5: continue
            if x3 >= new_w - 5 and y3 >= new_h - 5: continue
            filtered.append((bbox, t, conf))
        results = filtered

        blocks = []
        for bbox, text, conf in results:
            x1, y1 = bbox[0]; x3, y3 = bbox[2]
            blocks.append({
                'text': text, 'x1': x1, 'x3': x3, 'cx': (x1 + x3) / 2,
                'y1': y1, 'y3': y3, 'cy': (y1 + y3) / 2, 'conf': conf
            })

        COL_W = new_w / 8
        blocks.sort(key=lambda b: b['cx'])
        col_buckets = {}
        for b in blocks:
            col_idx = int(b['cx'] // COL_W)
            col_buckets.setdefault(col_idx, []).append(b)
        sorted_cols = sorted(col_buckets.items())

        merged_cols = []
        for col_idx, col_blocks in sorted_cols:
            if merged_cols and col_idx - merged_cols[-1][0] <= 1:
                prev_blocks = merged_cols[-1][1]
                prev_max_x = max(b['x3'] for b in prev_blocks)
                cur_min_x = min(b['x1'] for b in col_blocks)
                if cur_min_x - prev_max_x < COL_W * 0.3:
                    merged_cols[-1] = (merged_cols[-1][0], prev_blocks + col_blocks)
                    continue
            merged_cols.append((col_idx, col_blocks))

        ROLE_KW = ['总负责', '场控', '后勤', '摄影']
        QUEUE_TITLE_KW = ['擎护旗', '队列']

        first_role_col_idx = None
        for i, (_, col_blocks) in enumerate(merged_cols):
            col_text = ' '.join(b['text'] for b in col_blocks)
            if any(kw in col_text for kw in ROLE_KW):
                first_role_col_idx = i; break

        if first_role_col_idx is not None:
            col_types = ['role' if i >= first_role_col_idx else 'queue' for i in range(len(merged_cols))]
        else:
            col_types = ['queue' if i < 4 else 'role' for i in range(len(merged_cols))]

        def _fuzzy(name, name_list):
            if not name or len(name) < 1: return name
            if name in name_list: return name
            nc = set(name); best, best_score = name, 0
            for ref in name_list:
                rc = set(ref); common = len(nc & rc)
                overlap = common / max(len(nc), len(rc))
                seq = SequenceMatcher(None, name, ref).ratio()
                score = overlap * 10 + seq * 5
                if score > best_score: best_score, best = score, ref
            return best if best_score >= 5 and best != name else name

        def _clean_name_block(blocks_in_col):
            blocks_in_col.sort(key=lambda b: b['y1'])
            merged = []
            i = 0
            while i < len(blocks_in_col):
                b = blocks_in_col[i]; t = b['text'].strip()
                if len(t) == 1 and i + 1 < len(blocks_in_col):
                    nb = blocks_in_col[i + 1]
                    char_h = b['y3'] - b['y1']
                    if char_h > 0 and nb['y1'] - b['y3'] < char_h * 3.0:
                        t = t + nb['text'].strip(); i += 1
                merged.append(t); i += 1
            cleaned = []
            for t in merged:
                if not t: continue
                if t in QUEUE_TITLE_KW: continue
                if t in ROLE_KW: continue
                for kw in QUEUE_TITLE_KW + ROLE_KW:
                    if kw in t and t != kw: t = t.replace(kw, '').strip()
                if t: cleaned.append(t)
            return cleaned

        queue_raw = []
        ocr_roles = {}

        all_role_blocks = []
        for idx, (_, col_blocks) in enumerate(merged_cols):
            if col_types[idx] == 'queue':
                for t in _clean_name_block(col_blocks):
                    corrected = _fuzzy(t, known_names)
                    if corrected not in queue_raw:
                        queue_raw.append(corrected)
            else:
                all_role_blocks.extend(col_blocks)

        # 角色列按 y 分段匹配角色关键词
        if all_role_blocks:
            all_role_blocks.sort(key=lambda b: b['y1'])
            role_markers = []
            name_blocks = []
            for b in all_role_blocks:
                t = b['text'].strip()
                if t in ROLE_KW:
                    role_markers.append((b['y1'], t))
                else:
                    name_blocks.append(b)
            if role_markers:
                role_markers.sort(key=lambda x: x[0])
                for _, rk in role_markers:
                    ocr_roles.setdefault(rk, [])
                for nb in name_blocks:
                    best_role, best_dy = None, float('inf')
                    for ry, rk in role_markers:
                        dy = abs(nb['y1'] - ry)
                        if dy < best_dy:
                            best_dy, best_role = dy, rk
                    if best_role:
                        for ct in _clean_name_block([nb]):
                            corrected = _fuzzy(ct, known_names)
                            if corrected not in ocr_roles[best_role]:
                                ocr_roles[best_role].append(corrected)

        # 从队列中移除角色人员
        role_people = set()
        for names in ocr_roles.values():
            role_people.update(names)
        queue_raw = [n for n in queue_raw if n not in role_people]

        # 全图碎片兜底
        all_text = ''.join(b['text'] for b in blocks)
        all_chars = set(all_text)
        for ref_name in known_names:
            if ref_name in queue_raw or ref_name in role_people: continue
            rc = set(ref_name)
            hit = all_chars & rc
            if len(hit) >= max(2, len(rc) * 0.5):
                for b in blocks:
                    matched = _fuzzy(b['text'], known_names)
                    if matched == ref_name:
                        if matched not in queue_raw and matched not in role_people:
                            col_i = int(b['cx'] // COL_W)
                            is_role = False
                            if first_role_col_idx is not None:
                                for mi, (mci, _) in enumerate(merged_cols):
                                    if col_i == mci and col_types[mi] == 'role':
                                        is_role = True; break
                            if is_role:
                                if ref_name not in role_people:
                                    queue_raw.append(ref_name)
                            else:
                                if ref_name not in queue_raw:
                                    queue_raw.append(ref_name)
                        break

        os.unlink(pre)
        return ('\n'.join(queue_raw) if queue_raw else '', ocr_roles)
    finally:
        try: os.unlink(ip)
        except: pass


# ═══════════════════════════════════════
#  页面渲染
# ═══════════════════════════════════════

@st.cache_data(ttl=86400)
def warehouse_data_json():
    """返回物资仓库的裸数据，供 JS 渲染用"""
    import json
    data = {
        "grid": [
            [{"code": "X44-05", "person": "孜英"}, {"code": "X44-04", "person": ""}, {"code": "X43-08", "person": "顾铭文"}, {"code": "X44-06", "person": "林育臣"}, {"code": "X39-02", "person": ""}, {"code": "N43/41-01", "person": "钱俊宇"}, {"code": "N43/41-02", "person": ""}, {"code": "N43/43-01", "person": "章晏华"}, {"code": "N43/43-02", "person": "张博宣"}, {"code": "N43/43-03", "person": "李承哲"}, {"code": "N43/46", "person": ""}, {"code": "N44/44.5", "person": "沈昊"}, {"code": "X42-22", "person": "肖玉茂"}, {"code": "X42-23", "person": "张子昂"}],
            [{"code": "X43-04", "person": "韩钧宇"}, {"code": "X43-05", "person": ""}, {"code": "X43-06", "person": "朱家兴"}, {"code": "X43-07", "person": "柳洋"}, {"code": "X39-01", "person": "楼震霆"}, {"code": "N42/42.5-02", "person": "韦景浩"}, {"code": "N42/42.5-03", "person": "冯敬栩"}, {"code": "N42/42.5-04", "person": "潘睿"}, {"code": "N42/43.5-02", "person": ""}, {"code": "N42/44", "person": ""}, {"code": "N42/44.5", "person": "叶非"}, {"code": "N45/45", "person": "戴傲"}, {"code": "X42-19", "person": "（鞋面损坏）"}, {"code": "X42-21", "person": ""}],
            [{"label": "置物架"}, None, None, {"code": "X43-03", "person": ""}, {"code": "X38-04", "person": "郭婷心"}, {"code": "N42/40.5-02", "person": ""}, {"code": "N42/40.5-03", "person": "王梓丞"}, {"code": "N42/40.5-04", "person": "陆荪桐"}, {"code": "N42/43-01", "person": ""}, {"code": "N42/43-02", "person": ""}, {"code": "N42/43.5-01", "person": ""}, {"label": "置物架"}, None, None],
            [None, None, None, {"code": "X43-02", "person": ""}, {"code": "X38-03", "person": "卢格妤"}, {"code": "N41/42-02", "person": "姚忻成"}, {"code": "N41/42-03", "person": "胡思源"}, {"code": "N41/42-04", "person": "徐顺烨"}, {"code": "N42/42-01", "person": "张煜琦"}, {"code": "N42/42-02", "person": "夏瑞泽"}, {"code": "N42/42.5-01", "person": "张子恒"}, None, None, None],
            [None, None, None, {"code": "X43-01", "person": ""}, {"code": "X38-02", "person": ""}, {"code": "N39/40.5", "person": "韩雅丽"}, {"code": "N40/39.5-01", "person": "孙鸣"}, {"code": "N40/39.5-02", "person": "余佳卉"}, {"code": "N41/42-05", "person": "叶宇轩"}, {"code": "N41/43.5", "person": ""}, {"code": "N42/40.5-01", "person": ""}, None, None, None],
            [None, None, None, {"code": "X42-24", "person": ""}, {"code": "J38-01", "person": ""}, {"code": "N38/40-01", "person": "许诺"}, {"code": "N38/40-02", "person": ""}, {"code": "N38/40-03", "person": "段茗萱"}, {"code": "N40/41", "person": "吕凯进"}, {"code": "N40/41.5", "person": "程锦添"}, {"code": "N41/42-01", "person": "（舟山）"}, None, None, None],
            [None, None, None, None, None, None, None, None, {"code": "N38/41", "person": "江文欣"}, {"code": "N39/38-01", "person": "吴桐"}, {"code": "N39/38-02", "person": "柯天翊"}, {"code": "X42-12", "person": ""}, {"code": "X42-13", "person": ""}, {"code": "X42-17", "person": ""}],
            [{"label": "置物架"}, None, None, None, {"code": "X37-05", "person": "王雨梦"}, {"code": "N37/40", "person": "俞恩祺"}, {"code": "N37/41", "person": "艾克达"}, {"code": "N38/37", "person": "王艺霏"}, {"code": "N38/37.5-01", "person": "张艺"}, {"code": "N38/37.5-02", "person": "章芮容"}, {"code": "N38/37.5-03", "person": "施东隅"}, {"code": "X42-06", "person": "李思齐"}, {"code": "X42-07", "person": "姚爽"}, {"code": "X42-09", "person": ""}],
            [None, None, None, {"code": "J41-03", "person": ""}, {"code": "X37-04", "person": "呼岳洋"}, {"code": "N37/37-04", "person": ""}, {"code": "N37/37.5-01", "person": "林珩"}, {"code": "N37/37.5-02", "person": ""}, {"code": "N37/38-01", "person": "陈涵予"}, {"code": "N37/38-02", "person": "祁子谦"}, {"code": "N37/39.5", "person": "杨阿丽雅"}, {"code": "X42-01", "person": "李秉均"}, {"code": "X42-03", "person": "张嘉靖"}, {"code": "X42-05", "person": "努尔加娜特"}],
            [None, None, {"code": "X45-01", "person": ""}, {"code": "J41-01", "person": "叶栩浩"}, {"code": "X36-01", "person": "纪博雅"}, {"code": "N36/38-02", "person": "马欣雅"}, {"code": "N36/39", "person": "李泽一"}, {"code": "N37/36.5", "person": ""}, {"code": "N37/37-01", "person": ""}, {"code": "N37/37-02", "person": "李诗诗"}, {"code": "N37/37-03", "person": ""}, {"code": "X41-01", "person": "项烨"}, {"code": "X41-02", "person": "吴昕桐"}, {"code": "X41-03", "person": "陈思骆"}],
            [None, None, {"code": "X39-03", "person": "张鹏"}, {"code": "X39-04", "person": ""}, {"code": "J36-02", "person": ""}, {"code": "N35/36-01", "person": "方嫄 董欢瑶"}, {"code": "N35/37", "person": "初姝彤"}, {"code": "N36/36-01", "person": "岳佳凝"}, {"code": "N36/36-02", "person": "方佳瑶"}, {"code": "N36/37", "person": "杨羽茜"}, {"code": "N36/38-01", "person": "邬家琪"}, {"code": "X39-06", "person": ""}, {"code": "X39-07", "person": "文科"}, {"code": "X39-09", "person": ""}]
        ],
        "belts": [
            {"code": "F01", "uniform_size": "F165/80-01", "person": "李泽一", "cabinet": "一柜"},
            {"code": "F02", "uniform_size": "F165/80-02", "person": "", "cabinet": "一柜"},
            {"code": "F03", "uniform_size": "F165/84-01", "person": "林珩", "cabinet": "一柜"},
            {"code": "F04", "uniform_size": "F165/84-02", "person": "方佳瑶", "cabinet": "一柜"},
            {"code": "F05", "uniform_size": "F165/84-03", "person": "纪博雅", "cabinet": "一柜"},
            {"code": "F06", "uniform_size": "F170/84-06", "person": "韩雅丽", "cabinet": "四柜"},
            {"code": "F07", "uniform_size": "F165/88-01", "person": "李诗诗、王雨梦", "cabinet": "二柜"},
            {"code": "F08", "uniform_size": "F165/88-02", "person": "郭婷心", "cabinet": "二柜"},
            {"code": "F09", "uniform_size": "F165/88-03", "person": "岳佳凝、吴桐", "cabinet": "二柜"},
            {"code": "F10", "uniform_size": "F165/88-04", "person": "江文欣、董欢瑶", "cabinet": "二柜"},
            {"code": "F11", "uniform_size": "F170/84-01", "person": "方嫄", "cabinet": "四柜"},
            {"code": "F12", "uniform_size": "F170/84-02", "person": "柯天翊", "cabinet": "四柜"},
            {"code": "F13", "uniform_size": "F170/84-03", "person": "祁子谦、段茗萱", "cabinet": "四柜"},
            {"code": "F14", "uniform_size": "F170/84-04", "person": "张艺", "cabinet": "四柜"},
            {"code": "F15", "uniform_size": "F170/88-01", "person": "施东隅", "cabinet": "四柜"},
            {"code": "F16", "uniform_size": "F170/88-02", "person": "呼岳洋、章芮容", "cabinet": "四柜"},
            {"code": "F17", "uniform_size": "F170/92-01", "person": "艾克达、卢格妤", "cabinet": "五柜"},
            {"code": "F18", "uniform_size": "F170/92-02", "person": "许诺、余佳卉", "cabinet": "五柜"},
            {"code": "F19", "uniform_size": "F175/88-01", "person": "马欣雅", "cabinet": "二柜"},
            {"code": "F20", "uniform_size": "F175/92-01", "person": "俞恩祺、陈涵予", "cabinet": "三柜"},
            {"code": "M01", "uniform_size": "M175/88-01", "person": "陈思骆", "cabinet": "一柜"},
            {"code": "M02", "uniform_size": "M175/88-02", "person": "吴昕桐", "cabinet": "一柜"},
            {"code": "M03", "uniform_size": "M175/92-01", "person": "楼震霆、潘睿", "cabinet": "一柜"},
            {"code": "M04", "uniform_size": "M175/96-01", "person": "张鹏", "cabinet": "三柜"},
            {"code": "M05", "uniform_size": "M175/96-02", "person": "叶宇轩", "cabinet": "三柜"},
            {"code": "M06", "uniform_size": "M175/96-03", "person": "文科", "cabinet": "三柜"},
            {"code": "M07", "uniform_size": "M175/96-04", "person": "冯敬栩", "cabinet": "三柜"},
            {"code": "M08", "uniform_size": "M180/92-01", "person": "张嘉靖", "cabinet": "三柜"},
            {"code": "M09", "uniform_size": "M180/92-02", "person": "肖玉茂", "cabinet": "三柜"},
            {"code": "M10", "uniform_size": "M180/92-03", "person": "夏瑞泽", "cabinet": "三柜"},
            {"code": "M11", "uniform_size": "M180/92-04", "person": "胡思源", "cabinet": "三柜"},
            {"code": "M12", "uniform_size": "M180/92-05", "person": "李思齐", "cabinet": "二柜"},
            {"code": "M13", "uniform_size": "M180/92-06", "person": "叶栩浩", "cabinet": "二柜"},
            {"code": "M14", "uniform_size": "M185/100-03", "person": "", "cabinet": "六柜"},
            {"code": "M15", "uniform_size": "M180/96-05", "person": "朱家兴、姚忻成", "cabinet": "五柜"},
            {"code": "M16", "uniform_size": "M180/96-01", "person": "项烨、姚爽", "cabinet": "五柜"},
            {"code": "M17", "uniform_size": "M180/96-02", "person": "柳洋、孙鸣", "cabinet": "五柜"},
            {"code": "M18", "uniform_size": "M180/96-03", "person": "章晏华", "cabinet": "五柜"},
            {"code": "M19", "uniform_size": "M180/96-04", "person": "沈昊、王梓丞", "cabinet": "五柜"},
            {"code": "M20", "uniform_size": "M175/92-02", "person": "吕凯进", "cabinet": "一柜"},
            {"code": "M21", "uniform_size": "M180/96-06", "person": "钱俊宇、努尔加娜特、张子恒", "cabinet": "四柜"},
            {"code": "M22", "uniform_size": "M180/96-07", "person": "张煜琦、孜英", "cabinet": "四柜"},
            {"code": "M23", "uniform_size": "M180/100-01", "person": "张子昂", "cabinet": "五柜"},
            {"code": "M24", "uniform_size": "M185/100-01", "person": "张博宣、戴傲", "cabinet": "六柜"},
            {"code": "M25", "uniform_size": "M185/100-02", "person": "顾铭文、韦景浩、林育臣", "cabinet": "六柜"},
            {"code": "M26", "uniform_size": "M185/96-01", "person": "李承哲", "cabinet": "六柜"},
            {"code": "M27", "uniform_size": "M185/96-02", "person": "韩钧宇、陆荪桐", "cabinet": "六柜"}
        ],
        "uniforms": [
            {"code": "F165/80-01", "belt": "F01", "person": "李泽一", "cabinet": "一柜"},
            {"code": "F165/80-02", "belt": "F02", "person": "", "cabinet": "一柜"},
            {"code": "F165/84-01", "belt": "F03", "person": "林珩", "cabinet": "一柜"},
            {"code": "F165/84-02", "belt": "F04", "person": "方佳瑶", "cabinet": "一柜"},
            {"code": "F165/84-03", "belt": "F05", "person": "纪博雅", "cabinet": "一柜"},
            {"code": "F165/84-04", "belt": "", "person": "", "cabinet": "一柜"},
            {"code": "F165/84-05", "belt": "", "person": "", "cabinet": "一柜"},
            {"code": "F165/88-01", "belt": "F07", "person": "李诗诗、王雨梦", "cabinet": "二柜"},
            {"code": "F165/88-02", "belt": "F08", "person": "郭婷心", "cabinet": "二柜"},
            {"code": "F165/88-03", "belt": "F09", "person": "岳佳凝、吴桐", "cabinet": "二柜"},
            {"code": "F165/88-04", "belt": "F10", "person": "江文欣、董欢瑶", "cabinet": "二柜"},
            {"code": "F170/84-01", "belt": "F11", "person": "方嫄", "cabinet": "四柜"},
            {"code": "F170/84-02", "belt": "F12", "person": "柯天翊", "cabinet": "四柜"},
            {"code": "F170/84-03", "belt": "F13", "person": "祁子谦、段茗萱", "cabinet": "四柜"},
            {"code": "F170/84-04", "belt": "F14", "person": "张艺", "cabinet": "四柜"},
            {"code": "F170/84-05", "belt": "", "person": "", "cabinet": "四柜"},
            {"code": "F170/84-06", "belt": "F06", "person": "韩雅丽", "cabinet": "四柜"},
            {"code": "F170/88-01", "belt": "F15", "person": "施东隅", "cabinet": "四柜"},
            {"code": "F170/88-02", "belt": "F16", "person": "呼岳洋、章芮容", "cabinet": "四柜"},
            {"code": "F170/88-03", "belt": "", "person": "", "cabinet": "四柜"},
            {"code": "F170/92-01", "belt": "F17", "person": "艾克达、卢格妤", "cabinet": "五柜"},
            {"code": "F170/92-02", "belt": "F18", "person": "许诺、余佳卉", "cabinet": "五柜"},
            {"code": "F175/88-01", "belt": "F19", "person": "马欣雅", "cabinet": "二柜"},
            {"code": "F175/88-02", "belt": "", "person": "", "cabinet": "三柜"},
            {"code": "F175/92-01", "belt": "F20", "person": "俞恩祺、陈涵予", "cabinet": "三柜"},
            {"code": "F175/92-02", "belt": "", "person": "", "cabinet": "三柜"},
            {"code": "M175/88-01", "belt": "M01", "person": "陈思骆", "cabinet": "一柜"},
            {"code": "M175/88-02", "belt": "M02", "person": "吴昕桐", "cabinet": "一柜"},
            {"code": "M175/92-01", "belt": "M03", "person": "楼震霆、潘睿", "cabinet": "一柜"},
            {"code": "M175/92-02", "belt": "M20", "person": "吕凯进", "cabinet": "一柜"},
            {"code": "M175/96-01", "belt": "M04", "person": "张鹏", "cabinet": "三柜"},
            {"code": "M175/96-02", "belt": "M05", "person": "叶宇轩", "cabinet": "三柜"},
            {"code": "M175/96-03", "belt": "M06", "person": "文科", "cabinet": "三柜"},
            {"code": "M175/96-04", "belt": "M07", "person": "冯敬栩", "cabinet": "三柜"},
            {"code": "M180/92-05", "belt": "M12", "person": "李思齐", "cabinet": "二柜"},
            {"code": "M180/92-06", "belt": "M13", "person": "叶栩浩", "cabinet": "二柜"},
            {"code": "M180/92-07", "belt": "", "person": "", "cabinet": "二柜"},
            {"code": "M180/92-08", "belt": "", "person": "", "cabinet": "二柜"},
            {"code": "M180/92-10", "belt": "", "person": "", "cabinet": "二柜"},
            {"code": "M180/92-11", "belt": "", "person": "", "cabinet": "二柜"},
            {"code": "M180/92-01", "belt": "M08", "person": "张嘉靖", "cabinet": "三柜"},
            {"code": "M180/92-02", "belt": "M09", "person": "肖玉茂", "cabinet": "三柜"},
            {"code": "M180/92-03", "belt": "M10", "person": "夏瑞泽", "cabinet": "三柜"},
            {"code": "M180/92-04", "belt": "M11", "person": "胡思源", "cabinet": "三柜"},
            {"code": "M180/92-09", "belt": "", "person": "", "cabinet": "五柜"},
            {"code": "M180/92-12", "belt": "", "person": "", "cabinet": "五柜"},
            {"code": "M180/96-06", "belt": "M21", "person": "钱俊宇、努尔加娜特、张子恒", "cabinet": "四柜"},
            {"code": "M180/96-07", "belt": "M22", "person": "张煜琦、孜英", "cabinet": "四柜"},
            {"code": "M180/96-01", "belt": "M16", "person": "项烨、姚爽", "cabinet": "五柜"},
            {"code": "M180/96-02", "belt": "M17", "person": "柳洋、孙鸣", "cabinet": "五柜"},
            {"code": "M180/96-03", "belt": "M18", "person": "章晏华", "cabinet": "五柜"},
            {"code": "M180/96-04", "belt": "M19", "person": "沈昊、王梓丞", "cabinet": "五柜"},
            {"code": "M180/96-05", "belt": "M15", "person": "朱家兴、姚忻成", "cabinet": "五柜"},
            {"code": "M180/96-08", "belt": "", "person": "", "cabinet": "六柜"},
            {"code": "M180/96-09", "belt": "", "person": "", "cabinet": "六柜"},
            {"code": "M180/100-01", "belt": "M23", "person": "张子昂", "cabinet": "五柜"},
            {"code": "M180/100-02", "belt": "", "person": "", "cabinet": "六柜"},
            {"code": "M180/100-03", "belt": "", "person": "", "cabinet": "六柜"},
            {"code": "M180/100-04", "belt": "", "person": "", "cabinet": "六柜"},
            {"code": "M180/100-05", "belt": "", "person": "", "cabinet": "六柜"},
            {"code": "M185/96-01", "belt": "M26", "person": "李承哲", "cabinet": "六柜"},
            {"code": "M185/96-02", "belt": "M27", "person": "韩钧宇、陆荪桐", "cabinet": "六柜"},
            {"code": "M185/100-01", "belt": "M24", "person": "张博宣、戴傲", "cabinet": "六柜"},
            {"code": "M185/100-02", "belt": "M25", "person": "顾铭文、韦景浩、林育臣", "cabinet": "六柜"},
            {"code": "M185/100-03", "belt": "M14", "person": "", "cabinet": "六柜"}
        ]
    }
    return json.dumps(data, ensure_ascii=False)


# ── HTML 渲染辅助函数（匹配 Flask 版表格样式）──
def make_person_rows(sorted_people, changed):
    rows = []
    for i, p in enumerate(sorted_people):
        ch = changed.get(p['name'], {})
        g = '女' if p.get('gender')=='F' else '男'
        items = ['uniform','hat','boots','belt']
        cells = []
        for k in items:
            v = p.get(k, '')
            cls = ' class="hl"' if ch.get(k) else ''
            cells.append(f'<td{cls}>{v}</td>')
        rows.append(f'<tr><td>{i+1}</td><td><b>{p["name"]}</b></td><td>{g}</td>{"".join(cells)}</tr>')
    return '\n'.join(rows)


def _wh_escape(s):
    """HTML-escape for inline onclick attributes"""
    return s.replace('&', '&amp;').replace("'", "&#39;").replace('"', '&quot;').replace('<', '&lt;').replace('>', '&gt;')

def _wh_fmt_name(nm):
    """Short name helper — match JS fmtName"""
    if not nm:
        return ''
    parts = [x.strip() for x in nm.replace('/', ',').replace('，', ',').replace('、', ',').split(',') if x.strip()]
    return f'{len(parts)}人' if len(parts) > 1 else parts[0]

SHIRT_SET = {
    'F165/80-01':1,'F165/80-02':1,'F165/84-02':1,'F165/88-01':1,'F165/88-02':1,
    'F165/88-03':1,'F165/88-04':1,'M180/92-05':1,'M175/96-02':1,'F170/84-01':1,
    'F170/84-02':1,'F170/84-03':1,'F170/84-06':1,'F170/88-01':1,'F170/88-02':1,
    'F170/92-01':1,
}

def _wh_build_boot_cell(cell):
    """Generate one boot cell HTML — data attributes only, JS event delegation handles click"""
    code = _wh_escape(str(cell.get('code', '')))
    person = str(cell.get('person', ''))
    nm = _wh_fmt_name(person)
    psn = _wh_escape(person)
    inner = f'<span class="cc">{code}</span>'
    if nm:
        inner += f'<span class="cn">{_wh_escape(nm)}</span>'
    return f'<div class="cell boot" data-code="{code}" data-person="{psn}" data-type="boot">{inner}</div>'

def _wh_render_boots(data):
    """Render the boots (grid) section"""
    grid = data.get('grid', [])
    flat = [c for r in grid for c in r if c and c.get('code')]
    cg = [{'n':1,'c':4},{'n':2,'c':4},{'n':3,'c':3},{'n':4,'c':3}]
    rg = [{'n':'A','r':2},{'n':'B','r':4},{'n':'C','r':5}]

    h = '<details class="section" open><summary class="section-header"><span class="tag" style="background:#E65100">🥾</span>马靴 · ' + str(len(flat)) + '库位 ▼</summary>'
    h += '<div class="table-wrap"><div class="grid-table" style="grid-template-columns:28px repeat(14,minmax(64px,1fr))">'
    h += '<div class="cell gh" style="background:#fafafa"></div>'
    for g in cg:
        h += f'<div class="cell gh" style="grid-column:span {g["c"]};background:#FFE0B2;color:#BF360C;font-weight:700">{g["n"]}</div>'

    ri = 0
    for rg2 in rg:
        rows = grid[ri:ri+rg2['r']]
        for lri, row in enumerate(rows):
            if lri == 0:
                h += f'<div class="cell gh" style="grid-row:span {rg2["r"]};writing-mode:vertical-lr;background:#FFE0B2;color:#BF360C;font-weight:700">{rg2["n"]}</div>'
            for ci in range(14):
                cell = row[ci] if ci < len(row) else None
                if cell and cell.get('code'):
                    h += _wh_build_boot_cell(cell)
                else:
                    h += '<div class="cell empty"></div>'
        ri += rg2['r']
    h += '</div></div></details>'
    return h

def _wh_belt_cell(item):
    """Generate one belt cell HTML — data attributes only, JS event delegation handles click"""
    code = _wh_escape(str(item.get('code', '')))
    person = str(item.get('person', ''))
    nm = _wh_fmt_name(person)
    cabinet = _wh_escape(str(item.get('cabinet', '')))
    uniform_size = _wh_escape(str(item.get('uniform_size', '')))
    gender = '女款' if code.startswith('F') else '男款'
    psn = _wh_escape(person)
    inner = f'<span class="cc">{code}</span>'
    if nm:
        inner += f'<span class="cn">{_wh_escape(nm)}</span>'
    return f'<div class="cell belt" data-code="{code}" data-person="{psn}" data-type="belt" data-gender="{gender}" data-cabinet="{cabinet}" data-uniform="{uniform_size}">{inner}</div>'

def _wh_uniform_cell(item):
    """Generate one uniform cell HTML — data attributes only, JS event delegation handles click"""
    code = _wh_escape(str(item.get('code', '')))
    person = str(item.get('person', ''))
    nm = _wh_fmt_name(person)
    cabinet = _wh_escape(str(item.get('cabinet', '')))
    belt = _wh_escape(str(item.get('belt', '')))
    has_shirt = '1' if code in SHIRT_SET else '0'
    psn = _wh_escape(person)
    inner = f'<span class="cc">{code}</span>'
    if nm:
        inner += f'<span class="cn">{_wh_escape(nm)}</span>'
    if has_shirt == '1':
        inner += '<span class="ct" style="background:#4CAF50">衬衫</span>'
    return f'<div class="cell uniform" data-code="{code}" data-person="{psn}" data-type="uniform" data-cabinet="{cabinet}" data-belt="{belt}" data-has-shirt="{has_shirt}">{inner}</div>'

def _wh_render_section(items, item_type, title, color):
    """Render a belts or uniforms section with cabinet grouping"""
    if not items:
        return ''
    cabs = ['一柜','二柜','三柜','四柜','五柜','六柜']
    ccl = ['#E65100','#FF8F00','#F9A825','#FFB300','#FFC107','#FFCA28']
    cell_fn = _wh_belt_cell if item_type == 'belt' else _wh_uniform_cell

    h = f'<details class="section"><summary class="section-header"><span class="tag" style="background:{color}">{title[:2]}</span>{title} · {len(items)}项 ▼</summary><div class="table-wrap">'
    for ci, cab in enumerate(cabs):
        grp = [it for it in items if it.get('cabinet') == cab]
        if not grp:
            continue
        h += f'<div class="cab-label" style="color:{ccl[ci]};border-color:{ccl[ci]}">{cab} · {len(grp)}项</div>'
        h += '<div class="grid-table" style="grid-template-columns:repeat(auto-fill,minmax(80px,1fr))">'
        for item in grp:
            h += cell_fn(item)
        h += '</div>'
    h += '</div></details>'
    return h

def render_warehouse_full():
    """Python 端预生成完整仓库 HTML，inline onclick，零外部文件依赖"""
    import json
    data = json.loads(warehouse_data_json())
    body = (
        _wh_render_boots(data) +
        _wh_render_section(data.get('belts', []), 'belt', '🎽 腰带', '#C62828') +
        _wh_render_section(data.get('uniforms', []), 'uniform', '👔 礼服', '#1565C0')
    )
    return '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1,user-scalable=no">
<title>物资仓库 — 国旗仪仗队</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:"Microsoft YaHei","PingFang SC",sans-serif;background:#f5f0ec;color:#333;min-height:100vh;font-size:14px}
.header{background:linear-gradient(135deg,#8B0000,#B71C1C);color:#fff;padding:12px 16px;position:sticky;top:0;z-index:50}
.header h1{font-size:1rem;font-weight:700}.header .sub{font-size:.65rem;opacity:.85}
.search-row{display:flex;gap:6px;margin-top:8px;position:relative}
.search-row input{flex:1;padding:10px 12px;border:none;border-radius:8px;font-size:.85rem;outline:none;background:#fff;min-width:0}
.search-row button{padding:10px 16px;border:none;border-radius:8px;background:#E65100;color:#fff;font-weight:700;cursor:pointer;font-size:.85rem;white-space:nowrap}
.suggestions{position:absolute;top:44px;left:0;right:0;background:#fff;border-radius:8px;box-shadow:0 4px 20px rgba(0,0,0,.2);z-index:100;max-height:180px;overflow-y:auto;display:none;font-size:.8rem;color:#222}
.suggestions .item{padding:10px 14px;cursor:pointer;border-bottom:1px solid #e8e4de;color:#222}
.suggestions .item:hover{background:#FFF3E0;color:#111}
.main{padding:8px}
.section{margin-bottom:12px;background:#fff;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.06);overflow:hidden;border:1px solid #e8e0d8}
details.section summary::-webkit-details-marker{display:none}
details.section summary{list-style:none;cursor:pointer;-webkit-tap-highlight-color:transparent;user-select:none}
details.section summary:active{opacity:.7}
.section-header{padding:8px 14px;font-size:.8rem;font-weight:700;border-bottom:2px solid #8B0000;display:flex;align-items:center;gap:8px;background:#fff}
.section-header .tag{padding:3px 10px;border-radius:4px;color:#fff;font-size:.65rem;font-weight:600}
.table-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;padding:4px}
.grid-table{display:grid;gap:2px;padding:2px}
.cell{border-radius:4px;display:flex;flex-direction:column;align-items:center;justify-content:center;padding:8px 6px;min-width:70px;min-height:50px;cursor:pointer;transition:all .15s;text-align:center;border:1px solid transparent;-webkit-tap-highlight-color:transparent}
.cell:active{transform:scale(.95)}
.cell.hl{border-color:#ff5722!important;box-shadow:0 0 0 2px #ff5722,0 0 12px rgba(255,87,34,.3)!important;z-index:10;animation:glow .8s ease infinite}
.cell.empty{background:#fafaf7;border-color:#e8e4de;cursor:default}
.cell.empty:active{transform:none}
.cell .cc{font-weight:700;font-size:.68rem;margin-bottom:2px}
.cell .cn{font-size:.6rem;max-width:100%;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:#666}
.cell .ct{font-size:.5rem;color:#fff;padding:1px 4px;border-radius:2px;margin-top:1px}
.cell.gh{min-width:auto;min-height:22px;padding:2px 3px;font-size:.65rem}
.cell.gh:active{transform:none}
.cell.boot{background:#FFF3E0;border-color:#FFE0B2}.cell.boot .cc{color:#BF360C}
.cell.belt{background:#FCE4EC;border-color:#F8BBD0}.cell.belt .cc{color:#880E4F}
.cell.uniform{background:#E3F2FD;border-color:#BBDEFB}.cell.uniform .cc{color:#0D47A1}
@keyframes glow{0%,100%{box-shadow:0 0 0 2px #ff5722}50%{box-shadow:0 0 0 5px #ff5722,0 0 16px rgba(255,87,34,.4)}}
.panel{position:sticky;top:0;background:#fff;border-radius:0 0 12px 12px;box-shadow:0 4px 20px rgba(0,0,0,.15);padding:16px;z-index:60;max-height:50vh;overflow-y:auto;display:none;border-bottom:3px solid #8B0000}
.panel.show{display:block}
.panel .ph{display:flex;justify-content:space-between;align-items:center;margin-bottom:8px}
.panel .ph h3{font-size:.85rem;color:#8B0000}
.panel .pc{font-size:.78rem;line-height:2;color:#555}
.panel .pc strong{color:#333}
.panel .close{font-size:1.2rem;color:#999;cursor:pointer;padding:4px 8px;background:none;border:none}
.overlay{position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,.3);z-index:55;display:none}
.overlay.show{display:block}
.cab-label{font-size:.65rem;font-weight:700;padding:3px 10px;border-left:3px solid;margin:4px 0 2px}
.tip{text-align:center;color:#999;font-size:.65rem;padding:8px}
</style>
</head>
<body>
<div class="header">
  <h1>🇨🇳 国旗仪仗队 · 物资仓库</h1>
  <div class="sub">马靴 · 腰带 · 礼服 存放位置</div>
  <div class="search-row">
    <input type="text" id="q" placeholder="搜索库位/姓名/尺码..." oninput="wh_si()">
    <button onclick="wh_ds()">🔍</button>
    <div class="suggestions" id="sug"></div>
  </div>
</div>
<div class="panel" id="pn"><div class="ph"><h3 id="pt"></h3><button class="close" onclick="wh_close()">✕</button></div><div class="pc" id="pb"></div></div>
<div class="main"><div id="ct">''' + body + '''</div><div class="tip">点击格子查看详情 | 左右滑动表格</div></div>
<div class="overlay" id="ov" onclick="wh_close()"></div>

<script>
function wh_open(title,body){document.getElementById('pt').innerHTML=title;document.getElementById('pb').innerHTML=body;document.getElementById('pn').classList.add('show');document.getElementById('ov').classList.add('show')}
function wh_close(){document.getElementById('pn').classList.remove('show');document.getElementById('ov').classList.remove('show')}

// Event delegation - no inline onclick (CSP-safe)
document.querySelector('.main').addEventListener('click',function(e){
  var el=e.target.closest('.cell[data-type]');
  if(!el)return;
  var t=el.dataset,code=t.code||'',person=t.person||'',cab=t.cabinet||'',belt=t.belt||'';
  if(t.type==='boot'){
    wh_open('🥾 马靴 '+code,'📍 库位：<strong>'+code+'</strong><br>👤 '+(person||'未分配'));
  }else if(t.type==='belt'){
    var g=t.gender||'';
    wh_open('🎽 腰带 '+code,g+' · '+cab+'<br>👔 礼服尺码：<strong>'+(t.uniform||'')+'</strong><br>👤 <strong>'+(person||'未分配')+'</strong>');
  }else if(t.type==='uniform'){
    var st=t.hasShirt==='1'?' <span style="background:#4CAF50;color:#fff;padding:1px 6px;border-radius:3px;font-size:.65rem">有配套衬衫</span>':'';
    wh_open('👔 礼服 '+code+st,'🏷️ '+cab+'<br>🎽 腰带：<strong>'+belt+'</strong><br>👤 <strong>'+(person||'未分配')+'</strong>');
  }
});

// Search index
(function(){
var ALL=[];
document.querySelectorAll('.cell[data-code]').forEach(function(el){
  ALL.push({el:el,type:el.dataset.type||'',code:el.dataset.code||'',person:el.dataset.person||''});
});
window.wh_si=function(){
  var q=document.getElementById('q').value.trim(),s=document.getElementById('sug');
  if(q.length<1){s.style.display='none';return}
  var ql=q.toLowerCase();
  var ms=ALL.filter(function(a){return a.code.toLowerCase().indexOf(ql)>=0||(a.person&&a.person.indexOf(q)>=0)}).slice(0,6);
  if(!ms.length){s.style.display='none';return}
  s.innerHTML=ms.map(function(m){var icon=m.type==='boot'?'🥾':m.type==='belt'?'🎽':'👔';return '<div class="item">'+icon+' '+m.code+'</div>'}).join('');
  s.style.display='block';
};
window.wh_ds=function(){
  var q=document.getElementById('q').value.trim();if(!q)return;
  document.getElementById('sug').style.display='none';
  var ql=q.toLowerCase();
  var ms=ALL.filter(function(a){return a.code.toLowerCase().indexOf(ql)>=0});
  if(!ms.length){wh_open('🔍 搜索','未找到 "'+q+'"');return}
  ms[0].el.scrollIntoView({behavior:'smooth',block:'center'});
  ms[0].el.click();
};
})();
</script>
</body>
</html>'''


# ═══════════════════════════════════════
#  主路由
# ═══════════════════════════════════════

if 'page' not in st.session_state:
    st.session_state.page = 'home'

inject_css()

st.markdown("""
<style>
/* ============================================================
   全局 CSS —— 与 Flask 本地版完全一致
   自动从 server.py FACULTY_HTML / HOME_HTML 移植
   ============================================================ */

/* CSS 变量 */
:root {
  --flag-red: #B81616; --flag-red-dark: #8E1010; --flag-red-disabled: #D9A4A4;
  --gold: #F5C518; --gold-light: #FDE68A;
  --green-900: #0f2518; --green-700: #1a3a2a; --green-500: #2d5a3f; --green-300: #5a8a6a;
  --cream: #F5F1E6; --paper-warm: #f0ebe3; --paper-card: #fefcf8;
  --ink: #2c1810; --ink-light: #5c4a3a; --ink-faint: #9c8a7a;
  --white: #ffffff; --gray-300: #d4d4cc; --gray-500: #8a8a80;
  --font-heading: 'SimSun','KaiTi','宋体','楷体','Microsoft YaHei',serif;
  --font-body: 'Microsoft YaHei','PingFang SC',system-ui,sans-serif;
  --font-mono: 'SimSun','Consolas','Courier New',monospace;
}

*{box-sizing:border-box;margin:0;padding:0}

/* 修复 Streamlit 默认 padding 和背景 */
.stApp {background-color: var(--cream)}
.block-container {padding: 0 !important; max-width: 100% !important}
section.main > .block-container {padding-top: 0 !important}
div[data-testid="stVerticalBlock"] {gap: 0 !important}

/* === 隐藏 Streamlit 原生元素 === */
#MainMenu, footer, header[data-testid="stHeader"] {visibility: hidden}

/* === PAGE BACKGROUND === */
.stApp::before {content:'';position:fixed;inset:0;pointer-events:none;z-index:0;
  background:radial-gradient(circle at 15% 20%, rgba(245,197,24,0.03) 0%, transparent 40%),
             radial-gradient(circle at 85% 60%, rgba(184,22,22,0.02) 0%, transparent 40%),
             radial-gradient(circle at 50% 90%, rgba(245,197,24,0.02) 0%, transparent 30%);
}

/* === MICRO ANIMATIONS === */
@keyframes fadeInUp{from{opacity:0;transform:translateY(20px)}to{opacity:1;transform:translateY(0)}}
@keyframes fadeInDown{from{opacity:0;transform:translateY(-16px)}to{opacity:1;transform:translateY(0)}}

/* === HEADER === */
.flag-header {
  background:var(--flag-red);color:var(--white);padding:20px 32px;
  display:flex;align-items:center;justify-content:space-between;
  position:relative;z-index:1;animation:fadeInDown .8s ease-out;
  border-bottom:2px solid rgba(245,197,24,.3);
  margin: 0 0 32px 0;
}
.flag-header .brand{display:flex;align-items:center;gap:12px}
.flag-header h1 {
  font-family:var(--font-heading);font-size:18px;font-weight:700;letter-spacing:.08em;
  color:var(--white);margin:0;
}
.flag-header .stars {display:flex;gap:4px;font-size:14px;color:var(--gold);line-height:1}

/* === MAIN CONTENT WRAPPER === */
.main-wrap {
  max-width:860px;margin:0 auto;padding:0 24px;position:relative;z-index:1;
}

/* === HOME PAGE ENTRIES GRID === */
.entries {
  display:grid;grid-template-columns:1fr 1fr;gap:20px;
  max-width:760px;margin:56px auto 48px;
  animation:fadeInUp .8s ease-out .15s both;
}
@media(max-width:640px){.entries{grid-template-columns:1fr}}

/* === HOME CARD (matches server.py HOME_HTML) === */
.home-card {
  display:block;background:var(--white);border:1px solid rgba(184,22,22,.08);
  border-left:3px solid var(--gold);border-radius:8px;padding:40px 28px;
  text-align:center;text-decoration:none;color:inherit;cursor:pointer;
  transition:all .3s ease;box-shadow:0 2px 12px rgba(0,0,0,.04);
  position:relative;overflow:hidden;
}
.home-card:hover {border-left-color:var(--flag-red);transform:translateY(-4px);box-shadow:0 8px 28px rgba(184,22,22,.1)}
.home-card .icon {font-size:42px;margin-bottom:14px;display:block;transition:transform .3s}
.home-card:hover .icon {transform:scale(1.05)}
.home-card .star-deco {display:inline-block;color:var(--gold);font-size:12px;margin:0 6px}
.home-card h3 {
  font-family:var(--font-heading);font-size:18px;font-weight:700;color:var(--green-900);
  margin-bottom:10px;letter-spacing:.06em;
}
.home-card p {font-size:13px;color:var(--ink-light);line-height:1.7}
/* Gold line accent */
.home-card::before {content:'';display:block;width:40px;height:2px;background:var(--gold);margin:0 auto 16px;transition:width .3s,background .3s}
.home-card:hover::before {width:56px;background:var(--flag-red)}

/* === FACULTY PAGE CARDS (matches server.py FACULTY_HTML) === */
.fac-card {
  background:var(--paper-card);border:1px solid rgba(184,22,22,.08);
  border-left:3px solid var(--gold);padding:28px;margin-bottom:20px;
  position:relative;animation:fadeInUp .7s ease-out both;box-shadow:0 2px 8px rgba(0,0,0,.03);
}
.fac-card:nth-child(1){animation-delay:.1s}
.fac-card:nth-child(2){animation-delay:.2s}
.fac-card h2 {
  font-family:var(--font-heading);font-size:17px;display:flex;align-items:center;gap:10px;
  margin-bottom:18px;padding-bottom:12px;border-bottom:2px solid rgba(184,22,22,.15);
  letter-spacing:.05em;color:var(--green-900);
}

/* === BADGE (step numbers) === */
.badge {
  width:28px;height:28px;background:var(--flag-red);color:var(--white);
  font-family:var(--font-heading);font-size:14px;font-weight:700;
  display:inline-flex;align-items:center;justify-content:center;flex-shrink:0;
}

/* === FORM LABELS === */
label.block {
  display:block;font-family:var(--font-heading);font-weight:600;
  margin-bottom:6px;margin-top:14px;font-size:14px;
  color:var(--green-900);letter-spacing:.04em;
}
label.block:first-child{margin-top:0}

/* === TEXTAREAS === */
textarea {
  width:100%;height:130px;border:1px solid var(--ink-faint);padding:12px;
  font:13px var(--font-mono);resize:vertical;background:var(--cream);
  color:var(--ink);line-height:1.8;transition:border-color .2s;
}
textarea:focus {outline:none;border-color:var(--flag-red);box-shadow:0 0 0 2px rgba(184,22,22,.1)}
textarea::placeholder {color:var(--ink-faint);font-style:italic}

/* === UPLOAD ZONE === */
.up {
  border:2px solid var(--ink-faint);padding:32px;text-align:center;
  cursor:pointer;transition:all .25s;background:var(--cream);position:relative;
}
.up::before {content:'';position:absolute;top:8px;left:8px;width:14px;height:14px;border-top:1px solid var(--flag-red);border-left:1px solid var(--flag-red);opacity:.3}
.up::after {content:'';position:absolute;bottom:8px;right:8px;width:14px;height:14px;border-bottom:1px solid var(--flag-red);border-right:1px solid var(--flag-red);opacity:.3}
.up:hover {border-color:var(--flag-red);background:var(--paper-card)}
.up:hover::before,.up:hover::after {border-color:var(--flag-red);opacity:.7}

/* === OR DIVIDER === */
.or {
  display:flex;align-items:center;gap:14px;margin:16px 0;
}
.or hr {flex:1;border:none;border-top:1px solid var(--ink-faint)}
.or span {
  font-family:var(--font-heading);color:var(--ink-faint);font-size:13px;
  letter-spacing:.1em;position:relative;padding:0 8px;
}
.or span::before {content:'\2605';font-size:8px;position:absolute;left:-4px;top:50%;transform:translateY(-50%);color:var(--gold)}
.or span::after {content:'\2605';font-size:8px;position:absolute;right:-4px;top:50%;transform:translateY(-50%);color:var(--gold)}

/* === BUTTONS === */
.btn {
  width:100%;padding:14px;border:none;font-family:var(--font-heading);
  font-size:16px;font-weight:700;letter-spacing:.08em;cursor:pointer;transition:all .25s;
}
.btn-b {
  background:var(--flag-red);color:var(--white);box-shadow:0 2px 6px rgba(184,22,22,.2);
  border-radius:6px;
}
.btn-b:hover:not(:disabled) {background:var(--flag-red-dark);box-shadow:0 4px 12px rgba(184,22,22,.3);transform:translateY(-1px)}
.btn-b:disabled {background:var(--flag-red-disabled);color:rgba(255,255,255,.6);cursor:not-allowed;box-shadow:none;transform:none}
.btn-g {
  background:var(--green-700);color:var(--gold);border:1px solid var(--gold);border-radius:6px;
}
.btn-g:hover:not(:disabled) {background:var(--green-500);transform:translateY(-1px)}
.btn-g:disabled {background:var(--gray-300);color:var(--gray-500);border-color:var(--gray-300);cursor:not-allowed;transform:none}

/* === SPINNER === */
.spin {
  width:22px;height:22px;border:2.5px solid rgba(245,197,24,.2);
  border-top-color:var(--gold);border-radius:50%;animation:s .7s linear infinite;
}
@keyframes s{to{transform:rotate(360deg)}}

/* === ROLES GRID === */
.roles-grid {display:grid;grid-template-columns:1fr 1fr;gap:14px}

/* === STATS GRID === */
.stats {
  display:grid;grid-template-columns:repeat(3,1fr);gap:0;
  border:1px solid var(--ink-faint);margin:16px 0;
}
.st {
  padding:20px 16px;text-align:center;border-right:1px solid var(--ink-faint);
  background:var(--paper-card);position:relative;
}
.st:last-child{border-right:none}
.st.c{border-top:3px solid var(--flag-red)}.st.r{border-top:3px solid var(--gold)}.st.p{border-top:3px solid var(--green-500)}
.st b {
  font-family:var(--font-heading);font-size:30px;font-weight:700;display:block;line-height:1.1;
}
.st.c b{color:var(--flag-red)}.st.r b{color:var(--gold)}.st.p b{color:var(--green-500)}
.st span {
  color:var(--ink-light);font-size:12px;letter-spacing:.08em;text-transform:uppercase;
}

/* === RESULT TABLES === */
table {
  width:100%;border-collapse:collapse;font-size:13px;margin-top:8px;
  border:1px solid var(--ink-faint);
}
th {
  background:var(--green-900);color:var(--gold-light);padding:8px 10px;text-align:center;
  font-family:var(--font-heading);font-weight:600;font-size:12px;letter-spacing:.05em;
  text-transform:uppercase;border:1px solid rgba(245,197,24,.15);
}
td {padding:8px 10px;border:1px solid var(--ink-faint);text-align:center}
tr:nth-child(even) td {background:var(--paper-warm)}
.old {text-decoration:line-through;color:var(--ink-faint);font-size:12px}
.hl {background:#fffde7;font-weight:600}

/* === HINT / ERROR === */
.hint {
  display:block;background:#fffde7;border-left:3px solid var(--gold);
  padding:10px 14px;margin-top:10px;font-size:13px;color:#6b5a20;
}
.err-msg {
  display:block;background:#fdf0f2;border-left:3px solid var(--flag-red);
  padding:12px 14px;color:var(--flag-red);margin-top:12px;font-size:13px;
}

/* === SECTION TITLES === */
.sect-title {
  font-family:var(--font-heading);font-size:15px;font-weight:700;
  margin:16px 0 6px;color:var(--green-900);letter-spacing:.05em;
}

/* === FOOTER === */
.flag-footer {
  text-align:center;padding:28px;font-family:var(--font-heading);
  font-size:12px;color:var(--green-700);letter-spacing:.06em;
  border-top:1px solid rgba(184,22,22,.08);margin:32px 0 0 0;
}
.flag-footer .star {color:var(--gold);font-size:10px;margin:0 6px}

/* === WAREHOUSE GRID === */
.section {margin-bottom:16px;background:var(--white);border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,.04);overflow:hidden;border:1px solid rgba(184,22,22,.06);border-left:3px solid var(--gold)}
.section-header {padding:10px 16px;font-size:13px;font-weight:700;border-bottom:2px solid rgba(184,22,22,.08);display:flex;align-items:center;gap:10px;background:var(--white);font-family:var(--font-heading);letter-spacing:.04em;color:#0f2518}
.section-header .tag {padding:4px 12px;border-radius:4px;color:var(--white);font-size:11px;font-weight:700}
.table-wrap {overflow-x:auto;-webkit-overflow-scrolling:touch;padding:6px}
.grid-table {display:grid;gap:2px;padding:4px}
.cell {border-radius:4px;display:flex;flex-direction:column;align-items:center;justify-content:center;padding:10px 8px;min-width:72px;min-height:52px;cursor:pointer;transition:all .15s;text-align:center;border:1px solid transparent}
.cell:hover {transform:translateY(-2px);box-shadow:0 3px 12px rgba(0,0,0,.1)}
.cell .cc {font-weight:700;font-size:11px;margin-bottom:2px}
.cell .cn {font-size:10px;max-width:100%;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:var(--ink-light)}
.cell .ct {font-size:9px;color:var(--white);padding:1px 5px;border-radius:3px;margin-top:2px}
.cell.gh {min-width:auto;min-height:24px;padding:2px 4px;font-size:11px}
.cell.gh:hover {transform:none;box-shadow:none}
.cell.empty {background:#fafaf7;border-color:#e8e0d8;cursor:default}
.cell.empty:hover {transform:none;box-shadow:none}
.cell.boot {background:#FFF3E0;border-color:#FFE0B2}.cell.boot .cc {color:#BF360C}
.cell.belt {background:#FCE4EC;border-color:#F8BBD0}.cell.belt .cc {color:#880E4F}
.cell.uniform {background:#E3F2FD;border-color:#BBDEFB}.cell.uniform .cc {color:#0D47A1}
.cab-label {font-size:11px;font-weight:700;padding:4px 12px;border-left:3px solid;margin:6px 0 2px;font-family:var(--font-heading);letter-spacing:.04em}

/* === WAREHOUSE PANEL & OVERLAY === */
.panel {position:fixed;bottom:0;left:0;right:0;background:var(--white);border-radius:12px 12px 0 0;box-shadow:0 -4px 20px rgba(0,0,0,.15);padding:20px;z-index:60;max-height:50vh;overflow-y:auto;display:none;border-top:3px solid var(--flag-red)}
.panel.show {display:block}
.panel .ph {display:flex;justify-content:space-between;align-items:center;margin-bottom:10px}
.panel .ph h3 {font-size:15px;color:var(--flag-red);font-family:var(--font-heading)}
.panel .pc {font-size:13px;line-height:2.2;color:var(--ink-light)}
.panel .pc strong {color:var(--ink)}
.panel .close {font-size:18px;color:var(--ink-faint);cursor:pointer;padding:4px 10px;background:none;border:none;border-radius:4px}
.panel .close:hover {color:var(--flag-red)}
.overlay {position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,.3);z-index:55;display:none}
.overlay.show {display:block}
.cell.hl {border-color:var(--flag-red)!important;box-shadow:0 0 0 2px var(--flag-red),0 0 12px rgba(184,22,22,.3)!important;z-index:10;animation:glow .8s ease infinite}
@keyframes glow{0%,100%{box-shadow:0 0 0 2px var(--flag-red)}50%{box-shadow:0 0 0 5px var(--flag-red),0 0 16px rgba(184,22,22,.4)}}

/* === Streamlit 组件样式覆盖 === */
.stButton > button {
  background-color: var(--flag-red) !important;
  color: var(--white) !important;
  border: none !important;
  font-weight: 700 !important;
  font-family: var(--font-heading) !important;
  font-size: 16px !important;
  letter-spacing: .08em !important;
  border-radius: 6px !important;
  padding: 14px 20px !important;
  transition: all .25s !important;
  box-shadow: 0 2px 6px rgba(184,22,22,.2) !important;
  width: auto !important;
}
.stButton > button:hover {
  background-color: var(--flag-red-dark) !important;
  box-shadow: 0 4px 12px rgba(184,22,22,.3) !important;
  transform: translateY(-1px) !important;
}
.stButton > button:disabled {
  background-color: var(--flag-red-disabled) !important;
  color: rgba(255,255,255,.6) !important;
  box-shadow: none !important;
  transform: none !important;
}

[data-testid="stFileUploader"] {
  background: var(--cream) !important;
  border: 2px solid var(--ink-faint) !important;
  border-radius: 0 !important;
  padding: 32px !important;
  text-align: center !important;
  position: relative !important;
}
/* corner bracket decorations — matching .up from Flask */
[data-testid="stFileUploader"]::before {
  content: '' !important; position: absolute !important;
  top: 8px !important; left: 8px !important;
  width: 14px !important; height: 14px !important;
  border-top: 1px solid var(--flag-red) !important;
  border-left: 1px solid var(--flag-red) !important;
  opacity: 0.3 !important; pointer-events: none !important;
}
[data-testid="stFileUploader"]::after {
  content: '' !important; position: absolute !important;
  bottom: 8px !important; right: 8px !important;
  width: 14px !important; height: 14px !important;
  border-bottom: 1px solid var(--flag-red) !important;
  border-right: 1px solid var(--flag-red) !important;
  opacity: 0.3 !important; pointer-events: none !important;
}
[data-testid="stFileUploader"]:hover {
  border-color: var(--flag-red) !important;
  background: var(--paper-card) !important;
}
[data-testid="stFileUploader"]:hover::before,
[data-testid="stFileUploader"]:hover::after {
  border-color: var(--flag-red) !important;
  opacity: 0.7 !important;
}

.stTextArea textarea {
  font-family: var(--font-mono) !important;
  font-size: 13px !important;
  line-height: 1.8 !important;
  height: 130px !important;
  border: 1px solid var(--ink-faint) !important;
  background: var(--cream) !important;
  color: var(--ink) !important;
}
.stTextArea textarea:focus {
  outline: none !important;
  border-color: var(--flag-red) !important;
  box-shadow: 0 0 0 2px rgba(184,22,22,.1) !important;
}

.stTextInput input {
  border: 1px solid var(--ink-faint) !important;
  background: var(--cream) !important;
  color: var(--ink) !important;
  font-family: var(--font-mono) !important;
}
.stTextInput input:focus {
  border-color: var(--flag-red) !important;
  box-shadow: 0 0 0 2px rgba(184,22,22,.1) !important;
}

[data-testid="stExpander"] {
  background: transparent !important;
  border: 1px solid var(--ink-faint) !important;
  border-radius: 0 !important;
}
</style>

<div class="flag-header">
  <div class="brand">
    <div class="stars">&#9733;</div>
    <h1>浙江大学国旗仪仗队</h1>
  </div>
  <div class="stars">&#9733;</div>
</div>
""", unsafe_allow_html=True)



# ════════════════════════════════
# 页面渲染
# ════════════════════════════════

if st.session_state.page == 'home':
    st.markdown('<div class="entries">', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("""
        <div class="home-card">
          <div class="icon">&#127891;</div>
          <h3>院系升旗礼服分配 <span class="star-deco">&#9733;</span></h3>
          <p>上传库存表 + 拍照上传人员安排表<br>自动识别 → 按尺寸排序 → 冲突检测 → 下载分配表</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("进入", key="btn_faculty", use_container_width=True):
            st.session_state.page = 'faculty'; st.rerun()

    with col2:
        st.markdown("""
        <div class="home-card">
          <div class="icon">&#128230;</div>
          <h3>物资仓库</h3>
          <p>马靴 · 腰带 · 礼服 存放位置<br>搜索库位 / 姓名 / 尺码</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("进入", key="btn_warehouse", use_container_width=True):
            st.session_state.page = 'warehouse'; st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("""
    <div class="flag-footer">
      <span class="star">&#9733;</span> 浙江大学国旗仪仗队 &copy; 2026 <span class="star">&#9733;</span>
    </div>
    """, unsafe_allow_html=True)


elif st.session_state.page == 'faculty':
    # 返回首页——页面切换在按钮回调里直接做
    back_home = st.button("← 返回首页", key="back_home")
    if back_home:
        st.session_state.fac_roster = ''
        st.session_state.fac_img = None
        st.session_state._ocr_tag = ''
        st.session_state.page = 'home'; st.rerun()

    st.markdown('<div class="main-wrap">', unsafe_allow_html=True)

    # ====== 先放 OCR 逻辑（必须在 text_area 创建之前） ======
    # 步骤① — fac-card
    st.markdown("""
    <div class="fac-card">
    <h2><span class="badge">1</span> 上传礼服库存表 (.xlsx)</h2>
    """, unsafe_allow_html=True)
    excel_file = st.file_uploader("拖拽或点击上传 .xlsx 文件", type=['xlsx'], key="fac_excel", label_visibility="collapsed")
    st.markdown('</div>', unsafe_allow_html=True)

    # 步骤② 照片上传
    st.markdown("""
    <div class="fac-card">
    <h2><span class="badge">2</span> 拍照上传人员安排表</h2>
    """, unsafe_allow_html=True)
    image_file = st.file_uploader("上传照片", type=['png','jpg','jpeg'], key="fac_img", label_visibility="collapsed")

    # —— OCR 自动触发（必须在 text_area 创建前） ——
    if excel_file is not None and image_file is not None:
        # 用文件名+大小作为指纹，文件变了就重新 OCR
        file_tag = f"{excel_file.name}:{excel_file.size}|{image_file.name}:{image_file.size}"
        if st.session_state.get('_ocr_tag', '') != file_tag:
            with st.spinner("AI 正在识别照片中的人员名单..."):
                ocr_text, ocr_roles = ocr_faculty_roster(image_file.getvalue(), excel_file.getvalue())
            st.session_state.fac_roster = ocr_text if ocr_text else ''
            st.session_state._ocr_tag = file_tag
            st.rerun()

    st.markdown('<div class="or"><hr><span>或手动输入 / 修正队列人员</span><hr></div>', unsafe_allow_html=True)
    st.markdown('<p style="color:var(--ink-light);font-size:13px;margin:8px 0">每行一人或顿号分隔。总负责、场控、后勤、摄影不参与分配。</p>', unsafe_allow_html=True)

    roster = st.text_area("队列人员",
                          placeholder="林珩\n韩雅丽\n艾克达\n张鹏\n夏瑞泽\n戴傲\n叶宇轩",
                          key="fac_roster", height=130, label_visibility="collapsed")

    with st.expander("其他角色（可选，不参与分配）"):
        c1, c2 = st.columns(2)
        m1 = c1.text_input("总负责", key="fm1")
        m2 = c2.text_input("场控", key="fm2")
        m3, m4 = st.columns(2)
        m3_ = m3.text_input("后勤", key="fm3")
        m4_ = m4.text_input("摄影", key="fm4")

    st.markdown('</div>', unsafe_allow_html=True)  # close fac-card

    # 生成按钮
    can_gen = excel_file is not None
    clicked = st.button("🔍 预览排序 & 生成分配表", disabled=not can_gen, use_container_width=True)

    # —— 点击按钮：直接生成 ——
    if clicked:
        excel_bytes = excel_file.getvalue()
        roster_val = roster.strip()
        if not roster_val:
            st.error("请先在文本框输入队列人员名单，或上传照片让 AI 自动识别")
            st.stop()

        with st.status("正在生成分配方案...", expanded=True):
            try:
                queue_names = []
                for chunk in re.split(r'[、，,\n\s]+', roster_val):
                    n = chunk.strip()
                    if len(n) >= 2: queue_names.append(n)

                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as t:
                    t.write(excel_bytes); tmp = t.name

                try:
                    wb = openpyxl.load_workbook(tmp)
                    all_persons = parse_equipment_sheet(wb['装备分配'])
                    name_map = {p['name']: p for p in all_persons}
                    pool = build_pool(wb)
                    # 为 boots/belt 提供空池（与 Flask 版一致）
                    for key in ['boots', 'belt']:
                        if key not in pool:
                            pool[key] = set()

                    faculty_persons = [dict(name_map[n]) for n in queue_names if n in name_map]
                    missing = [n for n in queue_names if n not in name_map]

                    conflicts = detect_faculty_conflicts(faculty_persons)
                    reassigns, changed = resolve_faculty_conflicts(conflicts, faculty_persons, pool)
                    b64 = generate_faculty_excel(faculty_persons, changed)
                    sorted_people = sort_people_by_uniform(faculty_persons)

                    msg = f"生成完成: {len(faculty_persons)} 人, {len(conflicts)} 冲突, {len(reassigns)} 重分配"
                    if missing: msg += f"（{len(missing)} 人未在库存中找到：{', '.join(missing)}）"

                    affected = len(set(r['person'] for r in reassigns))

                    # Build result HTML to match Flask version exactly
                    result_html = '<div class="fac-card" style="animation:fadeInUp .5s ease-out">'
                    result_html += '<h2><span class="badge">3</span> 生成结果</h2>'
                    result_html += f'<p style="color:var(--ink-light);margin-bottom:16px">{msg}</p>'
                    result_html += '<div class="stats">'
                    result_html += f'<div class="st c"><b>{len(conflicts)}</b><span>冲突数</span></div>'
                    result_html += f'<div class="st r"><b>{len(reassigns)}</b><span>重分配</span></div>'
                    result_html += f'<div class="st p"><b>{affected}</b><span>受影响人数</span></div>'
                    result_html += '</div>'
                    result_html += '<div class="sect-title">排序后人员列表</div>'
                    result_html += '<table><tr><th>序号</th><th>姓名</th><th>性别</th><th>礼服</th><th>礼帽</th><th>马靴</th><th>腰带</th></tr>'
                    result_html += make_person_rows(sorted_people, changed)
                    result_html += '</table>'
                    result_html += '</div>'
                    st.markdown(result_html, unsafe_allow_html=True)

                    st.download_button("📥 下载礼服分配表 (.xlsx)",
                                      data=base64.b64decode(b64),
                                      file_name="院系升旗礼服分配.xlsx",
                                      mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                      use_container_width=True)

                finally:
                    os.unlink(tmp)

            except Exception as e:
                import traceback
                st.error(f"生成失败: {e}")
                with st.expander("详细错误"): st.code(traceback.format_exc())


elif st.session_state.page == 'warehouse':
    if st.button("← 返回首页", key="wh_back"):
        st.session_state.page = 'home'; st.rerun()

    st.markdown("---")

    # 搜索栏
    st.text_input("搜索库位 / 姓名 / 尺码...", key="wh_search",
                   placeholder="搜索库位 / 姓名 / 尺码...", label_visibility="collapsed")

    # 用 st.components.v1.html 才能跑 JS（st.markdown 会过滤 script）
    wh_full = render_warehouse_full()
    st.components.v1.html(wh_full, height=800, scrolling=True)
