import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook('verify_debug.xlsx')
ws = wb['装备分配']

print('Checking verify_debug.xlsx:')
for row in range(2, ws.max_row + 1):
    name = ws.cell(row=row, column=2).value
    if not name: continue
    name = str(name).strip()
    uniform = str(ws.cell(row=row, column=3).value or '')
    hat = str(ws.cell(row=row, column=4).value or '')
    boots = str(ws.cell(row=row, column=5).value or '')
    belt = str(ws.cell(row=row, column=6).value or '')

    # Check fill colors
    fills = []
    for col in [3, 4, 5, 6]:
        cell = ws.cell(row=row, column=col)
        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
            rgb = str(cell.fill.start_color.rgb)
            if 'FFFF00' in rgb or 'FFFFFF00' in rgb:
                fills.append(col)

    if fills or name in ('柳洋', '孙鸣', '马欣雅', '张博宣'):
        print(f'  Row {row}: {name} | {uniform} | {hat} | {boots} | {belt} | FILLS={fills}')

# Also check what the template looks like
wb2 = openpyxl.load_workbook('本月礼服分配.xlsx')
ws2 = wb2['装备分配']
print('\nTemplate (original) for comparison:')
for row in range(2, ws2.max_row + 1):
    name = ws2.cell(row=row, column=2).value
    if not name: continue
    name = str(name).strip()
    if name in ('柳洋', '孙鸣', '马欣雅', '张博宣'):
        uniform = str(ws2.cell(row=row, column=3).value or '')
        hat = str(ws2.cell(row=row, column=4).value or '')
        print(f'  Row {row}: {name} | {uniform} | {hat}')
