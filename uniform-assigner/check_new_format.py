import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook('verify_debug.xlsx')
ws = wb['装备分配']

print('=== verify_debug.xlsx (new format) ===')
print(f'Max row: {ws.max_row}')
for row in range(1, ws.max_row + 1):
    vals = []
    fills = []
    for col in range(1, 7):
        v = ws.cell(row=row, column=col).value
        vals.append(str(v)[:30] if v else '')
        cell = ws.cell(row=row, column=col)
        rgb = str(cell.fill.start_color.rgb) if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb else ''
        fills.append('Y' if 'FFFF00' in rgb or 'FFFFFF00' in rgb else '')
    print(f'  Row {row}: {" | ".join(vals)}')
    if any(fills):
        print(f'          YELLOW: {fills}')
