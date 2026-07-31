import openpyxl, json
wb = openpyxl.load_workbook(r'C:\Users\夏瑞泽\xwechat_files\wxid_vnl456wm0r9h12_0c22\msg\file\2026-07\2025-2026学年夏五周到夏八周礼服分配(1)(1).xlsx')
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'=== {sn} ===')
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        rows.append([str(c) if c is not None else '' for c in row])
    print(json.dumps(rows, ensure_ascii=False))
    print()
wb.close()
