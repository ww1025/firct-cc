import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json

EXCEL_PATH = r"c:\Users\夏瑞泽\xwechat_files\wxid_vnl456wm0r9h12_0c22\msg\file\2026-07\2025-2026学年夏五周到夏八周礼服分配(1)(1).xlsx"

wb = openpyxl.load_workbook(EXCEL_PATH)
print("=== SHEET NAMES:", wb.sheetnames)
print()

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"=== SHEET: {sname} ===")
    print(f"  Dimensions: {ws.dimensions}")
    print(f"  Max Row: {ws.max_row}, Max Col: {ws.max_column}")
    merges = [str(m) for m in ws.merged_cells.ranges]
    print(f"  Merged cells: {len(merges)}")
    for m in merges[:60]:
        print(f"    {m}")
    print(f"  Frozen panes: {ws.freeze_panes}")
    print()
