"""礼服分配系统 - Streamlit 版（Streamlit Cloud 入口）"""
import streamlit as st
import base64, io, os, re, tempfile, sys
from difflib import SequenceMatcher

st.set_page_config(page_title="礼服自动分配系统", page_icon="🎖️", layout="wide", initial_sidebar_state="collapsed")

YELLOW = 'FFFFFF00'

# 模板文件路径
FACULTY_TEMPLATE = os.path.join(os.path.dirname(__file__), 'uniform-assigner', '院系升旗装备分配表模板.xlsx')

if 'page' not in st.session_state:
    st.session_state.page = 'home'

@st.cache_resource
def get_ocr_reader():
    import easyocr
    return easyocr.Reader(['ch_sim'], gpu=False)

# ═══ 业务逻辑（全部从 server.py 复制）═══

def fuzzy_match(ocr_name, known_names):
    if not known_names: return ocr_name
    if ocr_name in known_names: return ocr_name
    cleaned = re.sub(r'[_|^~\s\d]+', '', ocr_name)
    if cleaned in known_names: return cleaned
    for ref in known_names:
        if len(ref) >= 2 and len(cleaned) >= 2 and cleaned[:2] == ref[:2]: return ref
    best, best_score = None, 0
    for ref in known_names:
        common = len(set(cleaned) & set(ref))
        ratio = SequenceMatcher(None, cleaned, ref).ratio()
        score = common * 3 + ratio * 5
        if score > best_score: best_score, best = score, ref
    if best and best_score >= 6: return best
    return re.sub(r'[_|^~\s\d]+', '', ocr_name)

def parse_equipment_sheet(ws):
    persons = []
    for row in range(2, ws.max_row+1):
        name = ws.cell(row=row, column=2).value
        if not name: continue
        name = str(name).strip()
        uni  = str(ws.cell(row=row, column=3).value or '').strip()
        hat  = str(ws.cell(row=row, column=4).value or '').strip()
        boots= str(ws.cell(row=row, column=5).value or '').strip()
        belt = str(ws.cell(row=row, column=6).value or '').strip()
        note = str(ws.cell(row=row, column=7).value or '').strip()
        persons.append(dict(name=name, gender='M' if uni.startswith('M') else 'F',
                            uniform=uni, hat=hat, boots=boots, belt=belt, note=note))
    return persons

def parse_shared_relations(persons):
    rels, seen = [], set()
    for p in persons:
        if not p['note']: continue
        for seg in re.split(r'[，；]', p['note']):
            seg = seg.strip()
            if not seg.startswith('和'): continue
            idx = seg.find('共用')
            if idx < 0: continue
            names = [n.strip() for n in re.split(r'[、，, ]+', seg[1:idx]) if n.strip()]
            after = seg[idx+2:]
            types = [t for t in [('uniform','礼服'),('hat','礼帽'),('boots','马靴'),('belt','腰带')] if t[1] in after]
            for n in names:
                for t, _ in types:
                    code = p[t]
                    key = f"{t}|{code}|{','.join(sorted([p['name'], n]))}"
                    if key not in seen:
                        seen.add(key)
                        rels.append(dict(item_code=code, item_type=t, shared_by=[p['name'], n]))
    return rels

def merge_relations(rels):
    m = {}
    for r in rels:
        k = f"{r['item_type']}|{r['item_code']}"
        m.setdefault(k, set()).update(r['shared_by'])
    return [dict(item_code=k.split('|')[1], item_type=k.split('|')[0], shared_by=list(v)) for k,v in m.items()]

def detect_conflicts(schedule, rels):
    conflicts = []
    for day in schedule:
        daily = set(day['people'])
        for r in rels:
            overlap = [p for p in r['shared_by'] if p in daily]
            if len(overlap) >= 2:
                for i in range(1, len(overlap)):
                    conflicts.append(dict(day=day['day'], item_type=r['item_type'], item_code=r['item_code'],
                                          person_to_move=overlap[i], person_to_keep=overlap[0]))
    return conflicts

def parse_uniform_code(code):
    m = re.match(r'^([FM])(\d+)/(\d+)-(\d+)$', code)
    return dict(gender=m.group(1), height=int(m.group(2)), chest=int(m.group(3))) if m else None

def parse_hat_code(code):
    m = re.match(r'^([FM])(\d{4})$', code)
    return dict(gender=m.group(1), number=int(m.group(2))) if m else None

def find_alternative(person, item_type, used_items, all_items):
    own = person[item_type]
    available = [x for x in all_items if x not in used_items and x != own]
    gender = person['gender']
    candidates = []
    for item in available:
        if item_type == 'uniform':
            info = parse_uniform_code(item)
            if not info or info['gender'] != gender: continue
            pi = parse_uniform_code(own) or dict(height=0, chest=0)
            score = abs(info['height'] - pi['height']) * 4 + abs(info['chest'] - pi['chest']) * 5
            candidates.append((score, item))
        elif item_type == 'hat':
            info = parse_hat_code(item)
            if not info or info['gender'] != gender: continue
            pi = parse_hat_code(own) or dict(number=0)
            candidates.append((abs(info['number'] - pi['number']), item))
        else:
            if item.startswith(gender): candidates.append((0, item))
    if not candidates:
        for item in available: candidates.append((999998, item))
    if not candidates:
        candidates = [(999999, x) for x in all_items if x.startswith(gender)]
    candidates.sort()
    return candidates[0][1] if candidates else None

def build_pool(wb):
    pool = dict(uniform=set(), hat=set())
    for row in range(2, wb['装备分配'].max_row+1):
        for col,key in [(3,'uniform'),(4,'hat')]:
            v = str(wb['装备分配'].cell(row=row, column=col).value or '').strip()
            if v: pool[key].add(v)
    for sn in ['礼服腰带摆放','礼帽摆放']:
        ws = wb[sn]
        for row in range(3, ws.max_row+1):
            for col in range(1, ws.max_column+1):
                v = str(ws.cell(row=row, column=col).value or '').strip()
                if re.match(r'^[FM]\d{2,3}/', v): pool['uniform'].add(v)
                if re.match(r'^[FM]\d{4}$', v): pool['hat'].add(v)
    return pool

def resolve(conflicts, persons, schedule, pool):
    reassigns, changed = [], {}
    by_day = {}
    for c in conflicts: by_day.setdefault(c['day'], []).append(c)
    for day, clist in by_day.items():
        people = next((set(s['people']) for s in schedule if s['day'] == day), set())
        used = {
            'uniform': {p['uniform'] for p in persons if p['name'] in people},
            'hat': {p['hat'] for p in persons if p['name'] in people}
        }
        for c in clist:
            person = next((p for p in persons if p['name'] == c['person_to_move']), None)
            if not person: continue
            t = c['item_type']
            alt = find_alternative(person, t, used[t], pool[t])
            if alt:
                used[t].discard(person[t]); used[t].add(alt)
                reassigns.append(dict(person=person['name'], item_type=t, old_item=person[t], new_item=alt))
                changed.setdefault(person['name'], {})[t] = alt
                person[t] = alt
    return reassigns, changed

def generate_excel(template_path, persons, changed):
    import openpyxl
    from openpyxl.styles import PatternFill
    wb = openpyxl.load_workbook(template_path)
    ws = wb['装备分配']
    name_map = {p['name']: p for p in persons}
    for row in range(2, ws.max_row+1):
        name = ws.cell(row=row, column=2).value
        if not name: continue
        name = str(name).strip()
        p = name_map.get(name)
        if not p: continue
        ch = changed.get(name, {})
        for col,key in [(3,'uniform'),(4,'hat'),(5,'boots'),(6,'belt'),(7,'note')]:
            cell = ws.cell(row=row, column=col)
            cell.value = p.get(key, '') or None
            if key in ch:
                cell.fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type='solid')
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return base64.b64encode(buf.read()).decode()

def parse_schedule(text):
    schedule = []
    day_patterns = [
        (r'周\s*一', '周一'), (r'周\s*二', '周二'),
        (r'周\s*三', '周三'), (r'周\s*四', '周四'), (r'周\s*五', '周五'),
    ]
    for line in text.strip().split('\n'):
        line = line.strip()
        if not line: continue
        day = None
        for pat, label in day_patterns:
            if re.search(pat, line): day = label; break
        if not day: continue
        rest = re.sub(r'周\s*[一二三四五]\s*[:：]?\s*', '', line)
        chunks = re.split(r'[、，,]', rest)
        names = []
        for chunk in chunks:
            n = re.sub(r'\s+', '', chunk)
            if len(n) >= 2: names.append(n)
        if names: schedule.append(dict(day=day, people=names))
    return schedule

def adjust_monday_schedule(schedule):
    """周一人员从周二~周五班抽调，只参加周一升旗。从原班移除。"""
    monday_people = set()
    for day in schedule:
        if day['day'] == '周一':
            monday_people = set(day['people'])
            break
    if not monday_people:
        return schedule
    adjusted = []
    for day in schedule:
        if day['day'] == '周一':
            adjusted.append(day)
        else:
            filtered = [p for p in day['people'] if p not in monday_people]
            if filtered:
                adjusted.append(dict(day=day['day'], people=filtered))
    return adjusted

def uniform_sort_key(person):
    suit = person.get('uniform', '')
    if not suit: return (2, 999, 999, 999)
    gender_code = 0 if suit.startswith('F') else 1
    nums = [int(n) for n in re.findall(r'\d+', suit)]
    while len(nums) < 3: nums.append(0)
    return (gender_code, nums[0], nums[1], nums[2])

def sort_people_by_uniform(people):
    return sorted(people, key=uniform_sort_key)

def detect_faculty_conflicts(persons):
    conflicts = []
    for eq_type in ['uniform', 'hat', 'boots', 'belt']:
        eq_map = {}
        for p in persons:
            code = p.get(eq_type, '')
            if not code: continue
            eq_map.setdefault(code, []).append(p['name'])
        for code, names in eq_map.items():
            if len(names) >= 2:
                for i in range(1, len(names)):
                    conflicts.append(dict(item_type=eq_type, item_code=code,
                                          person_to_move=names[i], person_to_keep=names[0]))
    return conflicts

def resolve_faculty_conflicts(conflicts, persons, pool):
    reassigns, changed = [], {}
    used = {}
    for eq_type in ['uniform', 'hat', 'boots', 'belt']:
        used[eq_type] = {p.get(eq_type, '') for p in persons if p.get(eq_type)}
    for c in conflicts:
        person = next((p for p in persons if p['name'] == c['person_to_move']), None)
        if not person: continue
        t = c['item_type']
        alt = find_alternative(person, t, used[t], pool.get(t, set()))
        if alt:
            used[t].discard(person[t]); used[t].add(alt)
            reassigns.append(dict(person=person['name'], item_type=t, old_item=person[t], new_item=alt))
            changed.setdefault(person['name'], {})[t] = alt
            person[t] = alt
    return reassigns, changed

def generate_faculty_excel(persons, changed):
    import openpyxl
    from openpyxl.styles import PatternFill
    wb = openpyxl.load_workbook(FACULTY_TEMPLATE)
    yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type='solid')
    sorted_persons = sort_people_by_uniform(persons)
    ws1 = wb['装备分配']
    for row in range(2, ws1.max_row + 1):
        for col in range(1, 7):
            ws1.cell(row=row, column=col).value = None
    for i, p in enumerate(sorted_persons):
        row = i + 2
        name_changed = changed.get(p['name'], {})
        ws1.cell(row=row, column=1).value = p['name']
        ws1.cell(row=row, column=2).value = p.get('uniform', '')
        ws1.cell(row=row, column=3).value = p.get('hat', '')
        ws1.cell(row=row, column=4).value = p.get('boots', '')
        ws1.cell(row=row, column=5).value = p.get('belt', '')
        if name_changed:
            notes = []
            type_names = {'uniform': '礼服', 'hat': '礼帽', 'boots': '马靴', 'belt': '腰带'}
            for t, new in name_changed.items():
                notes.append(f'{type_names.get(t, t)}→{new}')
            ws1.cell(row=row, column=6).value = '；'.join(notes)
            change_map = {'uniform': 2, 'hat': 3, 'boots': 4, 'belt': 5}
            for eq_type, col in change_map.items():
                if eq_type in name_changed:
                    ws1.cell(row=row, column=col).fill = yellow_fill
    for row in range(len(sorted_persons) + 2, ws1.max_row + 1):
        for col in range(1, 7):
            ws1.cell(row=row, column=col).value = None
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return base64.b64encode(buf.read()).decode()

# ═══ OCR ═══

def ocr_daily_schedule(img_bytes, excel_bytes):
    from PIL import Image
    import numpy as np
    with tempfile.NamedTemporaryFile(suffix='.jpg', delete=False) as t:
        t.write(img_bytes); ip = t.name
    try:
        img = Image.open(ip).convert('RGB')
        w, h = img.size
        import openpyxl
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as et:
            et.write(excel_bytes); etmp = et.name
        try:
            wb_tmp = openpyxl.load_workbook(etmp)
            known_names = []
            for row in range(2, wb_tmp['装备分配'].max_row + 1):
                n = wb_tmp['装备分配'].cell(row=row, column=2).value
                if n and str(n).strip(): known_names.append(str(n).strip())
            wb_tmp.close()
        finally:
            os.unlink(etmp)
        scale = max(1, 800 // w)
        img = img.resize((w * scale, h * scale), Image.LANCZOS)
        pre = ip + '_pre.png'; img.save(pre)
        reader = get_ocr_reader()
        results = reader.readtext(np.array(img), detail=1)
        days_info = []
        name_entries = []
        for bbox, text, conf in results:
            x1, y1 = bbox[0]; x3, y3 = bbox[2]
            cx = (x1 + x3) / 2
            t = text.strip()
            if any(d in t for d in ['周一','周二','周三','周四','周五']):
                days_info.append({'day': t, 'x1': x1, 'x2': x3, 'cx': cx})
            elif len(t) >= 2:
                name_entries.append({'name': t, 'x': cx, 'y': (y1+y3)/2, 'conf': conf})
        days_info.sort(key=lambda d: d['x1'])
        grouped = {d['day']: [] for d in days_info}
        if days_info:
            col_width = (days_info[-1]['x2'] - days_info[0]['x1']) / len(days_info)
            for n in name_entries:
                best_d, best_dist = None, 99999
                for d in days_info:
                    dist = abs(n['x'] - d['cx'])
                    if dist < best_dist: best_dist, best_d = dist, d
                if best_d and best_dist < col_width * 1.3:
                    grouped[best_d['day']].append((n['y'], n['name']))
        ocr_schedule = []
        for day_label, nl in grouped.items():
            seen = set(); unique = []
            for _, raw_name in sorted(nl):
                corrected = fuzzy_match(raw_name, known_names)
                if corrected not in seen and corrected in known_names:
                    seen.add(corrected); unique.append(corrected)
            if unique: ocr_schedule.append(dict(day=day_label, people=unique))
        os.unlink(pre)
        if ocr_schedule:
            return '\n'.join(d['day'] + '：' + '、'.join(d['people']) for d in ocr_schedule)
        return ''
    finally:
        try: os.unlink(ip)
        except: pass

def ocr_faculty_roster(img_bytes, excel_bytes):
    from PIL import Image
    import numpy as np
    with tempfile.NamedTemporaryFile(suffix='.jpg', delete=False) as t:
        t.write(img_bytes); ip = t.name
    try:
        img = Image.open(ip).convert('RGB')
        w, h = img.size
        import openpyxl
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as et:
            et.write(excel_bytes); etmp = et.name
        try:
            wb_tmp = openpyxl.load_workbook(etmp)
            known_names = []
            for row in range(2, wb_tmp['装备分配'].max_row + 1):
                n = wb_tmp['装备分配'].cell(row=row, column=2).value
                if n and str(n).strip(): known_names.append(str(n).strip())
            wb_tmp.close()
        finally:
            os.unlink(etmp)
        scale = max(1, 1200 // w)
        img = img.resize((w * scale, h * scale), Image.LANCZOS)
        pre = ip + '_pre.png'; img.save(pre)
        reader = get_ocr_reader()
        results = reader.readtext(np.array(img), detail=1)
        HEADER_KW = ['擎护旗', '队列', '总负责', '场控', '后勤', '摄影']
        header_x_positions = []
        for bbox, text, conf in results:
            x1, y1 = bbox[0]; x3, y3 = bbox[2]
            t = text.strip()
            if any(kw in t for kw in HEADER_KW):
                header_x_positions.append({'text': t, 'x1': x1, 'x2': x3, 'cx': (x1 + x3) / 2})
        queue_raw = []
        scaled_w = w * scale
        if header_x_positions:
            header_x_positions.sort(key=lambda h: h['x1'])
            queue_headers = [h for h in header_x_positions if any(kw in h['text'] for kw in ['擎护旗', '队列'])]
            other_headers = [h for h in header_x_positions if not any(kw in h['text'] for kw in ['擎护旗', '队列'])]
            if queue_headers and other_headers:
                divide_x = other_headers[0]['x1'] - 100
            elif queue_headers:
                divide_x = queue_headers[-1]['x2'] + 250
            else:
                divide_x = other_headers[0]['x1'] - 250
            for bbox, text, conf in sorted(results, key=lambda r: r[0][0][1]):
                x1, y1 = bbox[0]; x3, y3 = bbox[2]
                t = text.strip()
                if len(t) < 2: continue
                if any(kw in t for kw in HEADER_KW): continue
                if x1 < 5 or y1 < 5: continue
                cx = (x1 + x3) / 2
                if cx < divide_x:
                    corrected = fuzzy_match(t, known_names)
                    if corrected not in queue_raw:
                        queue_raw.append(corrected)
        else:
            divide_x = int(0.6 * scaled_w)
            for bbox, text, conf in sorted(results, key=lambda r: r[0][0][1]):
                x1, y1 = bbox[0]; x3, y3 = bbox[2]
                t = text.strip()
                if len(t) < 2: continue
                if any(kw in t for kw in HEADER_KW): continue
                if x1 < 5 or y1 < 5: continue
                cx = (x1 + x3) / 2
                if cx < divide_x:
                    corrected = fuzzy_match(t, known_names)
                    if corrected not in queue_raw:
                        queue_raw.append(corrected)
        os.unlink(pre)
        return '\n'.join(queue_raw) if queue_raw else ''
    finally:
        try: os.unlink(ip)
        except: pass

# ═══ CSS ═══

def inject_css():
    st.markdown("""
    <style>
    /* ═══ Reset & Global ═══ */
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+SC:wght@400;500;600;700&display=swap');
    #MainMenu, footer, header, [data-testid="stToolbar"] {visibility: hidden !important; display: none !important}
    .stApp {background-color: #ECEEEF}
    .block-container {max-width: 1200px !important; padding: 0 !important}
    body, p, div, span, label, input, textarea, select, button {
        font-family: "Noto Sans SC", "PingFang SC", "Microsoft YaHei", "STHeiti", sans-serif !important
    }

    /* ═══ Top Header Bar ═══ */
    .header-bar {
        background: #14282B; padding: 22px 36px; margin: 0;
        display: flex; align-items: center; justify-content: space-between
    }
    .header-bar-left h1 {color: #FFFFFF !important; font-size: 1.35rem !important; font-weight: 700 !important; margin: 0 !important; letter-spacing: 3px}
    .header-bar-left .sub {color: #7A8F91; font-size: 0.8rem; margin-top: 4px; letter-spacing: 1px}
    .header-bar-right {color: #5A6F71; font-size: 0.75rem; text-align: right; letter-spacing: 1px}
    .header-red-line {height: 3px; background: #C41E3A; margin: 0}

    /* ═══ Page Header (sub pages) ═══ */
    .page-header {background: #FFFFFF; padding: 16px 36px; border-bottom: 1px solid #D0D3D4; margin-bottom: 24px}
    .page-header-title {font-size: 1.2rem; font-weight: 700; color: #14282B; letter-spacing: 1px}
    .page-header-desc {font-size: 0.85rem; color: #777; margin-top: 2px}
    .page-header .back-btn {margin-bottom: 8px}

    /* ═══ Headings ═══ */
    h1 {font-size: 1.8rem !important; font-weight: 700 !important; color: #14282B !important; letter-spacing: 1px}
    h2 {font-size: 1.3rem !important; font-weight: 600 !important; color: #1F3A3D !important}
    h3 {font-size: 1.1rem !important; font-weight: 600 !important; color: #2C4A4D !important; margin-top: 1rem !important}
    h4, h5 {color: #444444 !important; font-weight: 600 !important}
    p, .stMarkdown, .stCaption {color: #555555 !important; font-size: 14px !important}
    .section-label {font-size: 0.9rem !important; font-weight: 600 !important; color: #2C4A4D !important; margin: 16px 0 6px 0 !important; letter-spacing: 0.5px}

    /* ═══ Content Area ═══ */
    .content-area {padding: 24px 36px}

    /* ═══ Mission Cards (Home) ═══ */
    .mission-card {
        background: #FAFBFC; border: 1px solid #D0D3D4; border-left: 4px solid #1F3A3D;
        padding: 32px 28px; border-radius: 2px; min-height: 220px; cursor: default;
        transition: border-color 0.15s, box-shadow 0.15s
    }
    .mission-card:hover {border-color: #1F3A3D; box-shadow: 0 2px 10px rgba(20,40,43,0.10)}
    .mission-card-icon {font-size: 2rem; margin-bottom: 14px}
    .mission-card-title {font-size: 1.1rem; font-weight: 700; color: #14282B; margin-bottom: 8px; letter-spacing: 0.5px}
    .mission-card-desc {font-size: 0.88rem; color: #777; line-height: 1.5}

    /* ═══ Footer ═══ */
    .footer {text-align: center; padding: 28px 0 16px 0; color: #999; font-size: 0.75rem; letter-spacing: 1px}

    /* ═══ Buttons ═══ */
    .stButton button {
        font-weight: 600 !important; border-radius: 2px !important;
        padding: 0.6rem 1.5rem !important; font-size: 14px !important;
        background-color: #1F3A3D !important; color: #FFFFFF !important;
        border: none !important; letter-spacing: 0.5px;
        transition: background-color 0.15s !important
    }
    .stButton button:hover {background-color: #2C4A4D !important; transform: none !important}
    .stButton button:disabled {background-color: #C0C4C5 !important; color: #888 !important}
    .stDownloadButton button {
        font-weight: 600 !important; border-radius: 2px !important;
        background-color: #C41E3A !important; color: #FFFFFF !important;
        width: 100% !important; padding: 0.75rem !important; font-size: 15px !important;
        border: none !important; letter-spacing: 1px;
        transition: background-color 0.15s !important
    }
    .stDownloadButton button:hover {background-color: #A01830 !important; transform: none !important}

    /* Back button variant — secondary style */
    .back-btn-row {margin-bottom: 0}

    /* ═══ Cards & Containers ═══ */
    div[data-testid="stVerticalBlockBorderWrapper"] {
        background: #FAFBFC !important; border: 1px solid #D0D3D4 !important;
        border-radius: 2px !important; border-left: 3px solid #1F3A3D !important
    }
    .stExpander {
        border: 1px solid #D0D3D4 !important; border-radius: 2px !important;
        background: #FAFBFC !important
    }
    .stExpander summary {color: #1F3A3D !important; font-weight: 600 !important}

    /* ═══ Tables ═══ */
    [data-testid="stTable"] table {border-collapse: collapse !important; width: 100% !important}
    [data-testid="stTable"] thead th {
        background-color: #1F3A3D !important; color: #FFFFFF !important;
        font-weight: 600 !important; padding: 10px 12px !important;
        border: none !important; font-size: 13px !important; letter-spacing: 0.5px
    }
    [data-testid="stTable"] tbody td {
        padding: 8px 12px !important; border-bottom: 1px solid #E0E3E4 !important;
        color: #333333 !important; font-size: 13px !important
    }
    [data-testid="stTable"] tbody tr:nth-child(even) {background-color: #F5F6F7 !important}
    [data-testid="stTable"] tbody tr:nth-child(odd) {background-color: #FFFFFF !important}

    /* Streamlit dataframe override */
    div[data-testid="stDataFrame"] table {border-collapse: collapse !important}
    div[data-testid="stDataFrame"] th {
        background-color: #1F3A3D !important; color: #FFFFFF !important;
        font-weight: 600 !important; padding: 8px 12px !important;
        border: none !important; font-size: 13px !important
    }
    div[data-testid="stDataFrame"] td {
        padding: 6px 12px !important; border-bottom: 1px solid #E0E3E4 !important;
        color: #333 !important; font-size: 13px !important
    }

    /* ═══ File Uploader ═══ */
    [data-testid="stFileUploader"] {
        border: 2px solid #D0D3D4 !important; border-radius: 2px !important;
        background: #FAFBFC !important; padding: 4px !important
    }
    [data-testid="stFileUploader"]:hover {border-color: #1F3A3D !important}
    [data-testid="stFileUploader"] button {
        background-color: #1F3A3D !important; color: #FFF !important;
        border-radius: 2px !important; font-size: 13px !important
    }

    /* ═══ Inputs & Textareas ═══ */
    .stTextArea textarea, .stTextInput input, .stSelectbox select {
        border: 1px solid #C0C4C5 !important; border-radius: 2px !important;
        color: #111111 !important; font-size: 14px !important;
        background: #FAFBFC !important
    }
    .stTextArea textarea:focus, .stTextInput input:focus {
        border-color: #1F3A3D !important; box-shadow: 0 0 0 2px rgba(31,58,61,0.10) !important
    }

    /* ═══ Metrics ═══ */
    div[data-testid="stMetricValue"] {
        font-size: 1.8rem !important; font-weight: 700 !important; color: #1F3A3D !important
    }
    div[data-testid="stMetricLabel"] {color: #777 !important; font-size: 0.8rem !important}
    div[data-testid="stMetric"] {
        background: #FAFBFC !important; border: 1px solid #E0E3E4 !important;
        border-radius: 2px !important; padding: 12px 16px !important
    }

    /* ═══ Dividers ═══ */
    hr {border: none !important; border-top: 1px solid #D0D3D4 !important; margin: 20px 0 !important}
    .section-divider {border-top: 2px solid #E0E3E4; margin: 24px 0}

    /* ═══ Alerts ═══ */
    div[data-testid="stAlert"] {
        border-radius: 2px !important; border-left: 3px solid #1F3A3D !important
    }
    div[data-testid="stAlert"][data-baseweb="notification"] {
        background: #FAFBFC !important
    }

    /* ═══ Spinner ═══ */
    .stSpinner > div {border-top-color: #1F3A3D !important}

    /* ═══ Sidebar (hidden) ═══ */
    section[data-testid="stSidebar"] {display: none !important}
    </style>
    """, unsafe_allow_html=True)

# ═══ 页面路由 ═══

def go_to(page):
    st.session_state.page = page
    st.rerun()

# ── 首页 ──
def page_home():
    inject_css()
    # 顶部标题栏
    st.markdown("""
    <div class="header-bar">
        <div class="header-bar-left">
            <h1>礼服自动分配系统</h1>
            <div class="sub">UNIFORM ASSIGNMENT SYSTEM</div>
        </div>
        <div class="header-bar-right">浙江大学国旗仪仗队<br>装备管理控制台</div>
    </div>
    <div class="header-red-line"></div>
    <div class="content-area">
    <p style="font-size:0.95rem;color:#444;margin-bottom:28px;">选择分配模式开始使用</p>
    """, unsafe_allow_html=True)
    c1, c2 = st.columns(2, gap="medium")
    with c1:
        st.markdown("""
        <div class="mission-card">
            <div class="mission-card-icon">🌄</div>
            <div class="mission-card-title">日常升旗班礼服分配</div>
            <div class="mission-card-desc">上传上月 Excel + 拍照上传排班表<br>自动检测共享冲突 → 重分配 → 下载</div>
        </div>
        """, unsafe_allow_html=True)
        if st.button("进入日常分配", key="go_daily", use_container_width=True): go_to('daily')
    with c2:
        st.markdown("""
        <div class="mission-card">
            <div class="mission-card-icon">🎖️</div>
            <div class="mission-card-title">院系升旗礼服分配</div>
            <div class="mission-card-desc">上传库存表 + 拍照上传人员安排<br>按尺寸排序 → 冲突检测 → 下载 Excel</div>
        </div>
        """, unsafe_allow_html=True)
        if st.button("进入院系分配", key="go_faculty", use_container_width=True): go_to('faculty')
    st.markdown('<div class="footer">浙江大学国旗仪仗队 &copy; 2026</div></div>', unsafe_allow_html=True)

# ── 日常升旗页 ──
def page_daily():
    inject_css()
    st.markdown("""
    <div class="header-bar">
        <div class="header-bar-left">
            <h1>日常升旗班礼服分配</h1>
            <div class="sub">DAILY FLAG-RAISING · EQUIPMENT ASSIGNMENT</div>
        </div>
        <div class="header-bar-right">浙江大学国旗仪仗队</div>
    </div>
    <div class="header-red-line"></div>
    <div class="content-area">
    """, unsafe_allow_html=True)
    if st.button("← 返回首页", key="daily_back"): go_to('home')
    st.caption("上传上月分配表 + 拍照上传排班表 → 自动检测冲突 → 下载新表")
    st.markdown('<p class="section-label">① 上传上月分配表 (.xlsx)</p>', unsafe_allow_html=True)
    excel_file = st.file_uploader("拖拽或点击上传上月 Excel", type=['xlsx'], key="daily_excel", label_visibility="collapsed")
    st.markdown('<p class="section-label">② 拍照上传排班表</p>', unsafe_allow_html=True)
    image_file = st.file_uploader("上传排班表照片（手写也可以）", type=['png','jpg','jpeg','webp'], key="daily_image", label_visibility="collapsed")
    st.caption("或手动输入/修正排班文字")
    schedule_text = st.text_area("排班文字", placeholder="周一：李诗诗、郭婷心、岳佳凝、江文欣\n周二：柯天翊、章芮容、余佳卉、马欣雅\n周三：李泽一、方佳瑶、纪博雅、董欢瑶\n周四：林珩、王雨梦、施东隅、艾克达\n周五：吴桐、段茗萱、张艺、许诺", key="daily_schedule_text", label_visibility="collapsed", height=140)
    ocr_warning = st.empty()
    can_generate = excel_file and (schedule_text.strip() or image_file)
    if st.button("🚀 生成本月分配表", key="daily_gen", use_container_width=True, disabled=not can_generate):
        if not excel_file: return
        with st.spinner("正在处理..."):
            try:
                excel_bytes = excel_file.read()
                ocr_parsed = ''
                if image_file:
                    img_bytes = image_file.read()
                    if len(img_bytes) > 0:
                        ocr_parsed = ocr_daily_schedule(img_bytes, excel_bytes)
                        if ocr_parsed:
                            schedule_text = ocr_parsed
                            ocr_warning.info("⚠️ OCR 识别结果已填入下方，请核对修正后再点生成")
                if not schedule_text.strip(): st.error("请提供排班文本或上传排班照片"); st.stop()
                schedule = parse_schedule(schedule_text.strip())
                schedule = adjust_monday_schedule(schedule)
                if not schedule: st.error(f"排班格式无法解析。\n{schedule_text[:300]}"); st.stop()
                import openpyxl
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as t:
                    t.write(excel_bytes); tmp = t.name
                try:
                    wb = openpyxl.load_workbook(tmp)
                    persons = parse_equipment_sheet(wb['装备分配'])
                    rels = parse_shared_relations(persons)
                    mrels = merge_relations(rels)
                    conflicts = detect_conflicts(schedule, mrels)
                    pool = build_pool(wb)
                    reassigns, changed = resolve(conflicts, persons, schedule, pool)
                    b64 = generate_excel(tmp, persons, changed)
                    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
                    st.markdown('<p class="section-label">③ 生成结果</p>', unsafe_allow_html=True)
                    x1, x2, x3 = st.columns(3)
                    x1.metric("冲突数", len(conflicts)); x2.metric("重分配", len(reassigns)); x3.metric("受影响人数", len(set(r['person'] for r in reassigns)))
                    if conflicts:
                        st.markdown('<p class="section-label">冲突详情</p>', unsafe_allow_html=True)
                        st.dataframe([{"日期":c['day'],"装备":'礼服' if c['item_type']=='uniform' else '礼帽',"编号":c['item_code'],"需变动":c['person_to_move'],"保留":c['person_to_keep']} for c in conflicts], use_container_width=True, hide_index=True)
                    if reassigns:
                        st.markdown('<p class="section-label">重分配方案</p>', unsafe_allow_html=True)
                        st.dataframe([{"姓名":r['person'],"装备":'礼服' if r['item_type']=='uniform' else '礼帽',"旧编号":r['old_item'],"新编号":r['new_item']} for r in reassigns], use_container_width=True, hide_index=True)
                    st.download_button("📥 下载本月分配表 (.xlsx)", data=base64.b64decode(b64), file_name="本月礼服分配.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    if ocr_parsed: st.text_area("OCR 识别结果（可复制到上方修正）", value=ocr_parsed, height=120, key="daily_ocr_result")
                finally:
                    os.unlink(tmp)
            except Exception as e:
                import traceback; st.error(f"生成失败: {e}")
                with st.expander("详细错误"): st.code(traceback.format_exc())

# ── 院系升旗页 ──
def page_faculty():
    inject_css()
    st.markdown("""
    <div class="header-bar">
        <div class="header-bar-left">
            <h1>院系升旗礼服分配</h1>
            <div class="sub">FACULTY FLAG-RAISING · EQUIPMENT ASSIGNMENT</div>
        </div>
        <div class="header-bar-right">浙江大学国旗仪仗队</div>
    </div>
    <div class="header-red-line"></div>
    <div class="content-area">
    """, unsafe_allow_html=True)
    if st.button("← 返回首页", key="faculty_back"): go_to('home')
    st.caption("输入队列人员 → 按尺寸排序 → 冲突检测 → 下载 4-sheet Excel")
    st.markdown('<p class="section-label">① 上传礼服库存表 (.xlsx)</p>', unsafe_allow_html=True)
    excel_file = st.file_uploader("上传包含所有人员装备信息的 Excel", type=['xlsx'], key="faculty_excel", label_visibility="collapsed")
    st.markdown('<p class="section-label">② 拍照上传人员安排表</p>', unsafe_allow_html=True)
    image_file = st.file_uploader("上传人员安排表照片（手写也可以）", type=['png','jpg','jpeg','webp'], key="faculty_image", label_visibility="collapsed")
    st.caption("AI 自动识别队列人员（总负责/场控/后勤/摄影不参与分配）")
    st.caption("或手动输入/修正队列人员（每行一人或顿号分隔）")
    roster_text = st.text_area("队列人员", placeholder="林珩\n韩雅丽\n艾克达\n张鹏\n夏瑞泽\n戴傲\n叶宇轩", key="faculty_roster", label_visibility="collapsed", height=130)
    ocr_warning2 = st.empty()
    with st.expander("其他角色（可选，不参与分配）"):
        c1, c2 = st.columns(2)
        meta1 = c1.text_area("总负责", key="f_meta1", height=52, placeholder="（可选）")
        meta2 = c2.text_area("场控", key="f_meta2", height=52, placeholder="（可选）")
        c3, c4 = st.columns(2)
        meta3 = c3.text_area("后勤", key="f_meta3", height=52, placeholder="（可选）")
        meta4 = c4.text_area("摄影", key="f_meta4", height=52, placeholder="（可选）")
    can_gen = excel_file and (roster_text.strip() or image_file)
    if st.button("🔍 预览排序 & 生成分配表", key="faculty_gen", use_container_width=True, disabled=not can_gen):
        if not excel_file: return
        with st.spinner("正在处理..."):
            try:
                excel_bytes = excel_file.read()
                ocr_parsed = ''
                if image_file:
                    img_bytes = image_file.read()
                    if len(img_bytes) > 0:
                        ocr_parsed = ocr_faculty_roster(img_bytes, excel_bytes)
                        if ocr_parsed:
                            roster_text = ocr_parsed
                            ocr_warning2.info("⚠️ OCR 识别结果已填入下方，请核对修正后再点生成")
                if not roster_text.strip(): st.error("请提供队列人员名单或上传照片"); st.stop()
                queue_names = []
                for chunk in re.split(r'[、，,\n\s]+', roster_text.strip()):
                    n = chunk.strip()
                    if len(n) >= 2: queue_names.append(n)
                import openpyxl
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as t:
                    t.write(excel_bytes); tmp = t.name
                try:
                    wb = openpyxl.load_workbook(tmp)
                    all_persons = parse_equipment_sheet(wb['装备分配'])
                    name_map = {p['name']: p for p in all_persons}
                    pool = build_pool(wb)
                    faculty_persons = []; missing = []
                    for name in queue_names:
                        p = name_map.get(name)
                        if p: faculty_persons.append(dict(p))
                        else: missing.append(name)
                    original_equip = {}
                    for p in faculty_persons:
                        for k in ['uniform','hat','boots','belt']:
                            original_equip[f'{p["name"]}|{k}'] = p.get(k,'')
                    conflicts = detect_faculty_conflicts(faculty_persons)
                    reassigns, changed = resolve_faculty_conflicts(conflicts, faculty_persons, pool)
                    b64 = generate_faculty_excel(faculty_persons, changed)
                    sorted_persons = sort_people_by_uniform(faculty_persons)
                    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
                    st.markdown('<p class="section-label">③ 生成结果</p>', unsafe_allow_html=True)
                    msg = f"生成完成: {len(faculty_persons)} 人, {len(conflicts)} 个冲突, {len(reassigns)} 个重分配"
                    if missing: msg += f"（{len(missing)} 人未在库存中找到）"
                    st.success(msg)
                    x1, x2, x3 = st.columns(3)
                    x1.metric("冲突数", len(conflicts)); x2.metric("重分配", len(reassigns)); x3.metric("受影响人数", len(set(r['person'] for r in reassigns)))
                    st.markdown('<p class="section-label">排序后人员列表</p>', unsafe_allow_html=True)
                    rows = []
                    for i, p in enumerate(sorted_persons):
                        ch = changed.get(p['name'], {})
                        rows.append({"序号":i+1,"姓名":p['name'],"性别":'女' if p['gender']=='F' else '男',"礼服":p.get('uniform',''),"礼帽":p.get('hat',''),"马靴":p.get('boots',''),"腰带":p.get('belt',''),"变更":'；'.join(f'{t}→{nv}' for t,nv in ch.items()) if ch else ''})
                    st.dataframe(rows, use_container_width=True, hide_index=True)
                    if conflicts:
                        st.markdown('<p class="section-label">冲突详情</p>', unsafe_allow_html=True)
                        st.dataframe([{"装备类型":{'uniform':'礼服','hat':'礼帽','boots':'马靴','belt':'腰带'}.get(c['item_type'],c['item_type']),"编号":c['item_code'],"需变动":c['person_to_move'],"保留":c['person_to_keep']} for c in conflicts], use_container_width=True, hide_index=True)
                    if reassigns:
                        st.markdown('<p class="section-label">重分配方案</p>', unsafe_allow_html=True)
                        st.dataframe([{"姓名":r['person'],"装备":{'uniform':'礼服','hat':'礼帽','boots':'马靴','belt':'腰带'}.get(r['item_type'],r['item_type']),"旧编号":r['old_item'],"新编号":r['new_item']} for r in reassigns], use_container_width=True, hide_index=True)
                    st.download_button("📥 下载礼服分配表 (.xlsx)", data=base64.b64decode(b64), file_name="院系升旗礼服分配.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    if ocr_parsed: st.text_area("OCR 识别结果（可复制到上方修正）", value=ocr_parsed, height=120, key="faculty_ocr_result")
                finally:
                    os.unlink(tmp)
            except Exception as e:
                import traceback; st.error(f"生成失败: {e}")
                with st.expander("详细错误"): st.code(traceback.format_exc())

# ═══ Router ═══
if st.session_state.page == 'home':
    page_home()
elif st.session_state.page == 'daily':
    page_daily()
elif st.session_state.page == 'faculty':
    page_faculty()
