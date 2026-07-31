import openpyxl
from openpyxl.utils import get_column_letter
import json

EXCEL_PATH = r"c:\Users\夏瑞泽\xwechat_files\wxid_vnl456wm0r9h12_0c22\msg\file\2026-07\2025-2026学年夏五周到夏八周礼服分配(1)(1).xlsx"

wb = openpyxl.load_workbook(EXCEL_PATH)

# ==========================================
# Sheet 1: 装备分配 — all data rows
# ==========================================
ws = wb['装备分配']
print("=" * 80)
print("=== SHEET 1: 装备分配 (ALL DATA) ===")
print("=" * 80)

for row in range(1, ws.max_row + 1):
    values = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        v = cell.value
        if v is not None:
            # Check for yellow fill (FFE26B6B is light red/auburn in Excel indexed, FFFFFF00 etc)
            fg = cell.fill.fgColor
            bg = cell.fill.bgColor
            has_yellow = False
            try:
                if fg and fg.rgb and 'FFFF00' in str(fg.rgb).upper():
                    has_yellow = True
            except:
                pass
            try:
                if bg and bg.rgb and 'FFFF00' in str(bg.rgb).upper():
                    has_yellow = True
                # Check indexed yellow
                if bg and bg.indexed == 5:
                    has_yellow = True
            except:
                pass
            marker = " 🟡" if has_yellow else ""
            values.append(f"{get_column_letter(col)}={repr(v)}{marker}")
    if values:
        print(f"  Row {row}: {' | '.join(values)}")

# ==========================================
# Sheet 2: 礼服腰带摆放
# ==========================================
print("\n" + "=" * 80)
print("=== SHEET 2: 礼服腰带摆放 (ALL DATA) ===")
print("=" * 80)
ws2 = wb['礼服腰带摆放']
for row in range(1, ws2.max_row + 1):
    values = []
    for col in range(1, ws2.max_column + 1):
        cell = ws2.cell(row=row, column=col)
        v = cell.value
        if v is not None:
            # Check red font
            is_red = False
            try:
                if cell.font.color and cell.font.color.rgb and 'FF0000' in str(cell.font.color.rgb):
                    is_red = True
            except:
                pass
            marker = " 🔴" if is_red else ""
            values.append(f"{get_column_letter(col)}={repr(v)}{marker}")
    if values:
        print(f"  Row {row}: {' | '.join(values)}")

# ==========================================
# Sheet 3: 礼帽摆放
# ==========================================
print("\n" + "=" * 80)
print("=== SHEET 3: 礼帽摆放 (ALL DATA) ===")
print("=" * 80)
ws3 = wb['礼帽摆放']
for row in range(1, ws3.max_row + 1):
    values = []
    for col in range(1, ws3.max_column + 1):
        cell = ws3.cell(row=row, column=col)
        v = cell.value
        if v is not None:
            is_red = False
            try:
                if cell.font.color and cell.font.color.rgb and 'FF0000' in str(cell.font.color.rgb):
                    is_red = True
            except:
                pass
            marker = " 🔴" if is_red else ""
            values.append(f"{get_column_letter(col)}={repr(v)}{marker}")
    if values:
        print(f"  Row {row}: {' | '.join(values)}")

# ==========================================
# Sheet 4: 马靴摆放
# ==========================================
print("\n" + "=" * 80)
print("=== SHEET 4: 马靴摆放 (ALL DATA) ===")
print("=" * 80)
ws4 = wb['马靴摆放']
for row in range(1, ws4.max_row + 1):
    values = []
    for col in range(1, ws4.max_column + 1):
        cell = ws4.cell(row=row, column=col)
        v = cell.value
        if v is not None:
            values.append(f"{get_column_letter(col)}={repr(v)}")
    if values:
        print(f"  Row {row}: {' | '.join(values)}")

print("\nDone.")
