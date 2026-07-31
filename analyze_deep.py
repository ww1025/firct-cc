import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json

EXCEL_PATH = r"c:\Users\夏瑞泽\xwechat_files\wxid_vnl456wm0r9h12_0c22\msg\file\2026-07\2025-2026学年夏五周到夏八周礼服分配(1)(1).xlsx"

wb = openpyxl.load_workbook(EXCEL_PATH)

# ============================================================
# Sheet 1: 装备分配 - deep analysis
# ============================================================
ws = wb['装备分配']
print("=" * 80)
print("SHEET: 装备分配")
print("=" * 80)

# Print every cell with content, style info
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            f = cell.font
            fill = cell.fill
            align = cell.alignment
            border = cell.border
            print(f"  [{row},{col}] value={repr(cell.value)}")
            print(f"    font: name={f.name}, size={f.size}, bold={f.bold}, italic={f.italic}, color={f.color}, vertAlign={f.vertAlign}")
            print(f"    fill: fgColor={fill.fgColor}, bgColor={fill.bgColor}, patternType={fill.patternType}")
            print(f"    align: h={align.horizontal}, v={align.vertical}, wrap={align.wrap_text}, shrink={align.shrink_to_fit}")
            if border:
                for side_name in ['left', 'right', 'top', 'bottom']:
                    s = getattr(border, side_name)
                    if s and s.style:
                        print(f"    border {side_name}: style={s.style}, color={s.color}")
            print()

# ============================================================
# Column widths and row heights
# ============================================================
print("Column widths:")
for col_letter in ['A','B','C','D','E','F','G','H']:
    if col_letter in ws.column_dimensions:
        print(f"  {col_letter}: width={ws.column_dimensions[col_letter].width}")

print("\nRow heights (first 65):")
for r in range(1, 66):
    if r in ws.row_dimensions:
        rd = ws.row_dimensions[r]
        print(f"  row {r}: height={rd.height}")
